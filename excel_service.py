"""
excel_service.py  —  Input parsing and formatted output for the verification flow
=================================================================================

Owns the workbook schema for the verification pipeline:

    Talent Name | Wikipedia URL |
    <Platform> | <Platform> Status | <Platform> Confidence | <Platform> Reason  (x5) |
    Confidence (overall)

Platforms follow ``social_urls.PLATFORMS`` order:
    Instagram, Facebook, YouTube, TikTok, X

Public interface:
    load_talent_table_from_path(path) -> DataFrame  (Talent Name + Wikipedia URL)
    build_talent_df(names)            -> DataFrame  (names-only entry point)
    empty_result_columns()            -> list[str]  (ordered output columns)
    save_results(df, output_dir)      -> Path       (formatted .xlsx)
    save_output(df, output_dir)       -> str        (api-compatible wrapper)
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

import social_urls
from verification_service import (
    STATUS_MANUAL,
    STATUS_NOT_FOUND,
    STATUS_STOPPED,
    STATUS_VERIFIED,
    STATUS_WRONG,
)

TALENT_COL = "Talent Name"
WIKI_COL = "Wikipedia URL"
OVERALL_CONF_COL = "Confidence"

# Working-only column holding a dict of the remaining input columns for a row.
# Used as identity context by the verifier (especially when no Wikipedia URL is
# given). It is NOT part of ``ordered_columns`` so it is dropped before export.
INPUT_META_COL = "_input_metadata"

# Working-only column holding {platform: profile_url} for handles the CLIENT
# already recorded in the input file. These are by far the strongest candidates
# available — the brand-definition export ships instagram_user / twitter_handle
# / facebook_page / tiktok_user / youtube_channel_username on many rows, and
# re-discovering them from scratch is both wasteful and less accurate.
# Not part of ``ordered_columns`` so it is dropped before export.
INPUT_HANDLES_COL = "_input_handles"

# Input column aliases that carry an existing handle/URL per platform. Matched
# case-insensitively against the header. Order matters: first hit wins.
_PLATFORM_HANDLE_COLUMNS: Dict[str, List[str]] = {
    "Instagram": ["instagram_user", "instagram_username", "instagram_handle",
                  "instagram_url", "instagram", "ig_handle"],
    "Facebook":  ["facebook_page", "facebook_url", "facebook_username",
                  "facebook", "fb_page"],
    "X":         ["twitter_handle", "twitter_url", "twitter_username",
                  "x_handle", "x_url", "twitter"],
    "YouTube":   ["youtube_channel_username", "youtube_channel_id",
                  "youtube_channel", "youtube_url", "youtube"],
    "TikTok":    ["tiktok_user", "tiktok_handle", "tiktok_url", "tiktok"],
}

# Output platform order (also the processing/UI order).
PLATFORM_ORDER: List[str] = list(social_urls.PLATFORMS.keys())


def link_col(platform: str) -> str:
    return platform


def status_col(platform: str) -> str:
    return f"{platform} Status"


def conf_col(platform: str) -> str:
    return f"{platform} Confidence"


def reason_col(platform: str) -> str:
    return f"{platform} Reason"


def source_col(platform: str) -> str:
    """Where the chosen link came from — SerpApi, Phase 0 bio, input file, etc."""
    return f"{platform} Source"


def ordered_columns() -> List[str]:
    cols: List[str] = [TALENT_COL, WIKI_COL]
    for platform in PLATFORM_ORDER:
        cols.extend([
            link_col(platform), status_col(platform), source_col(platform),
            conf_col(platform), reason_col(platform),
        ])
    cols.append(OVERALL_CONF_COL)
    return cols


def empty_result_columns() -> List[str]:
    return ordered_columns()


# ────────────────────────────────────────────────────────────────────────────
#  Input parsing
# ────────────────────────────────────────────────────────────────────────────

def _find_column(raw: pd.DataFrame, *candidates: str) -> Optional[str]:
    """Exact (case-insensitive) header match against known aliases."""
    cmap = {str(c).strip().lower(): c for c in raw.columns}
    for cand in candidates:
        if cand.lower() in cmap:
            return cmap[cand.lower()]
    return None


def _find_column_contains(raw: pd.DataFrame, *keywords: str,
                          exclude: Optional[set] = None) -> Optional[str]:
    """Match the first header that contains any of the given keywords."""
    exclude = exclude or set()
    for col in raw.columns:
        if col in exclude:
            continue
        header = str(col).strip().lower()
        if any(kw in header for kw in keywords):
            return col
    return None


_WIKI_URL_RE = re.compile(r"wikipedia\.org/wiki/", re.I)


def _detect_wiki_column(raw: pd.DataFrame, exclude: Optional[set] = None) -> Optional[str]:
    """
    Find the Wikipedia column by its CONTENT, not its header — so the header can
    be named anything. Picks the first column whose cells contain
    'wikipedia.org/wiki/...' links.
    """
    exclude = exclude or set()
    for col in raw.columns:
        if col in exclude:
            continue
        for value in raw[col].dropna().astype(str).head(25):
            if _WIKI_URL_RE.search(value):
                return col
    return None


def _clean_str(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


# Internal client tag appended to names (e.g. "Chris Brown - DAR"); strip it so
# it doesn't pollute Apify/Serper searches or the displayed name.
_NAME_TAG_SUFFIX_RE = re.compile(r"\s*[-–—]\s*DAR\s*$", re.I)


def _clean_talent_name(name: str) -> str:
    return _NAME_TAG_SUFFIX_RE.sub("", name).strip()


def _new_frame(rows: List[Dict[str, str]]) -> pd.DataFrame:
    """Create a DataFrame with the full output schema, seeded with input rows."""
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[TALENT_COL, WIKI_COL])
    for col in ordered_columns():
        if col not in df.columns:
            df[col] = ""
    for working_col in (INPUT_META_COL, INPUT_HANDLES_COL):
        if working_col not in df.columns:
            df[working_col] = ""
    df = df[ordered_columns() + [INPUT_META_COL, INPUT_HANDLES_COL]]
    # Force object dtype so numeric confidences (and the metadata dict) can share
    # columns with strings — some pandas versions infer a strict `str` dtype for
    # all-string columns, which then rejects int/dict assignment.
    return df.astype(object)


def load_talent_table_from_path(excel_path: Path) -> pd.DataFrame:
    """Read an .xlsx/.xls/.csv into the verification schema (Talent Name + Wikipedia URL)."""
    excel_path = Path(excel_path)
    if not excel_path.is_file():
        raise ValueError(f"File not found: {excel_path}")
    suffix = excel_path.suffix.lower()
    try:
        raw = pd.read_csv(excel_path) if suffix == ".csv" else pd.read_excel(excel_path)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Could not read spreadsheet: {exc}") from exc
    if raw.empty:
        raise ValueError("The file has no rows.")

    # Name column: known aliases → any header mentioning talent/name/title/brand
    # → fall back to the first column.
    name_col = (
        _find_column(raw, TALENT_COL, "Talent", "Name", "Title", "title")
        or _find_column_contains(raw, "talent", "name", "title", "brand", "entity")
        or raw.columns[0]
    )

    # Wikipedia column, resilient to arbitrary header names:
    #   1) known aliases, 2) any header containing "wiki",
    #   3) sniff the cell values for wikipedia.org/wiki/ links.
    wiki_col = (
        _find_column(
            raw, WIKI_COL, "wikipedia_url", "wikipedia_page", "Wikipedia Page",
            "Wikipedia", "Wiki URL", "wiki_url", "Wiki", "Wikipedia Link", "wiki_page",
        )
        or _find_column_contains(raw, "wiki", exclude={name_col})
        or _detect_wiki_column(raw, exclude={name_col})
    )

    # Columns that already carry a known profile handle/URL per platform.
    handle_cols: Dict[str, str] = {}
    for platform, aliases in _PLATFORM_HANDLE_COLUMNS.items():
        col = _find_column(raw, *aliases)
        if col is not None:
            handle_cols[platform] = col

    # Every column that is neither the name nor the Wikipedia URL becomes
    # per-row identity metadata (used by the verifier, e.g. when no wiki link).
    meta_cols = [c for c in raw.columns if c not in (name_col, wiki_col)]

    rows: List[Dict[str, object]] = []
    handle_hits = 0
    for i in range(len(raw)):
        name = _clean_talent_name(_clean_str(raw.iloc[i][name_col]))
        if not name:
            continue
        wiki = _clean_str(raw.iloc[i][wiki_col]) if wiki_col else ""
        metadata = {
            str(c): _clean_str(raw.iloc[i][c])
            for c in meta_cols
            if _clean_str(raw.iloc[i][c])
        }
        known: Dict[str, str] = {}
        for platform, col in handle_cols.items():
            # _clean_str first: a bare pandas NaN stringifies to "nan", which
            # would otherwise be accepted as the handle "nan".
            url = social_urls.coerce_profile_url(_clean_str(raw.iloc[i][col]), platform)
            if url:
                known[platform] = url
        handle_hits += len(known)
        rows.append({
            TALENT_COL: name, WIKI_COL: wiki,
            INPUT_META_COL: metadata, INPUT_HANDLES_COL: known,
        })

    if not rows:
        raise ValueError("No valid talent names found.")
    if handle_cols:
        print(f"[INPUT] Handle columns detected: "
              f"{ {p: str(c) for p, c in handle_cols.items()} }")
        print(f"[INPUT] {handle_hits} known profile URL(s) supplied across "
              f"{len(rows)} row(s) — these are verified, not re-discovered.")
    return _new_frame(rows)


def build_talent_df(names: List[str]) -> pd.DataFrame:
    """Names-only entry point (no Wikipedia URLs)."""
    rows = [{TALENT_COL: str(n).strip(), WIKI_COL: ""} for n in names if str(n).strip()]
    if not rows:
        raise ValueError("At least one non-empty name is required.")
    return _new_frame(rows)


# ────────────────────────────────────────────────────────────────────────────
#  Formatted output
# ────────────────────────────────────────────────────────────────────────────

# Status → fill colour (hex, openpyxl solid fill).
_STATUS_FILL = {
    STATUS_VERIFIED: "E2EFDA",   # green
    STATUS_MANUAL: "FFF2CC",     # amber
    STATUS_WRONG: "FCE4D6",      # red
    STATUS_NOT_FOUND: "F2F2F2",  # grey
    STATUS_STOPPED: "DDEBF7",    # blue — never searched, not an absence
}


def save_results(
    df: pd.DataFrame,
    output_dir: Optional[Path] = None,
    filename_prefix: str = "Talent_Social_Lookup",
) -> Path:
    """Write the results workbook with header styling and status colour bands."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = Path(output_dir) if output_dir is not None else Path(__file__).resolve().parent
    base_dir.mkdir(parents=True, exist_ok=True)
    output_path = base_dir / f"{filename_prefix}_{timestamp}.xlsx"

    # Keep only known columns, in canonical order.
    export_df = df.reindex(columns=ordered_columns())

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        export_df.to_excel(output_path, index=False)
        print(f"\n[EXPORT] Saved (no formatting): {output_path}")
        return output_path

    export_df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [str(cell.value or "") for cell in ws[1]]
    status_col_indices = {
        headers.index(status_col(p)): p for p in PLATFORM_ORDER if status_col(p) in headers
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for idx in status_col_indices:
            cell = row[idx]
            fill_hex = _STATUS_FILL.get(str(cell.value or "").strip())
            if fill_hex:
                cell.fill = PatternFill("solid", fgColor=fill_hex)

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_cell in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row_cell:
                try:
                    max_len = max(max_len, len(str(c.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n[EXPORT] Saved: {output_path}")
    return output_path


def save_output(
    df: pd.DataFrame,
    output_dir: Optional[Path] = None,
    filename_prefix: str = "Talent_Social_Lookup",
) -> str:
    """API-compatible wrapper used by api_server.py."""
    return str(save_results(df, output_dir=output_dir, filename_prefix=filename_prefix))
