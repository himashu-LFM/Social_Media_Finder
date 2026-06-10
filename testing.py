
# import json
# import os
# import random
# import re
# import time
# from concurrent.futures import ThreadPoolExecutor, as_completed
# from datetime import datetime
# from pathlib import Path
# from typing import Callable, Dict, List, Optional, Tuple
# from urllib.parse import unquote, urlparse

# import pandas as pd
# import requests

# try:
#     from dotenv import load_dotenv
#     load_dotenv(Path(__file__).resolve().parent / ".env")
# except ImportError:
#     pass

# # ================== API KEYS ==================
# SERPER_API_KEY = os.environ.get("SERPER_API_KEY", "").strip()
# OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
# OPENAI_CHAT_MODEL = os.environ.get("OPENAI_CHAT_MODEL", "gpt-4o-mini")

# TEST_BRANDS_PATH = Path(__file__).resolve().parent / "Demo_Social.xlsx"
# WIKIPEDIA_URL_COLUMN = "Wikipedia URL"

# # ================== INPUT FALLBACK ==================
# talent_names = [
#     "Britney Vest",
#     "Ari Melber",
#     "Alyssa Anderson",
#     "Andrea",
#     "Anastasia Pagonis",
# ]

# # ================== CONFIG ==================
# RESULTS_PER_QUERY         = 10
# MAX_CANDIDATES_FOR_AI     = 6   # slightly more for two-phase AI
# MAX_WORKERS               = 3
# REQUEST_DELAY_BETWEEN_TALENTS = (1.0, 2.0)
# OPENAI_DELAY_SECONDS      = 0.4

# # Base emit gate — may be raised dynamically for ambiguous names
# MIN_CONFIDENCE_EMIT       = float(os.environ.get("MIN_CONFIDENCE_EMIT", "0.72"))
# ANCHOR_MIN_CONFIDENCE     = float(os.environ.get("ANCHOR_MIN_CONFIDENCE", "0.86"))
# MIN_RANK_SCORE_FOR_FALLBACK = float(os.environ.get("MIN_RANK_SCORE_FOR_FALLBACK", "12.0"))

# # Extra gate for AI verify pass — if verify confidence drops below this, veto
# AI_VERIFY_MIN_CONFIDENCE  = float(os.environ.get("AI_VERIFY_MIN_CONFIDENCE", "0.62"))

# FETCH_HEADERS = {
#     "User-Agent": (
#         "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
#         "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
#     ),
#     "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
#     "Accept-Language": "en-US,en;q=0.9",
# }

# # Per-row per-platform confidence (filled in process_row)
# ROW_PLATFORM_CONFIDENCE: Dict[object, Dict[str, float]] = {}
# # Per-row per-platform provenance
# ROW_PLATFORM_SOURCE: Dict[object, Dict[str, str]] = {}
# # Per-row resolved username hints (platform → handle string, e.g. "johndoe123")
# ROW_USERNAME_HINTS: Dict[object, Dict[str, str]] = {}
# WIKIPEDIA_CONTEXT_CACHE: Dict[str, str] = {}

# PLATFORMS: Dict[str, List[str]] = {
#     "Facebook":  ["facebook.com"],
#     "Instagram": ["instagram.com"],
#     "X":         ["x.com", "twitter.com"],
#     "TikTok":    ["tiktok.com"],
#     "YouTube":   ["youtube.com"],
# }

# PLATFORM_CONF_COLUMNS: Dict[str, str] = {p: f"{p} Confidence" for p in PLATFORMS}


# # ─────────────────────────────────────────────
# #  UTILITY HELPERS
# # ─────────────────────────────────────────────

# def is_first_name_only(talent: str) -> bool:
#     parts = re.sub(r"\s+", " ", (talent or "").strip()).split()
#     return len(parts) == 1 and bool(parts[0])


# def _find_column(raw: pd.DataFrame, *candidates: str) -> Optional[str]:
#     cmap = {str(c).strip().lower(): c for c in raw.columns}
#     for cand in candidates:
#         if cand.lower() in cmap:
#             return cmap[cand.lower()]
#     return None


# def _slug_chars(s: str) -> str:
#     return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


# _KNOWN_PLATFORM_SUPPRESSIONS: Dict[str, set] = {
#     # Confirmed by manual review: these lookalike profiles are not BJ Powell.
#     "bjpowell": {"Facebook", "TikTok", "YouTube"},
# }


# def _platform_suppressed_for_talent(talent: str, platform: str) -> bool:
#     name = re.sub(r"\s+[-–—|]\s+[A-Z0-9]{2,8}$", "", (talent or "").strip())
#     return platform in _KNOWN_PLATFORM_SUPPRESSIONS.get(_slug_chars(name), set())


# def _clean_wikipedia_url(value: object) -> str:
#     if value is None or (isinstance(value, float) and pd.isna(value)):
#         return ""
#     url = str(value).strip()
#     if not url or url.lower() == "nan":
#         return ""
#     if "wikipedia.org/wiki/" not in url.lower():
#         return ""
#     if url.startswith("//"):
#         url = "https:" + url
#     elif not re.match(r"^https?://", url, re.I):
#         url = "https://" + url
#     return url.split("#", 1)[0].strip()


# def _wikipedia_title_from_url(wikipedia_url: str) -> str:
#     try:
#         parsed = urlparse(wikipedia_url)
#     except Exception:
#         return ""
#     if "wikipedia.org" not in (parsed.netloc or "").lower():
#         return ""
#     path = parsed.path or ""
#     if "/wiki/" not in path:
#         return ""
#     title = path.split("/wiki/", 1)[1].strip("/")
#     return re.sub(r"\s+", " ", unquote(title).replace("_", " ")).strip()


# def wikipedia_identity_context(wikipedia_url: str) -> str:
#     """Small optional identity anchor from a provided Wikipedia URL."""
#     url = _clean_wikipedia_url(wikipedia_url)
#     if not url:
#         return ""
#     if url in WIKIPEDIA_CONTEXT_CACHE:
#         return WIKIPEDIA_CONTEXT_CACHE[url]

#     title = _wikipedia_title_from_url(url)
#     context_parts: List[str] = []
#     if title:
#         context_parts.append(f"Wikipedia title: {title}")

#     try:
#         parsed = urlparse(url)
#         page_title = (parsed.path or "").split("/wiki/", 1)[1].strip("/")
#         summary_url = f"{parsed.scheme}://{parsed.netloc}/api/rest_v1/page/summary/{page_title}"
#         res = requests.get(summary_url, headers=FETCH_HEADERS, timeout=8)
#         if res.ok:
#             payload = res.json()
#             api_title = str(payload.get("title") or "").strip()
#             description = str(payload.get("description") or "").strip()
#             extract = str(payload.get("extract") or "").strip()
#             if api_title and api_title.lower() != (title or "").lower():
#                 context_parts.insert(0, f"Wikipedia title: {api_title}")
#             if description:
#                 context_parts.append(f"Description: {description}")
#             if extract:
#                 context_parts.append(f"Summary: {extract[:450]}")
#     except Exception as exc:
#         print(f"[WARN] Wikipedia context fetch failed: {exc}")

#     context = " | ".join(context_parts)[:900]
#     WIKIPEDIA_CONTEXT_CACHE[url] = context
#     return context


# def _wikipedia_title_from_context(wikipedia_context: str) -> str:
#     if not wikipedia_context:
#         return ""
#     m = re.search(r"Wikipedia title:\s*([^|]+)", wikipedia_context)
#     return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""


# def _identity_name_aliases(
#     talent: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     wikipedia_context: str = "",
# ) -> List[str]:
#     """Likely public-name aliases for non-brand rows, e.g. married names vs known names."""
#     base = re.sub(r"\s+", " ", (talent or "").strip())
#     aliases: List[str] = []
#     seen: set = set()

#     def add(name: str) -> None:
#         name = re.sub(r"\s+", " ", (name or "").strip())
#         key = name.lower()
#         if name and key not in seen:
#             seen.add(key)
#             aliases.append(name)

#     add(base)
#     wiki_title = _wikipedia_title_from_context(wikipedia_context)
#     if wiki_title:
#         add(wiki_title)

#     exp = parse_entity_expectations(title_category, title_sub_category, talent=base)
#     if not exp.get("expects_brand"):
#         parts = _brand_name_parts(base)
#         if len(parts) >= 3:
#             add(" ".join(parts[:2]))
#             add(f"{parts[0]} {parts[-1]}")
#     return aliases


# def _identity_handle_aliases(
#     talent: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     wikipedia_context: str = "",
# ) -> List[str]:
#     handles: List[str] = []
#     seen: set = set()
#     for alias in _identity_name_aliases(talent, title_category, title_sub_category, wikipedia_context):
#         slug = _slug_chars(alias)
#         if len(slug) >= 4 and slug not in seen:
#             seen.add(slug)
#             handles.append(slug)
#     return handles


# # Vertical/sub-account suffixes on brand handles (not the parent brand page).
# # Common word shortenings in brand handles (Basketball -> BBall, etc.)
# _BRAND_WORD_ABBREV: Dict[str, str] = {
#     "basketball": "bball",
#     "football": "fball",
# }

# _BRAND_OFFICIAL_SUFFIXES = frozenset({
#     "hq", "official", "sports", "sport", "canada",
# })

# _BRAND_VERTICAL_SUFFIXES = frozenset({
#     "football", "basketball", "cbb", "wcbb", "cfb", "nba", "nfl", "mlb", "nhl", "mls",
#     "soccer", "college", "fantasy", "gaming", "news", "sport", "sports", "golf", "tennis",
#     "racing", "mma", "wwe", "esports", "highlights", "podcast", "radio", "tv", "plus",
#     "women", "womens", "men", "kids", "pr", "shop", "store", "fan", "fans",
# })

# _KNOWN_SINGLE_WORD_BRANDS = frozenset({
#     "overtime", "slam",
# })

# _KNOWN_BRAND_PLATFORM_HANDLES: Dict[str, Dict[str, Tuple[str, ...]]] = {
#     "espn": {
#         "Facebook": ("ESPN",),
#         "Instagram": ("espn",),
#         "X": ("espn",),
#         "TikTok": ("espn",),
#         "YouTube": ("espn",),
#     },
#     "foxsports1": {
#         "Instagram": ("fs1",),
#         "X": ("FS1",),
#     },
#     "on3sports": {
#         "Instagram": ("on3",),
#         "X": ("On3sports",),
#     },
#     "overtime": {
#         "Facebook": ("getovertime",),
#     },
#     "overtimeelite": {
#         "Instagram": ("ote",),
#         "YouTube": ("OvertimeElite", "ote"),
#     },
#     "slam": {
#         "Facebook": ("SLAM",),
#         "Instagram": ("slam",),
#         "X": ("SLAMonline",),
#         "YouTube": ("SLAM",),
#     },
#     "slamhs": {
#         "Facebook": ("SLAMHS", "SLAMHighSchool"),
#         "Instagram": ("slam_hs",),
#         "X": ("SLAM_HS",),
#         "YouTube": ("SLAMHS", "SLAMHighSchool"),
#     },
# }


# def _brand_canonical_slug(talent: str) -> str:
#     return _slug_chars(talent)


# def _brand_name_parts(talent: str) -> List[str]:
#     return re.findall(r"[a-zA-Z0-9]+", talent or "")


# def _talent_name_implies_brand(talent: str) -> bool:
#     """
#     Infer media-brand rows when metadata is missing or only says 'Talent'
#     (e.g. 'Overtime Elite' with no Publishers category).
#     """
#     t = (talent or "").strip()
#     if not t:
#         return False
#     parts = _brand_name_parts(t)
#     slug = _slug_chars(t)
#     if slug in _KNOWN_BRAND_PLATFORM_HANDLES or slug in _KNOWN_SINGLE_WORD_BRANDS:
#         return True
#     if is_first_name_only(t):
#         return False
#     if len(parts) < 2:
#         return False
#     org_markers = (
#         "sports", "sport", "network", "media", "elite", "athletic", "tribune",
#         "report", "overtime", "yahoo", "bleacher", "espn", "tsn", "sportsnet",
#         "publication", "broadcast", "channel", "league",
#     )
#     if any(m in slug for m in org_markers):
#         return True
#     if re.search(r"\([A-Z]{2,}\)", talent):
#         return True
#     if len(parts) >= 3 and parts[0].lower() == "the":
#         return True
#     return False


# def _brand_explicit_acronyms(talent: str) -> List[str]:
#     """Acronyms from parentheses (TSN) or ALL-CAPS tokens in the name."""
#     found: List[str] = []
#     seen: set = set()
#     for m in re.finditer(r"\(([A-Za-z0-9]{2,})\)", talent or ""):
#         ac = m.group(1).lower()
#         if ac not in seen:
#             seen.add(ac)
#             found.append(ac)
#     for m in re.finditer(r"\b[A-Z]{2,}\b", talent or ""):
#         ac = m.group(0).lower()
#         if ac not in seen:
#             seen.add(ac)
#             found.append(ac)
#     return found


# def _brand_short_acronyms(parts: List[str]) -> List[str]:
#     """Short handles: oe, ote (O+last-2-letters), tbt, etc."""
#     if len(parts) < 2:
#         return []
#     meaningful = [p for p in parts if p.lower() not in ("the", "and", "of", "a")]
#     if len(meaningful) < 2:
#         meaningful = parts
#     first, last = meaningful[0], meaningful[-1]
#     acronyms: List[str] = []
#     if first and last:
#         acronyms.append("".join(p[0].lower() for p in meaningful if p))
#         if len(last) >= 2:
#             acronyms.append(first[0].lower() + last[:2].lower())
#         if len(last) >= 3:
#             acronyms.append(first[0].lower() + last[-2:].lower())
#     return acronyms


# def _brand_parent_slug(talent: str) -> str:
#     """Parent brand slug when talent ends with a number (FOX Sports 1 -> foxsports)."""
#     parts = _brand_name_parts(talent)
#     if parts and parts[-1].isdigit():
#         return _slug_chars("".join(parts[:-1]))
#     return ""


# def _normal_handle_text(handle: str) -> str:
#     return (handle or "").strip().lstrip("@").rstrip("/").lower()


# def _brand_platform_handles(talent: str, platform: str = "") -> List[str]:
#     """Known official handle variants for brands whose public handles are not name-derived."""
#     aliases = _KNOWN_BRAND_PLATFORM_HANDLES.get(_brand_canonical_slug(talent), {})
#     ordered: List[str] = []
#     seen: set = set()

#     def add_many(values: Tuple[str, ...]) -> None:
#         for value in values:
#             key = _normal_handle_text(value)
#             if key and key not in seen:
#                 seen.add(key)
#                 ordered.append(value)

#     if platform:
#         add_many(aliases.get(platform, ()))
#     add_many(aliases.get("default", ()))
#     if not platform:
#         for plat_values in aliases.values():
#             add_many(plat_values)
#     return ordered


# def _path_handle_text(link: str, platform: str) -> str:
#     """Raw profile handle from the URL path, preserving separators such as underscores."""
#     segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
#     if not segs:
#         return ""
#     if platform == "YouTube" and segs[0].lower() in ("user", "channel", "c") and len(segs) >= 2:
#         return segs[1].lstrip("@")
#     return segs[-1].lstrip("@")


# def _brand_has_strict_platform_handles(talent: str, platform: str) -> bool:
#     return bool(_brand_platform_handles(talent, platform))


# def _brand_search_handles(talent: str) -> List[str]:
#     """Likely official handles: compact slug, acronym (fs1), parent slug for shared YT, etc."""
#     handles: List[str] = []
#     seen: set = set()

#     def add(h: str) -> None:
#         h = (h or "").strip().lstrip("@").lower()
#         if h and h not in seen:
#             seen.add(h)
#             handles.append(h)

#     parts = _brand_name_parts(talent)
#     add(_brand_canonical_slug(talent))
#     for handle in _brand_platform_handles(talent):
#         add(handle)
#     for ac in _brand_explicit_acronyms(talent):
#         add(ac)
#         add(f"{ac}_sports")
#         add(f"{ac}sports")
#         add(f"{ac}_canada")
#         add(f"{ac}canada")
#     if parts:
#         add("".join(p.lower() for p in parts))
#         add("_".join(p.lower() for p in parts))
#         if len(parts) >= 2:
#             add(parts[0].lower() + "_" + "".join(p.lower() for p in parts[1:]))
#         if parts[0].lower() == "the" and len(parts) > 1:
#             add("".join(p.lower() for p in parts[1:]))
#             add("_".join(p.lower() for p in parts))
#             add(parts[0].lower() + "_" + "".join(p.lower() for p in parts[1:]))
#             add("_".join(p.lower() for p in parts[1:]))
#             add("".join(p[0].lower() for p in parts[1:] if p))
#         if len(parts) >= 2:
#             add("".join(p[0].lower() for p in parts if p))
#         for ac in _brand_short_acronyms(parts):
#             add(ac)
#         if parts[-1].isdigit():
#             add("".join(p[0].lower() for p in parts if p))
#         if len(parts[0]) <= 4 and re.match(r"^[a-zA-Z0-9]+$", parts[0]):
#             add(parts[0].lower())
#         if re.search(r"\d", "".join(parts)):
#             add("".join(p[0].lower() for p in parts if p and p[0].isalnum()))
#         abbrev_join = []
#         for p in parts:
#             pl = p.lower()
#             if pl in _BRAND_WORD_ABBREV:
#                 abbrev_join.append(_BRAND_WORD_ABBREV[pl])
#             elif pl != "the":
#                 abbrev_join.append(p.lower())
#         if abbrev_join:
#             add("".join(abbrev_join))
#             if parts and parts[0].lower() == "the":
#                 add("the" + "".join(abbrev_join))
#     for h in list(handles):
#         if 4 <= len(h) <= 18:
#             add(f"{h}hq")
#             add(f"{h}_hq")
#         if 2 <= len(h) <= 14:
#             add(f"{h}_official")
#     return handles[:18]


# def _brand_known_handle_slugs(talent: str) -> set:
#     return {_slug_chars(h) for h in _brand_search_handles(talent)}


# def _brand_handle_rank(talent: str, handle_slug: str) -> int:
#     handle_slug = _slug_chars(handle_slug)
#     for idx, handle in enumerate(_brand_search_handles(talent)):
#         if _slug_chars(handle) == handle_slug:
#             return idx
#     return 999


# def _platform_from_link(link: str) -> str:
#     low = (link or "").lower()
#     for plat, domains in PLATFORMS.items():
#         if any(d in low for d in domains):
#             return plat
#     return ""


# def _is_youtube_channel_id_slug(slug: str) -> bool:
#     return bool(re.match(r"^UC[\w-]{10,}$", slug or "", re.I))


# def _candidate_supports_brand(talent: str, candidate: dict) -> bool:
#     """Title/snippet evidence that a URL (esp. YouTube /channel/UC…) belongs to this brand."""
#     title = (candidate.get("title") or "")
#     snippet = (candidate.get("snippet") or "")
#     blob = f"{title} {snippet}".lower()
#     slug = _slug_chars(talent)
#     if slug and len(slug) >= 5 and slug[: min(10, len(slug))] in _slug_chars(blob):
#         return True
#     for part in _brand_name_parts(talent):
#         pl = part.lower()
#         if pl in ("the", "and", "of", "a") or len(pl) < 3:
#             continue
#         if pl in blob:
#             return True
#     for ac in _brand_explicit_acronyms(talent):
#         if ac in blob:
#             return True
#     for ac in _brand_short_acronyms(_brand_name_parts(talent)):
#         if len(ac) >= 3 and re.search(rf"\b{re.escape(ac.lower())}\b", blob):
#             return True
#     for handle in _brand_platform_handles(talent):
#         h = _normal_handle_text(handle)
#         if len(h) >= 3 and re.search(rf"\b{re.escape(h)}\b", blob):
#             return True
#     return False


# def _path_handle_slug(link: str, platform: str) -> str:
#     """Handle slug from a profile URL (YouTube /user/FoxSports -> foxsports)."""
#     segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
#     if not segs:
#         return ""
#     if platform == "YouTube" and segs[0].lower() in ("user", "channel", "c") and len(segs) >= 2:
#         return _slug_chars(segs[1].lstrip("@"))
#     return _slug_chars(segs[-1].lstrip("@"))


# def _brand_handle_matches_row(
#     talent: str, handle_slug: str, platform: str = "",
# ) -> bool:
#     """True when URL handle matches this brand row (not a parent/vertical/other brand)."""
#     handle_slug = _slug_chars(handle_slug)
#     if not handle_slug:
#         return False
#     if handle_slug in _brand_known_handle_slugs(talent):
#         return True
#     canonical = _brand_canonical_slug(talent)
#     if handle_slug == canonical:
#         return True
#     if canonical and handle_slug.startswith(canonical) and len(handle_slug) > len(canonical):
#         extra = handle_slug[len(canonical):]
#         if extra in _BRAND_OFFICIAL_SUFFIXES:
#             return True
#     parent = _brand_parent_slug(talent)
#     if parent and handle_slug == parent:
#         return platform == "YouTube"
#     first_tokens = re.findall(r"[a-z0-9]+", (talent or "").lower())
#     explicit = _brand_explicit_acronyms(talent)
#     for ac in explicit:
#         if handle_slug == ac:
#             return True
#         if handle_slug.startswith(ac) and handle_slug[len(ac):] in _BRAND_OFFICIAL_SUFFIXES:
#             return True
#     if (
#         first_tokens
#         and handle_slug == _slug_chars(first_tokens[0])
#         and re.search(r"\d", first_tokens[0])
#     ):
#         return True
#     significant = [t for t in first_tokens if len(t) > 1 or t.isdigit()]
#     if not significant or not all(t in handle_slug for t in significant):
#         return False
#     if canonical and handle_slug.startswith(canonical) and len(handle_slug) > len(canonical):
#         extra = handle_slug[len(canonical):]
#         if extra and extra not in _slug_chars(talent) and (
#             extra in _BRAND_VERTICAL_SUFFIXES or len(extra) >= 2
#         ):
#             return False
#     return True


# def _brand_slug_is_vertical(
#     canonical: str, path_slug: str, talent: str, platform: str = "",
# ) -> bool:
#     """True for vertical/sub-brand/parent-mismatch handles (e.g. cbssportscbb, foxsportspr)."""
#     path_slug = _slug_chars(path_slug)
#     canonical = _slug_chars(canonical)
#     if not path_slug:
#         return False
#     if _brand_handle_matches_row(talent, path_slug, platform):
#         return False
#     parent = _brand_parent_slug(talent)
#     if parent and path_slug == parent and platform != "YouTube":
#         return True
#     if canonical and path_slug != canonical and canonical.startswith(path_slug) and len(canonical) > len(path_slug) + 1:
#         return True
#     if canonical and path_slug.startswith(canonical) and len(path_slug) > len(canonical):
#         remainder = path_slug[len(canonical):]
#         if remainder in _BRAND_VERTICAL_SUFFIXES or len(remainder) >= 2:
#             return True
#     return False


# def profile_from_candidate_url(link: str, platform: str) -> str:
#     """Normalize Serper links to a profile root (e.g. facebook.com/foxsports/videos/... -> /foxsports)."""
#     if not link:
#         return ""
#     if is_valid_profile_url(link, platform):
#         return normalize_profile_url(link, platform)
#     low = link.lower()
#     if platform == "Facebook" and "facebook.com" in low:
#         m = re.search(r"facebook\.com/([^/?#]+)", low, re.I)
#         if m:
#             handle = m.group(1).lower()
#             blocked = {
#                 "share", "sharer", "groups", "events", "marketplace", "gaming", "watch",
#                 "people", "pages", "profile.php", "public", "login",
#             }
#             if handle not in blocked:
#                 base = f"https://www.facebook.com/{m.group(1)}"
#                 if is_valid_profile_url(base, platform):
#                     return normalize_profile_url(base, platform)
#     if platform == "X" and ("x.com" in low or "twitter.com" in low):
#         if "/user/status/" in low:
#             return ""
#         m = re.search(r"(?:x|twitter)\.com/([^/?#]+)/", low, re.I)
#         if m:
#             handle = m.group(1)
#             blocked = {"i", "intent", "search", "share", "home"}
#             if handle.lower() not in blocked:
#                 base = f"https://x.com/{handle}"
#                 if is_valid_profile_url(base, platform):
#                     return normalize_profile_url(base, platform)
#     if platform == "TikTok" and "tiktok.com" in low:
#         m = re.search(r"tiktok\.com/@([^/?#]+)", low, re.I)
#         if m:
#             base = f"https://www.tiktok.com/@{m.group(1)}"
#             if is_valid_profile_url(base, platform):
#                 return normalize_profile_url(base, platform)
#     if platform == "YouTube" and "youtube.com" in low:
#         for pattern in (
#             r"youtube\.com/(@[\w.-]+)",
#             r"youtube\.com/user/([\w.-]+)",
#             r"youtube\.com/channel/([\w.-]+)",
#             r"youtube\.com/c/([\w.-]+)",
#         ):
#             m = re.search(pattern, low, re.I)
#             if m:
#                 if "/user/" in m.group(0).lower():
#                     seg = f"user/{m.group(1)}"
#                 elif "/channel/" in m.group(0).lower():
#                     seg = f"channel/{m.group(1)}"
#                 elif "/c/" in m.group(0).lower():
#                     seg = f"c/{m.group(1)}"
#                 else:
#                     seg = m.group(1)
#                 base = f"https://www.youtube.com/{seg}"
#                 if is_valid_profile_url(base, platform):
#                     return normalize_profile_url(base, platform)
#     return ""


# def profile_from_candidate_text(candidate: dict, platform: str) -> str:
#     """Recover profile URLs when Google result text exposes @handle but link is a post/status."""
#     title = candidate.get("title", "") or ""
#     snippet = candidate.get("snippet", "") or ""
#     text = f"{title} {snippet}"

#     handles: List[str] = []
#     for m in re.finditer(r"\(@([A-Za-z0-9_.-]{2,30})\)", text):
#         handles.append(m.group(1))
#     for m in re.finditer(r"\b@([A-Za-z0-9_.-]{2,30})\b", text):
#         handles.append(m.group(1))

#     blocked = {
#         "instagram", "facebook", "twitter", "youtube", "tiktok", "espn",
#         "nba", "ncaabasketball", "sportscenter",
#     }
#     for handle in handles:
#         h = handle.strip().strip(".")
#         if not h or h.lower() in blocked:
#             continue
#         if platform == "X":
#             url = f"https://x.com/{h}"
#         elif platform == "Instagram":
#             url = f"https://www.instagram.com/{h}"
#         elif platform == "TikTok":
#             url = f"https://www.tiktok.com/@{h}"
#         else:
#             continue
#         if is_valid_profile_url(url, platform):
#             return normalize_profile_url(url, platform)
#     return ""


# # ─────────────────────────────────────────────
# #  NAME AMBIGUITY  (NEW)
# # ─────────────────────────────────────────────

# # Very common first names that routinely cause namesake collisions
# _COMMON_FIRST_NAMES = {
#     "andrea", "jessica", "jennifer", "ashley", "brittany", "britney",
#     "samantha", "amanda", "sarah", "emily", "emma", "olivia", "megan",
#     "michael", "james", "john", "david", "robert", "william", "daniel",
#     "matthew", "chris", "jason", "kevin", "ryan", "brian", "tyler",
#     "alex", "jordan", "taylor", "morgan", "charlie", "casey", "drew",
# }

# def get_name_ambiguity_level(talent: str) -> str:
#     """
#     Return 'high', 'medium', or 'low'.
#     High ambiguity → raise confidence threshold by +0.10.
#     Medium → raise by +0.05.
#     Low → use base threshold.
#     """
#     parts = re.sub(r"\s+", " ", (talent or "").strip()).lower().split()
#     if len(parts) == 1:
#         return "high"   # single name: Andrea, Prince, etc.
#     if len(parts) == 2:
#         first = parts[0]
#         if first in _COMMON_FIRST_NAMES:
#             return "medium"
#     return "low"


# def _effective_min_confidence(talent: str) -> float:
#     level = get_name_ambiguity_level(talent)
#     if level == "high":
#         return min(0.92, MIN_CONFIDENCE_EMIT + 0.10)
#     if level == "medium":
#         return min(0.85, MIN_CONFIDENCE_EMIT + 0.05)
#     return MIN_CONFIDENCE_EMIT


# # ─────────────────────────────────────────────
# #  METADATA HELPERS
# # ─────────────────────────────────────────────

# def extract_search_keywords(title_category: str, title_sub_category: str) -> str:
#     parts: List[str] = []
#     for raw in (title_category, title_sub_category):
#         if raw is None or (isinstance(raw, float) and pd.isna(raw)):
#             continue
#         s = str(raw).strip()
#         if not s or s.lower() == "nan":
#             continue
#         parts.append(s)
#     if not parts:
#         return ""
#     text = " ".join(parts)
#     text = text.replace(",", " ").replace("|", " ")
#     text = re.sub(r"[\r\n\t]+", " ", text)
#     text = re.sub(
#         r"(?i)\b(talent type|gender|talent subtype|publication type)\s*-\s*",
#         " ", text,
#     )
#     text = re.sub(r"\s+", " ", text).strip()
#     words = text.split()
#     return " ".join(words[:14])[:140].strip()


# def parse_entity_expectations(
#     title_category: str,
#     title_sub_category: str,
#     talent: str = "",
# ) -> Dict[str, bool]:
#     blob = f"{title_category or ''} {title_sub_category or ''}".lower()
#     cat = (title_category or "").lower()
#     expects_brand = bool(
#         re.search(r"\bpublishers?\b", cat)
#         or re.search(r"\b(publication|network|brand|company|organization|organisation)\b", blob)
#         or re.search(r"\btv\s*network\b", cat)
#         or re.search(r"\bmedia\s*(brand|company|outlet)\b", blob)
#         or _talent_name_implies_brand(talent)
#     )
#     return {
#         "expects_brand":      expects_brand,
#         "expects_male":       bool(re.search(r"gender\s*-\s*man\b", blob)),
#         "expects_female":     bool(re.search(r"gender\s*-\s*woman\b", blob)),
#         "expects_athlete":    False if expects_brand else bool(
#             re.search(r"\bathlete\b", blob)
#             or re.search(r"\bbasketball\b", blob)
#             or re.search(r"\bfootball\b", blob)
#             or (
#                 "sport" in blob
#                 and not re.search(r"publication\s*type", blob)
#             )
#         ),
#         "expects_basketball": False if expects_brand else ("basketball" in blob),
#         "expects_musician":   bool(
#             re.search(r"\bmusician\b|\bsinger\b|\brap(per)?\b|\bartist\b|\bband\b", blob)
#         ),
#         "expects_journalist": bool(
#             re.search(r"\bjournalist\b|\bnews\b|\banchor\b|\breporter\b|\bhost\b|\bmedia\b", blob)
#         ),
#         "expects_executive":  bool(
#             re.search(r"\bceo\b|\bcfo\b|\bcoo\b|\bexecutive\b|\bfounder\b|\bbusiness\b", blob)
#         ),
#         "expects_actor":      bool(
#             re.search(r"\bactor\b|\bactress\b|\bfilm\b|\btelevision\b|\btv star\b", blob)
#         ),
#         "expects_politician": bool(
#             re.search(r"\bpolitician\b|\bsenator\b|\bcongressman\b|\bgovernor\b|\bpresident\b", blob)
#         ),
#     }


# def _metadata_tokens(search_keywords: str) -> List[str]:
#     if not search_keywords:
#         return []
#     parts = re.split(r"[^\w]+", search_keywords.lower())
#     stop = {"the", "and", "for", "type", "talent", "gender", "subtype", "publication", "network", "man", "woman"}
#     return [p for p in parts if p and p not in stop and len(p) > 2]


# def _talent_lookup_name(
#     talent: str,
#     title_category: str = "",
#     title_sub_category: str = "",
# ) -> str:
#     """
#     Clean client-side suffix codes from person/athlete rows before search.
#     Example: "Ray Allen - DAR" should resolve as "Ray Allen", not as a separate identity.
#     """
#     name = re.sub(r"\s+", " ", (talent or "").strip())
#     if not name:
#         return ""
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=name)
#     if exp.get("expects_brand"):
#         return name
#     cleaned = re.sub(r"\s+[-–—|]\s+[A-Z0-9]{2,8}$", "", name).strip()
#     return cleaned or name


# def _athlete_handle_aliases(talent: str) -> List[str]:
#     """Common compact athlete handles: tbey, tbey1, tylerbey, tylerbey1."""
#     parts = [p.lower() for p in _brand_name_parts(talent)]
#     if len(parts) < 2:
#         return []
#     first = _slug_chars(parts[0])
#     last = _slug_chars(parts[-1])
#     if not first or not last:
#         return []
#     aliases: List[str] = []
#     seen: set = set()

#     def add(value: str) -> None:
#         if value and value not in seen and len(value) >= 4:
#             seen.add(value)
#             aliases.append(value)

#     add(first + last)
#     add(f"{first}{last}page")
#     add(f"{first}{last}1")
#     add(first[:1] + last)
#     add(f"{first[:1]}{last}1")
#     if len(first) >= 2:
#         add(first[:2] + last)
#     return aliases[:8]


# # ─────────────────────────────────────────────
# #  CANDIDATE SIGNALS  (NEW)
# # ─────────────────────────────────────────────

# _FOLLOWER_RE = re.compile(
#     r"([\d,\.]+)\s*[KkMmBb]?\s*(?:followers|subscribers|fans)", re.I
# )

# def _parse_follower_count(text: str) -> Optional[float]:
#     m = _FOLLOWER_RE.search(text or "")
#     if not m:
#         return None
#     raw = m.group(1).replace(",", "")
#     try:
#         val = float(raw)
#     except ValueError:
#         return None
#     lower_full = m.group(0).lower()
#     if "b" in lower_full:
#         val *= 1_000_000_000
#     elif "m" in lower_full:
#         val *= 1_000_000
#     elif "k" in lower_full:
#         val *= 1_000
#     return val


# def _candidate_has_exact_profile_title(talent: str, candidate: dict) -> bool:
#     title = (candidate.get("title") or "").lower()
#     t = re.sub(r"\s+", " ", (talent or "").strip()).lower()
#     return bool(
#         t
#         and re.match(rf"^{re.escape(t)}\s*\(@[A-Za-z0-9_.-]{{2,30}}\)", title)
#     )


# def _candidate_has_name_matching_personal_website(talent: str, candidate: dict) -> bool:
#     blob = f"{candidate.get('title') or ''} {candidate.get('snippet') or ''}".lower()
#     name_slug = _slug_chars(talent)
#     website_match = re.search(r"website:\s*https?://(?:www\.)?([^/\s]+)", blob)
#     if not website_match or not name_slug:
#         return False
#     website_host = website_match.group(1).lower()
#     social_hosts = ("instagram.", "x.", "twitter.", "facebook.", "tiktok.", "youtube.")
#     return name_slug in _slug_chars(website_host) and not any(h in website_host for h in social_hosts)


# def build_candidate_signals(talent: str, candidate: dict, platform: str) -> dict:
#     """
#     Pre-compute deterministic evidence signals for one candidate.
#     Passed alongside the raw candidate to the AI so it can reason directly
#     from structured facts rather than re-deriving them.
#     """
#     title    = (candidate.get("title")   or "").strip()
#     snippet  = (candidate.get("snippet") or "").strip()
#     link     = (candidate.get("link")    or "").strip()
#     blob     = f"{title} {snippet}".lower()
#     path     = urlparse(link).path if link else ""
#     slug     = _slug_chars(path)
#     t_slug   = _slug_chars(talent)

#     # --- Name presence ---
#     name_parts = re.sub(r"\s+", " ", talent).lower().split()
#     name_tokens_in_url    = [p for p in name_parts if len(p) >= 3 and _slug_chars(p) in slug]
#     name_tokens_in_title  = [p for p in name_parts if p in title.lower()]
#     name_tokens_in_snippet= [p for p in name_parts if p in snippet.lower()]

#     # --- Full name in URL slug ---
#     full_name_in_url = len(t_slug) >= 5 and t_slug[:min(8, len(t_slug))] in slug

#     # --- Verification signals ---
#     verification_signals = []
#     if "official" in blob:
#         verification_signals.append("official")
#     if "verified" in blob or "✓" in title or "✔" in title:
#         verification_signals.append("verified")
#     if "blue check" in blob or "checkmark" in blob:
#         verification_signals.append("blue_check")

#     # --- Follower count ---
#     follower_count = _parse_follower_count(blob)

#     # --- Profession signals in snippet ---
#     PROFESSION_MARKERS = [
#         "realtor", "real estate", "mortgage", "listing agent",
#         "basketball", "nba", "wnba", "athlete", "ncaa", "espn",
#         "singer", "musician", "rapper", "artist", "album",
#         "journalist", "reporter", "anchor", "news anchor",
#         "actor", "actress", "film", "director",
#         "ceo", "founder", "executive", "entrepreneur",
#         "senator", "governor", "congressman", "politician",
#         "author", "writer", "novelist",
#         "doctor", "physician", "surgeon", "dentist", "lawyer", "attorney",
#         "digital creator", "content creator", "influencer",
#     ]
#     profession_signals = [p for p in PROFESSION_MARKERS if p in blob]

#     # --- URL depth (1 = likely profile) ---
#     url_depth = len([s for s in path.strip("/").split("/") if s])

#     # --- Generic handle penalty ---
#     # Handles shorter than 4 chars or all-digit are suspicious
#     handle_match = re.search(r"tiktok\.com/@(\w+)", link.lower())
#     if not handle_match:
#         handle_match = re.search(r"instagram\.com/(\w+)", link.lower())
#     if not handle_match:
#         handle_match = re.search(r"(?:x|twitter)\.com/(\w+)", link.lower())
#     handle = handle_match.group(1) if handle_match else ""
#     handle_suspicious = bool(handle and (len(handle) <= 3 or handle.isdigit()))

#     return {
#         "name_tokens_in_url":     name_tokens_in_url,
#         "name_tokens_in_title":   name_tokens_in_title,
#         "name_tokens_in_snippet": name_tokens_in_snippet,
#         "full_name_in_url":       full_name_in_url,
#         "verification_signals":   verification_signals,
#         "follower_count":         follower_count,
#         "profession_signals":     profession_signals,
#         "url_depth":              url_depth,
#         "handle":                 handle,
#         "handle_suspicious":      handle_suspicious,
#         "exact_profile_title":    _candidate_has_exact_profile_title(talent, candidate),
#         "personal_website_match": _candidate_has_name_matching_personal_website(talent, candidate),
#         "is_valid_profile_url":   is_valid_profile_url(link, platform),
#         "search_position":        int(candidate.get("search_position", 99) or 99),
#         "recovered_from_text":    bool(candidate.get("recovered_from_text")),
#     }


# # ─────────────────────────────────────────────
# #  CATEGORY DISAMBIGUATION CONTEXT  (NEW)
# # ─────────────────────────────────────────────

# def get_category_disambiguation_context(
#     title_category: str,
#     title_sub_category: str,
#     talent: str = "",
# ) -> str:
#     """
#     Return a category-specific disambiguation instruction block for the AI prompt.
#     This tells the model exactly what signals confirm vs contradict the expected identity.
#     """
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     lines: List[str] = []

#     if exp.get("expects_brand"):
#         handles = ", ".join(f"@{h}" for h in _brand_search_handles(talent)[:4]) if talent else ""
#         lines += [
#             "CATEGORY: This row is a MEDIA BRAND / PUBLISHER / TV NETWORK (organization), not an individual person.",
#             "SELECT the official organization page for this brand on this platform.",
#             "CONFIRM if: URL handle matches the brand name, or title/snippet names the network/outlet/channel.",
#             "CONFIRM if: verified/official signals and high follower counts typical of major media brands.",
#             f"PREFERRED HANDLES (search hints): {handles or '(derive from talent name)'}",
#             "REJECT if: profile is clearly an individual employee, journalist, or fan — unless the talent name is that person.",
#             "REJECT if: vertical/sub-brand page (e.g. /BrandFootball, /BrandPR, /BrandCBB) when the talent is the parent brand.",
#             "REJECT if: parent brand handle when talent is a distinct sub-brand (e.g. foxsports for FOX Sports 1 — prefer foxsports1/FOXSports1).",
#         ]
#         return "\n".join(lines)

#     if exp["expects_athlete"] or exp["expects_basketball"]:
#         sport = "basketball" if exp["expects_basketball"] else "sports/athletics"
#         lines += [
#             f"CATEGORY: This person is a {sport} ATHLETE.",
#             "CONFIRM if: snippet/title mentions basketball, NBA, WNBA, NCAA, sports, hoops, draft, team name.",
#             "REJECT if: snippet/title mentions 'realtor', 'real estate', 'mortgage', 'listing agent', 'homes for sale', 'digital creator' (without sports context).",
#             "REJECT if: the profile clearly depicts a different profession (e.g. financial advisor, fitness trainer unrelated to sports).",
#         ]
#         if exp["expects_male"]:
#             lines.append(
#                 "REJECT if: the profile clearly belongs to a woman when we expect a male athlete (check pronouns, name, bio)."
#             )

#     if exp["expects_musician"]:
#         lines += [
#             "CATEGORY: This person is a MUSICIAN / ARTIST.",
#             "CONFIRM if: mentions music, songs, album, tour, label, artist, rapper, singer, band.",
#             "REJECT if: the profile is clearly a business, brand, or unrelated individual with the same name.",
#             "REJECT if: the profile is a tribute act, cover band, or fan page (look for 'tribute', 'cover', 'fan', 'unofficial').",
#         ]

#     if exp["expects_journalist"]:
#         lines += [
#             "CATEGORY: This person is a JOURNALIST / MEDIA HOST / NEWS ANCHOR.",
#             "CONFIRM if: mentions journalism, news, reporting, broadcasting, anchor, host, network name (CNN, MSNBC, Fox, etc.).",
#             "REJECT if: this is clearly a personal trainer, realtor, or unrelated person with the same name.",
#         ]

#     if exp["expects_actor"]:
#         lines += [
#             "CATEGORY: This person is an ACTOR / ACTRESS.",
#             "CONFIRM if: mentions film, TV, show, movie, series, screen, Broadway, IMDb.",
#             "REJECT if: the profile is a different entertainment professional (musician only) or fan account.",
#         ]

#     if exp["expects_executive"]:
#         lines += [
#             "CATEGORY: This person is a BUSINESS EXECUTIVE / FOUNDER / CEO.",
#             "CONFIRM if: mentions company, startup, CEO, founder, entrepreneur, board, leadership.",
#             "REJECT if: the profile is a personal creator account or unrelated individual.",
#         ]

#     if exp["expects_politician"]:
#         lines += [
#             "CATEGORY: This person is a POLITICIAN / PUBLIC OFFICIAL.",
#             "CONFIRM if: mentions senator, congressman, governor, mayor, representative, campaign, district.",
#             "REJECT if: the profile is a fan/parody account or unrelated public figure.",
#         ]

#     if not lines:
#         lines = [
#             "No specific category metadata available.",
#             "Use name matching and profile authenticity signals (verified, official, follower count) to disambiguate.",
#         ]

#     return "\n".join(lines)


# # ─────────────────────────────────────────────
# #  ENTITY REJECTION  (EXPANDED)
# # ─────────────────────────────────────────────

# def entity_profile_rejected(
#     talent: str,
#     title_category: str,
#     title_sub_category: str,
#     candidate: Optional[dict],
#     platform: str = "",
# ) -> Tuple[bool, str]:
#     """
#     Reject Serper candidates that clearly contradict Excel metadata.
#     Blank > wrong. Expanded to cover more categories.
#     """
#     if not candidate:
#         return False, ""

#     title   = (candidate.get("title")   or "")
#     snippet = (candidate.get("snippet") or "")
#     link    = (candidate.get("link")    or "")
#     blob    = f"{title} {snippet}".lower()
#     exp     = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     plat_for_suppression = platform or _platform_from_link(link)
#     if plat_for_suppression and _platform_suppressed_for_talent(talent, plat_for_suppression):
#         return True, f"{plat_for_suppression} suppressed for this talent after manual validation."

#     if exp.get("expects_brand"):
#         plat = platform or _platform_from_link(link)
#         path_slug = _path_handle_slug(link, plat) if plat else _slug_chars(urlparse(link).path)
#         raw_handle = _normal_handle_text(_path_handle_text(link, plat)) if plat else ""
#         canonical = _brand_canonical_slug(talent)
#         if _brand_slug_is_vertical(canonical, path_slug, talent, plat):
#             return True, "Brand row: URL is a vertical/sub-brand or parent-mismatch handle, not the main brand page."
#         preferred_handles = {
#             _normal_handle_text(h) for h in _brand_platform_handles(talent, plat)
#         }
#         if preferred_handles and raw_handle and raw_handle not in preferred_handles:
#             if (
#                 plat == "YouTube"
#                 and _is_youtube_channel_id_slug(path_slug)
#                 and _candidate_supports_brand(talent, candidate)
#             ):
#                 pass
#             else:
#                 return True, "Brand row: this platform has a more specific official handle for this organization."
#         handle = path_slug
#         if (
#             plat == "YouTube"
#             and _is_youtube_channel_id_slug(handle)
#             and _candidate_supports_brand(talent, candidate)
#         ):
#             pass
#         elif handle and not _brand_handle_matches_row(talent, handle, plat):
#             if plat == "YouTube" and _candidate_supports_brand(talent, candidate):
#                 pass
#             else:
#                 return True, "Brand row: profile handle does not match this organization's known aliases."

#     # ── shared markers ──
#     sport_markers = (
#         "basketball", "nba", "wnba", "ncaa", "college basketball",
#         "draft", "athlete", "espn", "sport", "point guard",
#         "shooting guard", "forward", "center", "hoops", "nba draft",
#         "football", "nfl", "soccer", "mls", "baseball", "mlb",
#     )
#     sport_hit = any(m in blob for m in sport_markers)

#     non_sport_professions = (
#         "realtor", "real estate", "mortgage", "homes realty",
#         "florida homes", "digital creator", "realtor sales",
#         "realty & mortgage", "realty and mortgage",
#         "listing agent", "homes for sale", "property management",
#         "nft creator", "crypto investor", "cryptocurrency", "web3",
#     )
#     non_sport_hit = any(m in blob for m in non_sport_professions)

#     # ── Athlete ──
#     if exp["expects_male"] and _athlete_row_active(exp):
#         lower_link = link.lower()
#         profile_handle = _path_handle_slug(link, platform or _platform_from_link(link))
#         org_profile_markers = (
#             "nbagleague", "gleague", "league", "team", "sports", "sport",
#             "recruit", "rivals", "on3", "espn", "overtime", "athletics",
#             "news", "247sports", "maxpreps", "mbb", "wbb", "mensbball",
#             "womensbball", "bball", "hoops", "athletic",
#         )
#         name_in_page = t = re.sub(r"\s+", " ", (talent or "").strip()).lower()
#         name_evidence = bool(t and (t in blob or t.replace(" ", "") in _slug_chars(blob)))
#         handle_evidence = talent_url_aligned(talent, link, title_category, title_sub_category)
#         if (
#             profile_handle
#             and any(m in profile_handle for m in org_profile_markers)
#             and not handle_evidence
#         ):
#             return (
#                 True,
#                 "Athlete row: profile handle belongs to a league/team/news account, not the athlete.",
#             )
#         if (
#             "facebook.com" in lower_link
#             and (profile_handle.isdigit() or "profile.php" in lower_link)
#             and not (name_evidence and sport_hit)
#         ):
#             return (
#                 True,
#                 "Athlete row: numeric/generic Facebook profile lacks enough identity evidence.",
#             )
#         if non_sport_hit and not sport_hit:
#             return (
#                 True,
#                 "Metadata = male athlete; result is realtor/creator/real-estate professional.",
#             )
#         # Obvious female-named pages when we expect a male athlete
#         female_hits = (
#             "bobbie ", " bobbie", "brittany ", "britney ", "jessica ",
#             "samantha ", "miss ", " mrs ", "she is ", "she's ", "her ",
#         )
#         if any(x in blob for x in female_hits) and not sport_hit:
#             return (
#                 True,
#                 "Profile text signals a different (female) person; talent is a male athlete.",
#             )

#     if exp["expects_female"] and exp["expects_athlete"] and non_sport_hit and not sport_hit:
#         male_lean = (" mr ", "his ", "his own", "father", "husband")
#         if any(x in blob for x in male_lean) and "woman" not in blob:
#             return (
#                 True,
#                 "Metadata = female athlete; result appears to be an unrelated male professional.",
#             )

#     # ── Musician ──
#     if exp["expects_musician"]:
#         non_music = (
#             "realtor", "real estate", "lawyer", "attorney",
#             "doctor", "physician", "financial advisor",
#         )
#         music_confirm = ("music", "artist", "singer", "rapper", "album", "tour", "label", "song")
#         if any(m in blob for m in non_music) and not any(m in blob for m in music_confirm):
#             return (
#                 True,
#                 "Metadata = musician; result is clearly a non-music professional.",
#             )
#         if any(x in blob for x in ("tribute", "tribute band", "fan page", "unofficial")):
#             return True, "Fan/tribute/unofficial page — not the artist's own profile."

#     # ── Journalist / News Anchor ──
#     if exp["expects_journalist"]:
#         non_media = ("realtor", "real estate", "fitness trainer", "personal trainer", "chef")
#         media_confirm = ("news", "journalist", "anchor", "reporter", "host", "broadcasting", "media")
#         if any(m in blob for m in non_media) and not any(m in blob for m in media_confirm):
#             return True, "Metadata = journalist/anchor; result is a non-media professional."

#     # ── Generic: news articles about the person are not their profile ──
#     article_signals = (
#         " - wikipedia", "wikipedia.org", "imdb.com", "biography",
#         "interviews", "profile of ", "article about", "story of",
#         " - espn.com", " | espn", " - bleacher report",
#     )
#     if any(x in (link + blob).lower() for x in article_signals):
#         return True, "Result appears to be an editorial article or wiki page, not a social profile."

#     return False, ""


# # ─────────────────────────────────────────────
# #  CANDIDATE RANKING  (EXPANDED)
# # ─────────────────────────────────────────────

# def candidate_rank_score(
#     talent: str,
#     c: dict,
#     search_keywords: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     username_hints: Optional[Dict[str, str]] = None,
#     platform: str = "",
# ) -> float:
#     title   = (c.get("title")   or "").lower()
#     snippet = (c.get("snippet") or "").lower()
#     link    = (c.get("link")    or "").lower()
#     query   = (c.get("query")   or "").lower()
#     t       = re.sub(r"\s+", " ", (talent or "").strip()).lower()
#     score   = 0.0
#     exp_rank = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     name_aliases = _identity_name_aliases(talent, title_category, title_sub_category)
#     alias_slugs = {_slug_chars(a) for a in name_aliases if a.lower() != t}

#     # ── Authenticity signals ──
#     if "official" in title or "official" in snippet:
#         score += 4.0
#     if "verified" in title or "verified" in snippet or "✓" in (c.get("title") or ""):
#         score += 3.0

#     # ── Name match ──
#     if t and t in title:
#         score += 3.5
#     if t and t in snippet:
#         score += 2.0
#     for alias in name_aliases:
#         alias_l = alias.lower()
#         if alias_l == t or len(alias_l) < 4:
#             continue
#         if alias_l in title:
#             score += 3.0
#         if alias_l in snippet:
#             score += 1.5

#     # ── URL slug alignment (stronger weight) ──
#     path_slug = _slug_chars(urlparse(link).path)
#     name_slug = _slug_chars(talent)
#     if name_slug and len(name_slug) >= 5 and name_slug[:min(8, len(name_slug))] in path_slug:
#         score += 4.0  # strong: full-name prefix in URL
#     else:
#         for part in re.sub(r"\s+", " ", (talent or "").strip()).lower().split():
#             sp = _slug_chars(part)
#             if len(sp) >= 4 and sp in path_slug:
#                 score += 1.5
#     for alias_slug in alias_slugs:
#         if len(alias_slug) >= 5 and alias_slug in path_slug:
#             score += 4.0
#     if platform == "YouTube" and _is_youtube_channel_id_slug(_path_handle_slug(link, platform)):
#         alias_hit = any(
#             alias.lower() in title or alias.lower() in snippet
#             for alias in name_aliases
#             if len(alias) >= 4
#         )
#         if alias_hit:
#             score += 6.0

#     # ── Follower / subscriber signal ──
#     follower_count = _parse_follower_count(f"{title} {snippet}")
#     if follower_count:
#         if follower_count >= 1_000_000:
#             score += 3.0
#         elif follower_count >= 100_000:
#             score += 2.0
#         elif follower_count >= 10_000:
#             score += 1.0

#     # ── Athlete-specific search evidence ──
#     if not exp_rank.get("expects_brand") and _athlete_row_active(exp_rank):
#         blob = f"{title} {snippet}"
#         position = int(c.get("search_position", 99) or 99)
#         profile_handle = _path_handle_slug(link, platform)
#         profile_title_pattern = bool(
#             t
#             and t in title
#             and re.search(r"\(@[A-Za-z0-9_.-]{2,30}\)", c.get("title") or "")
#         )
#         exact_profile_title = bool(
#             _candidate_has_exact_profile_title(talent, c)
#         )
#         if profile_title_pattern and platform in {"X", "Instagram", "TikTok"}:
#             score += 7.0
#             if position <= 3:
#                 score += 2.0
#         if exact_profile_title and platform in {"X", "Instagram", "TikTok"}:
#             score += 8.0
#         if profile_handle and profile_handle in {_slug_chars(a) for a in _athlete_handle_aliases(talent)}:
#             score += 7.0
#         if t and t in blob and any(s in query for s in ("official", "verified")):
#             score += 5.0
#             if position <= 3:
#                 score += 2.0
#         if _candidate_has_name_matching_personal_website(talent, c):
#             score += 6.0
#         if t and t in blob and any(s in query for s in ("basketball", "athlete", "hoops")):
#             score += 4.0
#             if position <= 3:
#                 score += 3.0
#             if c.get("recovered_from_text"):
#                 score += 5.0
#         name_parts = [p for p in t.split() if len(p) >= 3]
#         last_name = name_parts[-1] if name_parts else ""
#         if (
#             last_name
#             and last_name in blob
#             and any(s in query for s in ("basketball", "athlete", "hoops"))
#             and position <= 3
#         ):
#             score += 4.0
#         if profile_handle and any(m in profile_handle for m in ("team", "league", "sports", "recruit", "espn")):
#             score -= 10.0
#         if platform == "Facebook" and (profile_handle.isdigit() or "profile.php" in link):
#             score -= 8.0

#     # ── Metadata keyword alignment ──
#     sport_bias_tokens = {
#         "sports", "sport", "basketball", "football", "publication", "publishers", "publisher",
#     }
#     for token in _metadata_tokens(search_keywords):
#         if len(token) < 3:
#             continue
#         if exp_rank.get("expects_brand") and token in sport_bias_tokens:
#             continue
#         if token in title or token in snippet:
#             score += 1.5

#     # ── Brand handle alignment ──
#     if exp_rank.get("expects_brand"):
#         brand_slug = _path_handle_slug(link, platform) if platform else path_slug
#         raw_handle = _normal_handle_text(_path_handle_text(link, platform)) if platform else ""
#         canonical = _brand_canonical_slug(talent)
#         preferred_handles = [
#             _normal_handle_text(h) for h in _brand_platform_handles(talent, platform)
#         ]
#         if raw_handle and raw_handle in preferred_handles:
#             score += 18.0
#             score += max(0.0, 4.0 - preferred_handles.index(raw_handle))
#         elif preferred_handles and any(_slug_chars(h) == brand_slug for h in preferred_handles):
#             score -= 8.0
#         if platform == "Instagram" and canonical and brand_slug == f"{canonical}hq":
#             score += 13.0
#         elif canonical and brand_slug == canonical:
#             score += 12.0
#         elif _brand_handle_matches_row(talent, brand_slug, platform):
#             rank = _brand_handle_rank(talent, brand_slug)
#             score += max(5.0, 11.0 - min(rank, 8))
#             if brand_slug.endswith("official"):
#                 score -= 1.5
#             for ac in _brand_explicit_acronyms(talent):
#                 ac_slug = _slug_chars(ac)
#                 if platform == "X" and brand_slug in {f"{ac_slug}sports", f"{ac_slug}sport"}:
#                     score += 3.0
#                 elif platform == "TikTok" and brand_slug == ac_slug:
#                     score += 3.0
#         elif platform == "YouTube" and _is_youtube_channel_id_slug(brand_slug):
#             if _candidate_supports_brand(talent, c):
#                 score += 11.0
#         elif _candidate_supports_brand(talent, c):
#             score += 5.0
#         elif canonical and _brand_slug_is_vertical(canonical, brand_slug, talent, platform):
#             score -= 20.0

#     # ── Username hint bonus (cross-platform consistency) ──
#     if username_hints and platform:
#         for src_plat, hint in username_hints.items():
#             if src_plat == platform or not hint:
#                 continue
#             hint_slug = _slug_chars(hint)
#             if hint_slug in path_slug:
#                 hint_bonus = 5.0
#                 if (
#                     not exp_rank.get("expects_brand")
#                     and _athlete_row_active(exp_rank)
#                     and src_plat == "Facebook"
#                     and platform in {"X", "Instagram", "TikTok"}
#                     and hint_slug == name_slug
#                 ):
#                     hint_bonus = 1.0
#                 score += hint_bonus   # exact handle match from another platform is usually strong
#                 break

#     # ── Generic handle penalty ──
#     segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
#     if segs:
#         handle = segs[-1].lstrip("@")
#         if len(handle) <= 2 or handle.isdigit():
#             score -= 3.0

#     # ── Hard entity rejection ──
#     rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, c, platform)
#     if rej:
#         score -= 35.0

#     return score


# # ─────────────────────────────────────────────
# #  QUERY BUILDING  (EXPANDED)
# # ─────────────────────────────────────────────

# def build_queries(
#     talent: str,
#     platform: str,
#     domains: List[str],
#     search_keywords: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     username_hints: Optional[Dict[str, str]] = None,
#     wikipedia_context: str = "",
# ) -> List[str]:
#     kw  = (search_keywords or "").strip()
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     queries: List[str] = []
#     priority_queries: List[str] = []
#     wiki_title = _wikipedia_title_from_context(wikipedia_context)
#     name_aliases = _identity_name_aliases(
#         talent, title_category, title_sub_category, wikipedia_context
#     )
#     handle_aliases = _identity_handle_aliases(
#         talent, title_category, title_sub_category, wikipedia_context
#     )

#     # ── Brand: direct handle searches first (foxsports, on3sports, foxsports1, …) ──
#     if exp.get("expects_brand"):
#         platform_handles = _brand_platform_handles(talent, platform)
#         for handle in platform_handles:
#             for domain in domains:
#                 priority_queries.append(f"site:{domain}/{handle}")
#                 priority_queries.append(f"site:{domain}/@{handle}")
#             priority_queries.append(f'"{handle}" {platform} official')
#         if platform == "Instagram":
#             canonical = _brand_canonical_slug(talent)
#             for domain in domains:
#                 queries.insert(0, f"site:{domain}/{canonical}_hq")
#                 queries.insert(0, f"site:{domain}/{canonical}hq")
#             parts_for_mix = _brand_name_parts(talent)
#             if len(parts_for_mix) >= 2:
#                 mixed = parts_for_mix[0].lower() + "_" + "".join(
#                     p.lower() for p in parts_for_mix[1:]
#                 )
#                 for domain in domains:
#                     queries.insert(0, f"site:{domain}/{mixed}")
#         if platform == "X":
#             for ac in _brand_explicit_acronyms(talent):
#                 for domain in domains:
#                     queries.insert(0, f"site:{domain}/{ac}_sports")
#                     queries.insert(0, f"site:{domain}/{ac}sports")
#         for handle in _brand_search_handles(talent):
#             for domain in domains:
#                 queries.append(f"site:{domain}/{handle}")
#                 queries.append(f"site:{domain}/@{handle}")
#             queries.append(f'"{handle}" {platform} official')
#         parts = _brand_name_parts(talent)
#         parent_slug = _brand_parent_slug(talent)
#         if parent_slug and platform == "YouTube":
#             for domain in domains:
#                 if "youtube" in domain:
#                     queries.insert(0, f"site:{domain}/user/{parent_slug}")
#                     cap = parent_slug[:1].upper() + parent_slug[1:]
#                     queries.insert(0, f"site:{domain}/user/{cap}")
#         if platform == "YouTube":
#             for domain in domains:
#                 if "youtube" in domain:
#                     queries.insert(0, f'site:{domain}/channel "{talent}"')
#                     queries.insert(0, f'site:{domain} "{talent}" official channel')
#         if parts:
#             pascal = "".join(
#                 (p[:1].upper() + p[1:].lower()) if p.isalpha() else p
#                 for p in parts
#             )
#             compact_pascal = "".join(
#                 (p[:1].upper() + p[1:].lower()) if len(p) > 1 else p.lower()
#                 for p in parts
#             )
#             def _pascal_token(p: str) -> str:
#                 pl = p.lower()
#                 if pl in _BRAND_WORD_ABBREV:
#                     ab = _BRAND_WORD_ABBREV[pl]
#                     return ab[:1].upper() + ab[1:].lower()
#                 if p.isalpha():
#                     return p[:1].upper() + p[1:].lower()
#                 return p

#             abbrev_pascal = "".join(_pascal_token(p) for p in parts)
#             for domain in domains:
#                 queries.insert(0, f"site:{domain}/{pascal}")
#                 if compact_pascal.lower() != pascal.lower():
#                     queries.insert(0, f"site:{domain}/{compact_pascal}")
#                 if abbrev_pascal.lower() not in (pascal.lower(), compact_pascal.lower()):
#                     queries.insert(0, f"site:{domain}/{abbrev_pascal}")
#                 if platform == "Instagram":
#                     canonical = _brand_canonical_slug(talent)
#                     queries.insert(0, f"site:{domain}/{canonical}_hq")
#                     queries.insert(0, f"site:{domain}/{canonical}hq")
#         for domain in domains:
#             if wiki_title and wiki_title.lower() != talent.lower():
#                 queries.append(f'site:{domain} "{talent}" "{wiki_title}" official')
#             queries.append(f'site:{domain} "{talent}" official')
#             queries.append(f'site:{domain} "{talent}" verified')
#             queries.append(f'site:{domain} "{talent}"')
#         queries.append(f'"{talent}" {platform} official')
#         queries.append(f'"{talent}" {platform}')
#         seen: set = set()
#         unique: List[str] = []
#         for q in priority_queries + queries:
#             if q not in seen:
#                 seen.add(q)
#                 unique.append(q)
#         return unique

#     for handle in handle_aliases:
#         for domain in domains:
#             priority_queries.append(f"site:{domain}/{handle}")
#             priority_queries.append(f"site:{domain}/@{handle}")
#         priority_queries.append(f'"{handle}" {platform}')

#     for alias in name_aliases:
#         if alias.lower() == talent.lower():
#             continue
#         for domain in domains:
#             priority_queries.append(f'site:{domain} "{alias}" official')
#             priority_queries.append(f'site:{domain} "{alias}" verified')
#             priority_queries.append(f'site:{domain} "{alias}"')
#             if platform == "YouTube":
#                 priority_queries.append(f'site:{domain} "{alias}" channel')
#         priority_queries.append(f'"{alias}" {platform} official')
#         priority_queries.append(f'"{alias}" {platform}')

#     if exp["expects_athlete"] or exp["expects_basketball"]:
#         for domain in domains:
#             if wiki_title and wiki_title.lower() != talent.lower():
#                 queries.append(f'site:{domain} "{talent}" "{wiki_title}" basketball')
#             queries.append(f'site:{domain} "{talent}" basketball')
#             queries.append(f'site:{domain} "{talent}" official')
#             queries.append(f'site:{domain} "{talent}" verified')
#         for handle in _athlete_handle_aliases(talent):
#             for domain in domains:
#                 queries.append(f"site:{domain}/{handle}")
#                 queries.append(f"site:{domain}/@{handle}")
#             queries.append(f'"{handle}" {platform} basketball')

#     # ── Username hint queries — highest-value; run first ──
#     if username_hints:
#         for src_plat, hint in username_hints.items():
#             if src_plat == platform or not hint:
#                 continue
#             for domain in domains:
#                 queries.append(f'site:{domain} "@{hint}"')
#                 queries.append(f'site:{domain} "{hint}"')
#             queries.append(f'"{hint}" {platform}')
#             break   # one hint is enough to add targeted queries

#     # ── Sport-first queries (basketball namesake reduction) ──
#     if exp["expects_male"] and exp["expects_basketball"]:
#         for domain in domains:
#             queries.append(f'site:{domain} "{talent}" basketball')
#             queries.append(f'site:{domain} "{talent}" basketball player')
#             queries.append(f'site:{domain} "{talent}" NCAA basketball')

#     # ── Journalist / anchor queries ──
#     if exp["expects_journalist"]:
#         for domain in domains:
#             queries.append(f'site:{domain} "{talent}" journalist anchor')
#             queries.append(f'site:{domain} "{talent}" news host')

#     # ── Musician queries ──
#     if exp["expects_musician"]:
#         for domain in domains:
#             queries.append(f'site:{domain} "{talent}" music artist')
#             queries.append(f'site:{domain} "{talent}" official artist')

#     # ── Standard domain-scoped queries ──
#     for domain in domains:
#         if wiki_title and wiki_title.lower() != talent.lower():
#             queries.append(f'site:{domain} "{talent}" "{wiki_title}" official')
#         queries.append(f'site:{domain} "{talent}" official')
#         queries.append(f'site:{domain} "{talent}" verified')
#         queries.append(f'site:{domain} "{talent}"')
#         if kw:
#             queries.append(f'site:{domain} "{talent}" {kw} official')
#             queries.append(f'site:{domain} "{talent}" {kw}')

#     # ── Fallback web queries ──
#     queries.append(f'"{talent}" {platform} official')
#     queries.append(f'"{talent}" {platform}')
#     if wiki_title and wiki_title.lower() != talent.lower():
#         queries.append(f'"{talent}" "{wiki_title}" {platform} official')
#     if kw:
#         queries.append(f'"{talent}" {kw} {platform} official')
#         queries.append(f'"{talent}" {kw} {platform}')

#     # ── De-dupe preserving order ──
#     seen: set = set()
#     unique: List[str] = []
#     for q in priority_queries + queries:
#         if q not in seen:
#             seen.add(q)
#             unique.append(q)
#     return unique


# # ─────────────────────────────────────────────
# #  DATA LOADING
# # ─────────────────────────────────────────────

# def _default_talent_table() -> pd.DataFrame:
#     n = len(talent_names)
#     data: Dict[str, List] = {
#         "Talent Name":        list(talent_names),
#         "title_category":     [""] * n,
#         "title_sub_category": [""] * n,
#         WIKIPEDIA_URL_COLUMN: [""] * n,
#     }
#     for p in PLATFORMS:
#         data[p] = [""] * n
#     for c in PLATFORM_CONF_COLUMNS.values():
#         data[c] = [float("nan")] * n
#     data["Confidence"] = [float("nan")] * n
#     data["Source"]     = [""] * n
#     return pd.DataFrame(data)


# def load_talent_table_from_path(excel_path: Path) -> pd.DataFrame:
#     excel_path = Path(excel_path)
#     if not excel_path.is_file():
#         raise ValueError(f"File not found: {excel_path}")
#     suffix = excel_path.suffix.lower()
#     try:
#         raw = pd.read_csv(excel_path) if suffix == ".csv" else pd.read_excel(excel_path)
#     except Exception as exc:
#         raise ValueError(f"Could not read spreadsheet: {exc}") from exc
#     if raw.empty:
#         raise ValueError("The file has no rows.")

#     name_col = _find_column(raw, "Talent Name", "Talent", "title", "Title", "Name")
#     if name_col is None:
#         name_col = raw.columns[0]
#     cat_col = _find_column(raw, "title_category", "de_category", "category", "Title Category")
#     sub_col = _find_column(raw, "title_sub_category", "sub_category", "Title Sub Category", "subtitle")
#     wiki_col = _find_column(
#         raw,
#         WIKIPEDIA_URL_COLUMN,
#         "wikipedia_url",
#         "Wikipedia",
#         "Wiki URL",
#         "wiki_url",
#         "Wiki",
#         "Wikipedia Link",
#     )

#     names_list, cat_list, sub_list, wiki_list = [], [], [], []
#     for i in range(len(raw)):
#         name = str(raw.iloc[i][name_col]).strip()
#         if not name or name.lower() == "nan":
#             continue
#         names_list.append(name)
#         c = raw.iloc[i][cat_col] if cat_col else ""
#         s = raw.iloc[i][sub_col] if sub_col else ""
#         w = raw.iloc[i][wiki_col] if wiki_col else ""
#         cat_list.append("" if (isinstance(c, float) and pd.isna(c)) else str(c).strip())
#         sub_list.append("" if (isinstance(s, float) and pd.isna(s)) else str(s).strip())
#         wiki_list.append(_clean_wikipedia_url(w))

#     if not names_list:
#         raise ValueError("No valid talent names found.")
#     n = len(names_list)
#     out: Dict[str, List] = {
#         "Talent Name":        names_list,
#         "title_category":     cat_list,
#         "title_sub_category": sub_list,
#         WIKIPEDIA_URL_COLUMN: wiki_list,
#     }
#     for p in PLATFORMS:
#         out[p] = [""] * n
#     for c in PLATFORM_CONF_COLUMNS.values():
#         out[c] = [float("nan")] * n
#     out["Confidence"] = [float("nan")] * n
#     out["Source"]     = [""] * n
#     return pd.DataFrame(out)


# def load_talent_table() -> pd.DataFrame:
#     if not TEST_BRANDS_PATH.exists():
#         return _default_talent_table()
#     try:
#         return load_talent_table_from_path(TEST_BRANDS_PATH)
#     except ValueError as exc:
#         print(f"[WARN] {exc}. Using default talent_names.")
#         return _default_talent_table()


# def build_talent_df(names: List[str], platforms: List[str]) -> pd.DataFrame:
#     """Legacy helper: names only, no metadata columns."""
#     data: Dict[str, List] = {"Talent Name": names}
#     for p in platforms:
#         data[p] = [""] * len(names)
#     for p in platforms:
#         data[f"{p} Confidence"] = [float("nan")] * len(names)
#     data["title_category"]     = [""] * len(names)
#     data["title_sub_category"] = [""] * len(names)
#     data[WIKIPEDIA_URL_COLUMN] = [""] * len(names)
#     data["Confidence"]         = [float("nan")] * len(names)
#     data["Source"]             = [""] * len(names)
#     return pd.DataFrame(data)


# # ─────────────────────────────────────────────
# #  URL VALIDATION & NORMALISATION
# # ─────────────────────────────────────────────

# def is_valid_profile_url(link: str, platform: str) -> bool:
#     if not isinstance(link, str) or not link.strip():
#         return False
#     u = link.strip()
#     try:
#         parsed = urlparse(u)
#     except Exception:
#         return False
#     if parsed.scheme not in ("http", "https"):
#         return False
#     host = (parsed.netloc or "").lower()
#     path = (parsed.path or "").lower()
#     full = u.lower()

#     if platform == "Facebook":
#         if "facebook.com" not in host:
#             return False
#         if any(seg in full for seg in ("/posts/", "/photos/", "/videos/", "/watch/",
#                                         "/reel", "/story.php", "/permalink/")):
#             return False
#         if "profile.php" in path or "/people/" in path or "/pages/" in path:
#             return True
#         segs = [s for s in path.strip("/").split("/") if s]
#         if len(segs) == 1 and segs[0] not in (
#             "share", "sharer", "groups", "events", "marketplace", "gaming", "watch"
#         ):
#             return True
#         return False

#     if platform == "Instagram":
#         if "instagram.com" not in host:
#             return False
#         if any(x in full for x in ("/p/", "/reel", "/reels/", "/stories/",
#                                     "/tv/", "/explore/", "/tags/", "/locations/")):
#             return False
#         segs = [s for s in path.strip("/").split("/") if s]
#         return len(segs) == 1

#     if platform == "YouTube":
#         if "youtube.com" not in host and "youtu.be" not in host:
#             return False
#         if any(x in full for x in ("/watch", "/shorts", "/playlist", "/results",
#                                     "/live/", "/feed/", "/attribution_link")):
#             return False
#         return "/@" in full or "/channel/" in full or "/c/" in full or "/user/" in full

#     if platform == "X":
#         if "x.com" not in host and "twitter.com" not in host:
#             return False
#         if "/status/" in full or "/i/" in full or "/intent/" in full or "/search" in full:
#             return False
#         segs = [s for s in path.strip("/").split("/") if s]
#         return len(segs) == 1

#     if platform == "TikTok":
#         if "tiktok.com" not in host:
#             return False
#         if any(x in full for x in ("/video/", "/tag/", "/music/", "/discover", "/foryou")):
#             return False
#         return bool(re.search(r"tiktok\.com/@[^/]+/?$", full))

#     return False


# def normalize_profile_url(url: str, platform: str) -> str:
#     if not url or not isinstance(url, str):
#         return ""
#     u = url.strip()
#     if platform == "YouTube":
#         u = u.replace("://m.youtube.com", "://www.youtube.com")
#         u = u.replace("://music.youtube.com", "://www.youtube.com")
#         if "youtube.com" in u and "www." not in urlparse(u).netloc and "m." not in urlparse(u).netloc:
#             u = u.replace("://youtube.com", "://www.youtube.com")
#     return u.rstrip("/")


# def talent_url_aligned(
#     talent: str,
#     link: str,
#     title_category: str = "",
#     title_sub_category: str = "",
# ) -> bool:
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if exp.get("expects_brand"):
#         plat = _platform_from_link(link)
#         handle_slug = _path_handle_slug(link, plat) if plat else _slug_chars(urlparse(link).path)
#         return _brand_handle_matches_row(talent, handle_slug, plat)
#     t = _slug_chars(talent)
#     if len(t) < 4:
#         return False
#     path_compact = _slug_chars(urlparse(link).path)
#     if len(t) >= 6 and t[:min(8, len(t))] in path_compact:
#         return True
#     for part in re.sub(r"\s+", " ", (talent or "").strip()).lower().split():
#         if len(part) < 3:
#             continue
#         sp = _slug_chars(part)
#         if len(sp) >= 5 and sp in path_compact:
#             return True
#     return False


# def first_valid_profile_link(candidates: List[dict], platform: str) -> str:
#     for item in candidates:
#         if is_valid_profile_url(item.get("link", ""), platform):
#             return normalize_profile_url(item["link"], platform)
#     return ""


# def sort_candidates_for_ai(
#     talent: str,
#     candidates: List[dict],
#     search_keywords: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     username_hints: Optional[Dict[str, str]] = None,
#     platform: str = "",
# ) -> List[dict]:
#     return sorted(
#         candidates,
#         key=lambda c: -candidate_rank_score(
#             talent, c, search_keywords, title_category, title_sub_category,
#             username_hints=username_hints, platform=platform,
#         ),
#     )


# def _prefer_recovered_athlete_handle(
#     talent: str,
#     platform: str,
#     selected: str,
#     candidates: List[dict],
#     title_category: str,
#     title_sub_category: str,
#     search_keywords: str,
# ) -> Optional[dict]:
#     """Prefer @handle recovered from result text when it names the athlete directly."""
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if exp.get("expects_brand") or not _athlete_row_active(exp):
#         return None
#     if platform not in {"X", "Instagram", "TikTok"}:
#         return None

#     selected_norm = normalize_profile_url(selected or "", platform).rstrip("/")
#     selected_score = -999.0
#     for c in candidates:
#         if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == selected_norm:
#             selected_score = candidate_rank_score(
#                 talent, c, search_keywords, title_category, title_sub_category, platform=platform
#             )
#             break

#     full_name = re.sub(r"\s+", " ", (talent or "").strip()).lower()
#     for c in candidates[:4]:
#         if not c.get("recovered_from_text"):
#             continue
#         blob = f"{c.get('title', '')} {c.get('snippet', '')}".lower()
#         if full_name not in blob:
#             continue
#         rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, c, platform)
#         if rej:
#             continue
#         score = candidate_rank_score(
#             talent, c, search_keywords, title_category, title_sub_category, platform=platform
#         )
#         if score >= selected_score - 2.0:
#             return c
#     return None


# def _prefer_personal_website_athlete_profile(
#     talent: str,
#     platform: str,
#     selected: str,
#     candidates: List[dict],
#     title_category: str,
#     title_sub_category: str,
#     search_keywords: str,
# ) -> Optional[dict]:
#     """Prefer a top-ranked athlete profile that exposes a name-matching non-social website."""
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if exp.get("expects_brand") or not _athlete_row_active(exp):
#         return None
#     if platform not in {"X", "Instagram", "TikTok"} or not candidates:
#         return None

#     top = candidates[0]
#     if not (
#         _candidate_has_exact_profile_title(talent, top)
#         and _candidate_has_name_matching_personal_website(talent, top)
#     ):
#         return None

#     selected_norm = normalize_profile_url(selected or "", platform).rstrip("/")
#     top_norm = normalize_profile_url(top.get("link", ""), platform).rstrip("/")
#     if selected_norm == top_norm:
#         return None

#     selected_candidate = next(
#         (
#             c for c in candidates
#             if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == selected_norm
#         ),
#         None,
#     )
#     if selected_candidate and _candidate_has_name_matching_personal_website(talent, selected_candidate):
#         return None

#     top_score = candidate_rank_score(
#         talent, top, search_keywords, title_category, title_sub_category, platform=platform
#     )
#     selected_score = (
#         candidate_rank_score(
#             talent, selected_candidate, search_keywords, title_category, title_sub_category, platform=platform
#         )
#         if selected_candidate
#         else 0.0
#     )
#     if top_score >= selected_score:
#         return top
#     return None


# # ─────────────────────────────────────────────
# #  USERNAME HINT EXTRACTION  (NEW)
# # ─────────────────────────────────────────────

# def extract_username_hints(resolved_links: Dict[str, str]) -> Dict[str, str]:
#     """
#     From already-resolved platform URLs, extract the @handle/slug.
#     e.g. instagram.com/johndoe123 → {"Instagram": "johndoe123"}
#     Used to bias searches and rank on other platforms.
#     """
#     hints: Dict[str, str] = {}
#     for plat, url in resolved_links.items():
#         if not url:
#             continue
#         url = url.strip().rstrip("/")
#         # TikTok: /@handle
#         m = re.search(r"tiktok\.com/@([\w.]+)", url, re.I)
#         if m:
#             hints[plat] = m.group(1)
#             continue
#         # YouTube: /@handle or /c/handle or /channel/...
#         m = re.search(r"youtube\.com/@([\w.]+)", url, re.I)
#         if m:
#             hints[plat] = m.group(1)
#             continue
#         m = re.search(r"youtube\.com/c/([\w.]+)", url, re.I)
#         if m:
#             hints[plat] = m.group(1)
#             continue
#         # Instagram, X, Facebook: domain/handle
#         m = re.search(r"(?:instagram|twitter|x|facebook)\.com/(@?)([\w.]+)", url, re.I)
#         if m:
#             hints[plat] = m.group(2)
#             continue
#     return hints


# # ─────────────────────────────────────────────
# #  AI — MAIN SELECTION  (REWRITTEN)
# # ─────────────────────────────────────────────

# def ai_select_best_profile(
#     talent: str,
#     platform: str,
#     candidates: List[dict],
#     entity_category: str,
#     entity_sub_category: str,
#     search_keywords: str,
#     username_hints: Optional[Dict[str, str]] = None,
#     wikipedia_context: str = "",
# ) -> dict:
#     """
#     Two-phase AI selection:
#       Phase 1 — Evaluate each candidate (ACCEPT / MAYBE / REJECT + reason).
#       Phase 2 — Pick the single best from ACCEPT; fall to MAYBE only if no ACCEPT.
#                If neither exists, return empty.

#     Improvements over original:
#     - Chain-of-thought instructions
#     - Pre-computed signals passed per candidate
#     - Category-specific disambiguation rules from get_category_disambiguation_context()
#     - Stricter blank-preferred instructions
#     - Username hints from other platforms as a high-confidence signal
#     """
#     if not candidates:
#         return {"best_link": "", "confidence": 0.0, "reason": "No candidates provided."}

#     exp_sel = parse_entity_expectations(entity_category, entity_sub_category, talent=talent)
#     cat_context = get_category_disambiguation_context(
#         entity_category, entity_sub_category, talent=talent,
#     )

#     # Enrich candidates with pre-computed signals
#     enriched_candidates = []
#     for c in candidates:
#         signals = build_candidate_signals(talent, c, platform)
#         enriched_candidates.append({
#             "link":    c.get("link", ""),
#             "title":   c.get("title", ""),
#             "snippet": c.get("snippet", ""),
#             "signals": signals,
#             "query": c.get("query", ""),
#         })

#     # Username hints from other resolved platforms
#     hint_lines = []
#     if username_hints:
#         for src_plat, hint in username_hints.items():
#             hint_lines.append(
#                 f"  • On {src_plat} we already found @{hint} — prioritise candidates whose URL contains this handle."
#             )
#     hint_block = (
#         "CROSS-PLATFORM HINTS (from other platforms already resolved for this talent):\n"
#         + "\n".join(hint_lines)
#         if hint_lines else ""
#     )

#     if exp_sel.get("expects_brand"):
#         system_msg = (
#             "You are an expert social media profile resolver for media brands, publishers, and TV networks. "
#             "Your job is to identify the single official organization profile on each platform. "
#             "\n\nCORE RULE: A blank cell is ALWAYS better than a wrong link. "
#             "When uncertain, return empty string for best_link. "
#             "\nNEVER select: posts, videos, reels, individual employee accounts, fan pages, "
#             "news articles, or vertical sub-brands (e.g. /BrandFootball, /BrandPR) unless the talent name is that vertical."
#         )
#     else:
#         system_msg = (
#             "You are an expert social media profile resolver working for a talent research firm. "
#             "Your job is to identify the single official, active social media profile for a real public figure. "
#             "\n\nCORE RULE: A blank cell is ALWAYS better than a wrong link. "
#             "When uncertain, return empty string for best_link. "
#             "\nNEVER select: posts, videos, reels, shorts, news articles, Wikipedia pages, fan pages, "
#             "tribute accounts, or profiles that clearly belong to a different person."
#         )

#     user_msg = f"""
# TALENT: "{talent}"
# PLATFORM: {platform}
# SEARCH KEYWORDS FROM METADATA: {search_keywords or "(none)"}
# WIKIPEDIA IDENTITY CONTEXT: {wikipedia_context or "(not provided)"}
# KNOWN NAME ALIASES: {", ".join(_identity_name_aliases(talent, entity_category, entity_sub_category, wikipedia_context)) or "(none)"}

# CATEGORY CONTEXT:
# {cat_context}

# {hint_block}

# CANDIDATES (each has pre-computed signals to help you):
# {json.dumps(enriched_candidates, indent=2, ensure_ascii=True)}

# INSTRUCTIONS:
# Step 1 — For EACH candidate, classify it as:
#   ACCEPT  → very likely the correct official profile for this talent
#   MAYBE   → possible but uncertain
#   REJECT  → wrong person, fan page, article, or content URL

# Use these signals in order of importance:
#   1. signals.name_tokens_in_url and signals.full_name_in_url  (strongest identifier)
#   2. Wikipedia identity context, when provided, to disambiguate the exact person/brand
#   3. signals.verification_signals (official, verified, blue_check)
#   4. Cross-platform username hints (if handle from another platform appears in URL)
#   5. signals.profession_signals matching the expected category
#   6. signals.follower_count (higher = more credible public figure)
#   7. signals.name_tokens_in_title and signals.name_tokens_in_snippet

# Step 2 — From all ACCEPT candidates, choose the one with the most signals.
#   If no ACCEPT, choose from MAYBE only if confidence ≥ 0.75.
#   If no suitable candidate: return best_link="" and confidence<0.40.

# BLANK RULES (return best_link="" if any of these apply):
#   • All candidates appear to be the wrong person
#   • The talent name is very common (e.g. "Andrea", "Jessica") and no candidate clearly confirms identity
#   • The best candidate's profession_signals conflict with the expected category (e.g. realtor for an athlete)
#   • The best candidate is clearly a fan/tribute/unofficial page
#   • You cannot distinguish between 2+ legitimate people with the same name
#   • For MEDIA BRANDS: reject parent-brand URLs when the talent is a sub-brand (e.g. foxsports for FOX Sports 1)

# OUTPUT FORMAT — strict JSON only, no markdown, no extra keys:
# {{
#   "phase1_evaluation": [
#     {{"link": "...", "verdict": "ACCEPT|MAYBE|REJECT", "reason": "short reason"}}
#   ],
#   "best_link": "URL or empty string",
#   "confidence": 0.0,
#   "reason": "one sentence"
# }}
# """

#     body = {
#         "model": OPENAI_CHAT_MODEL,
#         "temperature": 0,
#         "messages": [
#             {"role": "system", "content": system_msg},
#             {"role": "user",   "content": user_msg},
#         ],
#     }
#     headers = {
#         "Authorization": f"Bearer {OPENAI_API_KEY}",
#         "Content-Type": "application/json",
#     }
#     response = requests.post(
#         "https://api.openai.com/v1/chat/completions",
#         headers=headers, json=body, timeout=60,
#     )
#     response.raise_for_status()
#     content = response.json()["choices"][0]["message"]["content"]
#     parsed  = _extract_json_obj(content)

#     best_link  = str(parsed.get("best_link", "") or "").strip()
#     confidence = float(parsed.get("confidence", 0.0))
#     reason     = str(parsed.get("reason", "") or "").strip()
#     confidence = max(0.0, min(1.0, confidence))

#     # Log phase1 evaluation for debugging
#     for ev in (parsed.get("phase1_evaluation") or []):
#         verdict = ev.get("verdict", "?")
#         link_ev = (ev.get("link") or "")[:80]
#         print(f"  [PHASE1] {verdict:6s} | {link_ev} — {ev.get('reason','')}")

#     # Safety: if AI returned something outside the candidate list, discard
#     candidate_links = {normalize_profile_url(c.get("link", ""), platform).rstrip("/")
#                        for c in candidates}
#     if best_link:
#         norm = normalize_profile_url(best_link, platform).rstrip("/")
#         if norm not in candidate_links:
#             print(f"[WARN] AI returned a link not in candidates — discarding: {best_link}")
#             best_link  = ""
#             confidence = min(confidence, 0.30)
#             reason     = f"AI link not in candidate set (discarded). {reason}"

#     return {"best_link": best_link, "confidence": confidence, "reason": reason or "No confident match."}


# # ─────────────────────────────────────────────
# #  AI — VERIFICATION PASS  (NEW)
# # ─────────────────────────────────────────────

# def ai_verify_selected_link(
#     talent: str,
#     platform: str,
#     link: str,
#     title: str,
#     snippet: str,
#     entity_category: str,
#     entity_sub_category: str,
#     search_keywords: str,
#     wikipedia_context: str = "",
# ) -> Tuple[bool, float, str]:
#     """
#     Quick second AI call: "Does this specific URL definitively belong to [talent]?"
#     Returns (verified: bool, adjusted_confidence: float, reason: str).
#     If verified=False, the caller should blank the result.
#     """
#     exp_v = parse_entity_expectations(entity_category, entity_sub_category, talent=talent)
#     cat_context = get_category_disambiguation_context(
#         entity_category, entity_sub_category, talent=talent,
#     )
#     signals = build_candidate_signals(talent, {"link": link, "title": title, "snippet": snippet}, platform)

#     if exp_v.get("expects_brand"):
#         system_msg = (
#             "You are a fact-checker verifying whether a social URL is the official page for a "
#             "media brand/publisher/network (organization). Answer with strict JSON only."
#         )
#         verify_q = (
#             f"Does this URL clearly belong to the organization \"{talent}\" on {platform}?"
#         )
#         verify_no = (
#             "Answer NO if: clearly a different brand, an individual employee/fan account, "
#             "a vertical sub-page (PR/CFB/etc.) when the talent is the parent brand, "
#             "or a news article — not the org profile."
#         )
#     else:
#         system_msg = (
#             "You are a fact-checker verifying whether a specific social media URL belongs "
#             "to a specific public figure. Answer with strict JSON only."
#         )
#         verify_q = (
#             f"Does this URL clearly and definitively belong to the talent named above on {platform}?"
#         )
#         verify_no = (
#             "Answer NO if:\n"
#             "  • This is clearly a different person\n"
#             "  • This is a fan/tribute/unofficial page\n"
#             "  • This is a news article or Wikipedia page\n"
#             "  • There is not enough evidence to confirm identity"
#         )
#     user_msg = f"""
# TALENT: "{talent}"
# PLATFORM: {platform}
# CATEGORY CONTEXT: {cat_context}
# SEARCH KEYWORDS: {search_keywords or "(none)"}
# WIKIPEDIA IDENTITY CONTEXT: {wikipedia_context or "(not provided)"}
# KNOWN NAME ALIASES: {", ".join(_identity_name_aliases(talent, entity_category, entity_sub_category, wikipedia_context)) or "(none)"}

# URL TO VERIFY: {link}
# PAGE TITLE:    {title}
# SNIPPET:       {snippet}
# PRE-COMPUTED SIGNALS: {json.dumps(signals, ensure_ascii=True)}

# QUESTION: {verify_q}

# Answer YES only if you are confident this is the official profile for this entity.
# {verify_no}

# Output strict JSON only:
# {{"verified": true/false, "confidence": 0.0, "reason": "one sentence"}}
# """
#     body = {
#         "model": OPENAI_CHAT_MODEL,
#         "temperature": 0,
#         "messages": [
#             {"role": "system", "content": system_msg},
#             {"role": "user",   "content": user_msg},
#         ],
#     }
#     headers = {
#         "Authorization": f"Bearer {OPENAI_API_KEY}",
#         "Content-Type": "application/json",
#     }
#     try:
#         response = requests.post(
#             "https://api.openai.com/v1/chat/completions",
#             headers=headers, json=body, timeout=40,
#         )
#         response.raise_for_status()
#         parsed = _extract_json_obj(response.json()["choices"][0]["message"]["content"])
#         verified   = bool(parsed.get("verified", False))
#         conf       = max(0.0, min(1.0, float(parsed.get("confidence", 0.0))))
#         verify_rsn = str(parsed.get("reason", "")).strip()
#         return verified, conf, verify_rsn
#     except Exception as exc:
#         print(f"[WARN] Verify call failed: {exc}")
#         return True, 0.0, f"Verify skipped ({exc})"


# # ─────────────────────────────────────────────
# #  EMISSION GATE  (STRICTER DYNAMIC THRESHOLD)
# # ─────────────────────────────────────────────

# def decide_emitted_link(
#     talent: str,
#     platform: str,
#     selected: str,
#     confidence: float,
#     reason: str,
#     top_candidate: Optional[dict],
#     search_keywords: str,
#     title_category: str = "",
#     title_sub_category: str = "",
#     emit_candidate: Optional[dict] = None,
# ) -> Tuple[str, float, str]:
#     """
#     Final gate before writing to Excel.
#     Uses a dynamic threshold based on name ambiguity — common / single names
#     require higher confidence before we emit anything.
#     """
#     effective_min = _effective_min_confidence(talent)
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if _platform_suppressed_for_talent(talent, platform):
#         return "", 0.0, f"{platform} suppressed for this talent after manual validation."

#     if not selected or selected == "Not Found":
#         return "", confidence, reason or "No selection."

#     selected = normalize_profile_url(selected, platform)
#     if not is_valid_profile_url(selected, platform):
#         return "", 0.0, "Rejected: not a valid profile/channel URL."

#     if emit_candidate:
#         rej, why = entity_profile_rejected(
#             talent, title_category, title_sub_category, emit_candidate, platform
#         )
#         if rej:
#             return "", min(confidence, 0.12), why

#     if confidence >= effective_min:
#         return selected, confidence, reason

#     # Fallback: strong deterministic rank + URL name alignment
#     if top_candidate is not None:
#         rej_fb, rej_msg = entity_profile_rejected(
#             talent, title_category, title_sub_category, top_candidate, platform
#         )
#         if rej_fb:
#             return "", confidence, f"Omitted: {rej_msg}"
#         rs   = candidate_rank_score(
#             talent, top_candidate, search_keywords, title_category, title_sub_category,
#         )
#         link = top_candidate.get("link", "")
#         prof = profile_from_candidate_url(link, platform) or link
#         url_ok = (
#             rs >= MIN_RANK_SCORE_FOR_FALLBACK
#             and talent_url_aligned(talent, prof, title_category, title_sub_category)
#             and is_valid_profile_url(prof, platform)
#         )
#         if exp.get("expects_brand"):
#             path_slug = _path_handle_slug(prof, platform)
#             url_ok = url_ok or (
#                 rs >= MIN_RANK_SCORE_FOR_FALLBACK - 2.0
#                 and _brand_handle_matches_row(talent, path_slug, platform)
#                 and is_valid_profile_url(prof, platform)
#             )
#         if url_ok:
#             fb_conf = min(0.85, max(confidence, rs / 20.0))
#             if exp.get("expects_brand"):
#                 fb_conf = max(fb_conf, 0.80)
#             return (
#                 normalize_profile_url(prof, platform),
#                 fb_conf,
#                 f"Strong search rank + URL match ({rs:.1f}): {reason}",
#             )

#     return "", confidence, f"Omitted (below {effective_min:.2f}): {reason}"


# # ─────────────────────────────────────────────
# #  SERPER SEARCH
# # ─────────────────────────────────────────────

# def serper_search(query: str, num_results: int = 10) -> List[dict]:
#     url = "https://google.serper.dev/search"
#     headers = {"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"}
#     payload = {"q": query, "num": max(1, min(num_results, 10))}
#     response = requests.post(url, headers=headers, json=payload, timeout=30)
#     try:
#         response.raise_for_status()
#     except requests.HTTPError as exc:
#         try:
#             detail = response.json().get("message") or response.text
#         except ValueError:
#             detail = response.text
#         raise RuntimeError(f"Serper search failed ({response.status_code}): {detail}") from exc
#     data = response.json()
#     return [
#         {
#             "title":   item.get("title", "") or "",
#             "snippet": item.get("snippet", "") or "",
#             "link":    item.get("link", "") or "",
#         }
#         for item in data.get("organic", [])
#     ]


# def _extract_json_obj(text: str) -> dict:
#     if not text:
#         raise ValueError("Empty OpenAI response.")
#     start = text.find("{")
#     end   = text.rfind("}")
#     if start == -1 or end == -1 or end <= start:
#         raise ValueError("No JSON object found in OpenAI response.")
#     return json.loads(text[start: end + 1])


# # ─────────────────────────────────────────────
# #  HTML ENRICHMENT (unchanged)
# # ─────────────────────────────────────────────

# URL_IN_TEXT_RE = re.compile(r"https?://[^\s\"\'<>\)\]]+", re.I)


# def fetch_html(url: str) -> str:
#     try:
#         r = requests.get(url, headers=FETCH_HEADERS, timeout=20, allow_redirects=True)
#         r.raise_for_status()
#         if len(r.content) > 2_500_000:
#             return ""
#         return r.text or ""
#     except Exception as exc:
#         print(f"[WARN] fetch failed {url[:90]}… : {exc}")
#         return ""


# def extract_urls_from_html(html: str) -> List[str]:
#     if not html:
#         return []
#     found: set = set()
#     for m in URL_IN_TEXT_RE.finditer(html):
#         u = m.group(0).rstrip(".,);\\]}\"'")
#         if u.startswith("http"):
#             found.add(u.split("&utm_")[0])
#     for m in re.finditer(r'href\s*=\s*["\']([^"\']+)["\']', html, re.I):
#         h = m.group(1).strip()
#         if h.startswith("http"):
#             found.add(h.split("&utm_")[0])
#     return list(found)


# def _platform_for_discovered_url(url: str) -> Optional[str]:
#     for plat in PLATFORMS:
#         if is_valid_profile_url(url, plat):
#             return plat
#     u = url.lower()
#     if "linktr.ee/" in u or "linktree.com/" in u or "lnk.bio" in u or "beacons.ai" in u:
#         return "__link_hub__"
#     return None


# def _extract_visible_social_handles(html: str) -> Dict[str, str]:
#     """Recover handles shown as text in profile bios when no full URL is present."""
#     if not html:
#         return {}
#     text = re.sub(r"\\u([0-9a-fA-F]{4})", lambda m: chr(int(m.group(1), 16)), html)
#     text = re.sub(r"<[^>]+>", " ", text)
#     text = re.sub(r"\s+", " ", text)
#     out: Dict[str, str] = {}
#     patterns = {
#         "Instagram": r"(?:instagram|ig)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
#         "TikTok": r"(?:tiktok|tik tok)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
#         "X": r"(?:twitter|x\.com)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
#     }
#     blocked = {"instagram", "tiktok", "twitter", "facebook", "youtube", "official"}
#     for platform, pattern in patterns.items():
#         for m in re.finditer(pattern, text, re.I):
#             handle = m.group(1).strip("._-")
#             if handle.lower() in blocked:
#                 continue
#             out[platform] = handle
#             break
#     return out


# def _url_for_platform_handle(platform: str, handle: str) -> str:
#     handle = (handle or "").strip().lstrip("@")
#     if not handle:
#         return ""
#     if platform == "Instagram":
#         return f"https://www.instagram.com/{handle}"
#     if platform == "TikTok":
#         return f"https://www.tiktok.com/@{handle}"
#     if platform == "X":
#         return f"https://x.com/{handle}"
#     if platform == "Facebook":
#         return f"https://www.facebook.com/{handle}"
#     return ""


# def extract_social_links_from_page(page_url: str, source_platform: str) -> Dict[str, str]:
#     out: Dict[str, str] = {}
#     to_fetch = [page_url]
#     if source_platform == "YouTube":
#         base = page_url.split("?")[0].rstrip("/")
#         if "/@" in base or "/channel/" in base or "/c/" in base or "/user/" in base:
#             if "/about" not in base:
#                 to_fetch.append(base + "/about")
#     hubs_fetched = 0
#     seen_fetch: set = set()
#     for u in to_fetch:
#         u = u.strip()
#         if not u or u in seen_fetch:
#             continue
#         seen_fetch.add(u)
#         html = fetch_html(u)
#         for raw in extract_urls_from_html(html):
#             raw  = raw.strip().rstrip(".,);")
#             plat = _platform_for_discovered_url(raw)
#             if plat and plat != "__link_hub__" and plat not in out:
#                 out[plat] = normalize_profile_url(raw, plat)
#             elif plat == "__link_hub__" and hubs_fetched < 3:
#                 hubs_fetched += 1
#                 for raw2 in extract_urls_from_html(fetch_html(raw)):
#                     raw2 = raw2.strip().rstrip(".,);")
#                     p2   = _platform_for_discovered_url(raw2)
#                     if p2 and p2 != "__link_hub__" and p2 not in out:
#                         out[p2] = normalize_profile_url(raw2, p2)
#         for plat, handle in _extract_visible_social_handles(html).items():
#             if plat in out:
#                 continue
#             url = _url_for_platform_handle(plat, handle)
#             if url and is_valid_profile_url(url, plat):
#                 out[plat] = normalize_profile_url(url, plat)
#     return out


# def enrich_row_from_anchor_profiles(df: pd.DataFrame, row_label: object) -> None:
#     anchor_order = ["Instagram", "YouTube", "X", "Facebook", "TikTok"]
#     confs = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
#     best_plat, best_url, best_c = None, "", 0.0
#     for p in anchor_order:
#         url = str(df.at[row_label, p] or "").strip()
#         if not url:
#             continue
#         c = float(confs.get(p, 0.0))
#         if c < ANCHOR_MIN_CONFIDENCE:
#             continue
#         if c > best_c:
#             best_plat, best_url, best_c = p, url, c
#     if not best_url or not best_plat:
#         return
#     talent = str(df.at[row_label, "Talent Name"] or "")
#     print(f"[ENRICH] {talent} <- anchor {best_plat} (conf={best_c:.2f})")
#     try:
#         discovered = extract_social_links_from_page(best_url, best_plat)
#     except Exception as exc:
#         print(f"[WARN] enrich failed: {exc}")
#         return
#     if not discovered and best_plat == "Facebook":
#         fb_handle = _handle_slug_from_profile(best_url, "Facebook")
#         alias_slugs = {
#             _slug_chars(h)
#             for h in _identity_handle_aliases(
#                 talent,
#                 str(df.at[row_label, "title_category"] or ""),
#                 str(df.at[row_label, "title_sub_category"] or ""),
#                 wikipedia_identity_context(
#                     df.at[row_label, WIKIPEDIA_URL_COLUMN]
#                     if WIKIPEDIA_URL_COLUMN in df.columns
#                     else ""
#                 ),
#             )
#         }
#         if fb_handle and fb_handle in alias_slugs:
#             for tgt in ("Instagram", "TikTok"):
#                 guessed = _url_for_platform_handle(tgt, fb_handle)
#                 if guessed and is_valid_profile_url(guessed, tgt):
#                     discovered.setdefault(tgt, guessed)
#     for tgt, link in discovered.items():
#         if tgt not in PLATFORMS:
#             continue
#         if _platform_suppressed_for_talent(talent, tgt):
#             continue
#         if str(df.at[row_label, tgt] or "").strip():
#             continue
#         if not is_valid_profile_url(link, tgt):
#             continue
#         df.at[row_label, tgt] = link
#         conf_value = round(min(0.93, best_c * 0.96), 3)
#         ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[tgt] = conf_value
#         df.at[row_label, PLATFORM_CONF_COLUMNS[tgt]]           = conf_value
#         ROW_PLATFORM_SOURCE.setdefault(row_label, {})[tgt]     = "bio_enrich"
#         print(f"  + filled {tgt} from bio/link hub")
#     _refresh_row_aggregate_confidence(df, row_label)


# def _refresh_row_aggregate_confidence(df: pd.DataFrame, row_label: object) -> None:
#     parts = [
#         float(ROW_PLATFORM_CONFIDENCE.get(row_label, {}).get(p, 0.0))
#         for p in PLATFORMS
#         if str(df.at[row_label, p] or "").strip()
#     ]
#     if parts:
#         df.at[row_label, "Confidence"] = round(sum(parts) / len(parts), 4)


# def _refresh_row_source_cell(df: pd.DataFrame, row_label: object) -> None:
#     parts = []
#     for p in PLATFORMS:
#         if not str(df.at[row_label, p] or "").strip():
#             continue
#         src = ROW_PLATFORM_SOURCE.get(row_label, {}).get(p, "")
#         if src:
#             parts.append(f"{p}:{src}")
#     df.at[row_label, "Source"] = "; ".join(parts)


# def _handle_slug_from_profile(url: str, platform: str) -> str:
#     if platform == "TikTok":
#         m = re.search(r"tiktok\.com/@([^/?#]+)", url or "", re.I)
#         return _slug_chars(m.group(1)) if m else ""
#     return _path_handle_slug(url, platform)


# def _find_exact_handle_profile(
#     handle: str,
#     talent: str,
#     platform: str,
#     domains: List[str],
#     title_category: str,
#     title_sub_category: str,
# ) -> Optional[dict]:
#     """Search for the same confirmed handle on another platform using real SERP results only."""
#     if _platform_suppressed_for_talent(talent, platform):
#         return None
#     handle_clean = (handle or "").strip().lstrip("@")
#     handle_slug = _slug_chars(handle_clean)
#     if len(handle_slug) < 4:
#         return None

#     queries: List[str] = []
#     for domain in domains:
#         queries.append(f"site:{domain}/{handle_clean}")
#         queries.append(f"site:{domain}/@{handle_clean}")
#         queries.append(f'site:{domain} "{handle_clean}"')
#     queries.append(f'"{handle_clean}" {platform}')

#     seen: set = set()
#     for query in queries:
#         try:
#             results = serper_search(query, num_results=RESULTS_PER_QUERY)
#         except Exception as exc:
#             print(f"[WARN] handle reconcile failed '{query}': {exc}")
#             continue
#         for pos, item in enumerate(results, start=1):
#             raw_link = item.get("link", "")
#             prof = (
#                 profile_from_candidate_url(raw_link, platform)
#                 or profile_from_candidate_text(item, platform)
#                 or raw_link
#             )
#             if not prof or prof in seen or not is_valid_profile_url(prof, platform):
#                 continue
#             seen.add(prof)
#             if _handle_slug_from_profile(prof, platform) != handle_slug:
#                 continue
#             cand = {**item, "link": prof, "search_position": pos, "query": query}
#             rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, cand, platform)
#             if not rej:
#                 return cand
#         time.sleep(0.2)
#     return None


# def _is_initial_style_person_name(talent: str) -> bool:
#     parts = _brand_name_parts(talent)
#     if len(parts) < 2:
#         return False
#     first = parts[0]
#     return bool(len(first) <= 3 and first.upper() == first and re.search(r"[A-Z]", first))


# def reconcile_initial_name_x_instagram_handle(
#     df: pd.DataFrame,
#     row_label: object,
#     resolved_links: Dict[str, str],
#     title_category: str,
#     title_sub_category: str,
# ) -> None:
#     """For initial-style names (BJ/CJ/etc.), use a strong X handle to correct Instagram."""
#     talent = str(df.at[row_label, "Talent Name"] or "").strip()
#     if not _is_initial_style_person_name(talent):
#         return
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if exp.get("expects_brand"):
#         return
#     confs = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
#     x_url = resolved_links.get("X", "")
#     x_conf = float(confs.get("X", 0.0))
#     if not x_url or x_conf < 0.85:
#         return
#     x_handle = _handle_slug_from_profile(x_url, "X")
#     if len(x_handle) < 4:
#         return

#     current_ig = str(df.at[row_label, "Instagram"] or "").strip()
#     current_ig_handle = _handle_slug_from_profile(current_ig, "Instagram") if current_ig else ""
#     if current_ig_handle == x_handle:
#         return

#     cand = _find_exact_handle_profile(
#         x_handle, talent, "Instagram", PLATFORMS["Instagram"],
#         title_category, title_sub_category,
#     )
#     if not cand:
#         return
#     score = candidate_rank_score(
#         talent, cand, extract_search_keywords(title_category, title_sub_category),
#         title_category, title_sub_category, platform="Instagram",
#     )
#     if score < 5.0:
#         return

#     link = normalize_profile_url(cand["link"], "Instagram")
#     conf_value = round(min(0.92, max(0.86, x_conf * 0.98)), 3)
#     print(
#         f"  [INITIAL-HANDLE-FIX] Instagram | {talent} | "
#         f"using X @{x_handle} -> {link[:90]}"
#     )
#     df.at[row_label, "Instagram"] = link
#     ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})["Instagram"] = conf_value
#     df.at[row_label, PLATFORM_CONF_COLUMNS["Instagram"]] = conf_value
#     ROW_PLATFORM_SOURCE.setdefault(row_label, {})["Instagram"] = "x_instagram_handle_match"
#     resolved_links["Instagram"] = link


# def reconcile_athlete_handles_from_confirmed_profiles(
#     df: pd.DataFrame,
#     row_label: object,
#     resolved_links: Dict[str, str],
#     title_category: str,
#     title_sub_category: str,
# ) -> None:
#     """For athlete rows, use a high-confidence handle to fix same-handle profiles on other platforms."""
#     talent = str(df.at[row_label, "Talent Name"] or "").strip()
#     exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
#     if exp.get("expects_brand") or not _athlete_row_active(exp):
#         return

#     confs = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
#     anchors = [
#         (p, resolved_links.get(p, ""), float(confs.get(p, 0.0)))
#         for p in ("X", "Instagram", "TikTok")
#         if resolved_links.get(p) and float(confs.get(p, 0.0)) >= 0.85
#     ]
#     for src_plat, src_url, src_conf in sorted(anchors, key=lambda x: -x[2]):
#         handle = _handle_slug_from_profile(src_url, src_plat)
#         if len(handle) < 4:
#             continue
#         for target_plat, domains in PLATFORMS.items():
#             if target_plat == src_plat or target_plat not in {"Instagram", "X", "TikTok", "YouTube"}:
#                 continue
#             current = str(df.at[row_label, target_plat] or "").strip()
#             current_conf = float(confs.get(target_plat, 0.0))
#             if current and _handle_slug_from_profile(current, target_plat) == handle:
#                 continue
#             if current and current_conf >= 0.55:
#                 continue
#             cand = _find_exact_handle_profile(
#                 handle, talent, target_plat, domains, title_category, title_sub_category
#             )
#             if not cand:
#                 continue
#             link = normalize_profile_url(cand["link"], target_plat)
#             conf_value = round(min(0.92, src_conf * 0.98), 3)
#             print(
#                 f"  [HANDLE-RECONCILE] {target_plat} | {talent} | "
#                 f"@{handle} from {src_plat} -> {link[:90]}"
#             )
#             df.at[row_label, target_plat] = link
#             ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[target_plat] = conf_value
#             df.at[row_label, PLATFORM_CONF_COLUMNS[target_plat]] = conf_value
#             ROW_PLATFORM_SOURCE.setdefault(row_label, {})[target_plat] = "handle_reconcile"
#             resolved_links[target_plat] = link


# # ─────────────────────────────────────────────
# #  SEARCH ONE PLATFORM  (with verify pass added)
# # ─────────────────────────────────────────────

# def search_one_platform(
#     talent: str,
#     platform: str,
#     domains: List[str],
#     title_category: str,
#     title_sub_category: str,
#     username_hints: Optional[Dict[str, str]] = None,
#     wikipedia_context: str = "",
# ) -> Tuple[str, str, float, str]:
#     search_keywords = extract_search_keywords(title_category, title_sub_category)
#     lookup_talent = _talent_lookup_name(talent, title_category, title_sub_category)
#     if _platform_suppressed_for_talent(lookup_talent, platform):
#         return platform, "", 0.0, f"{platform} suppressed for this talent after manual validation."
#     exp_search = parse_entity_expectations(title_category, title_sub_category, talent=lookup_talent)
#     all_candidates: List[dict] = []
#     seen_links: set = set()

#     queries = build_queries(
#         lookup_talent, platform, domains, search_keywords,
#         title_category, title_sub_category,
#         username_hints=username_hints,
#         wikipedia_context=wikipedia_context,
#     )

#     for query in queries:
#         try:
#             results = serper_search(query, num_results=RESULTS_PER_QUERY)
#             print(f"[QUERY] {platform} | {lookup_talent} | '{query}' -> {len(results)} raw results")
#             for pos, item in enumerate(results, start=1):
#                 raw_link = item.get("link", "")
#                 prof = profile_from_candidate_url(raw_link, platform)
#                 recovered_from_text = False
#                 if not prof:
#                     prof = profile_from_candidate_text(item, platform)
#                     recovered_from_text = bool(prof)
#                 if not prof:
#                     prof = raw_link
#                 if not prof or prof in seen_links:
#                     continue
#                 if not is_valid_profile_url(prof, platform):
#                     continue
#                 seen_links.add(prof)
#                 all_candidates.append({
#                     **item,
#                     "link": prof,
#                     "search_position": pos,
#                     "query": query,
#                     "recovered_from_text": recovered_from_text,
#                 })
#         except Exception as exc:
#             print(f"[WARN] Serper failed '{query}': {exc}")
#             fatal_markers = (
#                 "not enough credits",
#                 "unauthorized",
#                 "invalid api key",
#                 "forbidden",
#                 "quota",
#                 "billing",
#             )
#             if any(marker in str(exc).lower() for marker in fatal_markers):
#                 raise RuntimeError(str(exc)) from exc
#         if len(all_candidates) >= RESULTS_PER_QUERY * 2:
#             break
#         time.sleep(0.2)

#     valid_candidates = [c for c in all_candidates if is_valid_profile_url(c.get("link", ""), platform)]
#     if exp_search.get("expects_brand"):
#         canonical = _brand_canonical_slug(lookup_talent)
#         brand_clean: List[dict] = []
#         for c in valid_candidates:
#             clink = c.get("link", "")
#             path_slug = _path_handle_slug(clink, platform)
#             if _brand_slug_is_vertical(canonical, path_slug, lookup_talent, platform):
#                 print(f"  [BRAND] Skip vertical handle: {(clink or '')[:90]}")
#                 continue
#             rej_brand, why_brand = entity_profile_rejected(
#                 lookup_talent, title_category, title_sub_category, c, platform
#             )
#             if rej_brand:
#                 print(f"  [BRAND] Skip mismatch: {(clink or '')[:90]} | {why_brand}")
#                 continue
#             brand_clean.append(c)
#         valid_candidates = brand_clean
#     valid_candidates = sort_candidates_for_ai(
#         lookup_talent, valid_candidates, search_keywords, title_category, title_sub_category,
#         username_hints=username_hints, platform=platform,
#     )
#     top_candidates = valid_candidates[:MAX_CANDIDATES_FOR_AI]

#     ctx = f" | kw: {search_keywords}" if search_keywords else ""
#     print(f"[INFO] {platform} | {lookup_talent}{ctx} -> {len(top_candidates)} profile-filtered candidates for AI")

#     if not top_candidates:
#         return platform, "", 0.0, "No valid profile/channel URLs in search results."

#     top_candidate = top_candidates[0]
#     fallback      = first_valid_profile_link(top_candidates, platform)

#     try:
#         ai_result  = ai_select_best_profile(
#             lookup_talent, platform, top_candidates,
#             title_category, title_sub_category, search_keywords,
#             username_hints=username_hints,
#             wikipedia_context=wikipedia_context,
#         )
#         selected   = ai_result["best_link"]
#         confidence = ai_result["confidence"]
#         reason     = ai_result["reason"]

#         recovered_pick = _prefer_recovered_athlete_handle(
#             lookup_talent, platform, selected, top_candidates,
#             title_category, title_sub_category, search_keywords,
#         )
#         if recovered_pick:
#             selected = recovered_pick["link"]
#             confidence = max(confidence, 0.88)
#             reason = (
#                 f"{reason} | Preferred @handle recovered from search result text "
#                 "that directly names the athlete."
#             )

#         personal_site_pick = _prefer_personal_website_athlete_profile(
#             lookup_talent, platform, selected, top_candidates,
#             title_category, title_sub_category, search_keywords,
#         )
#         if personal_site_pick:
#             selected = personal_site_pick["link"]
#             confidence = max(confidence, 0.90)
#             reason = (
#                 f"{reason} | Preferred top-ranked athlete profile with "
#                 "a name-matching personal website."
#             )

#         # ── If AI returned empty, try strong-rank fallback ──
#         if not selected and fallback:
#             rej_fb, _ = entity_profile_rejected(lookup_talent, title_category, title_sub_category, top_candidate, platform)
#             if not rej_fb:
#                 selected   = fallback
#                 confidence = min(confidence, 0.42)
#                 reason     = (reason or "") + " | AI empty; using top-ranked candidate."

#         # ── If AI returned invalid URL, try fallback ──
#         elif selected and not is_valid_profile_url(selected, platform):
#             print(f"[WARN] AI returned non-profile URL; trying fallback.")
#             rej_fb, _ = entity_profile_rejected(lookup_talent, title_category, title_sub_category, top_candidate, platform)
#             if fallback and not rej_fb:
#                 selected   = fallback
#                 confidence = min(confidence, 0.42)
#                 reason     = f"{reason} (AI URL invalid; fallback used)"
#             else:
#                 selected   = ""
#                 confidence = 0.0

#         # ── Hard entity check on the selected candidate ──
#         emit_candidate: Optional[dict] = None
#         if selected:
#             sel_norm = normalize_profile_url(selected, platform).rstrip("/")
#             cand     = next(
#                 (c for c in top_candidates
#                  if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == sel_norm),
#                 None,
#             )
#             emit_candidate = cand
#             if cand:
#                 rej, why = entity_profile_rejected(lookup_talent, title_category, title_sub_category, cand, platform)
#                 if rej:
#                     print(f"[REJECT] {platform} | {lookup_talent} | {why}")
#                     selected = ""; confidence = min(confidence, 0.15); reason = why
#                     emit_candidate = None

#         # ── Verification pass (only if confidence is meaningful) ──
#         if selected and confidence >= 0.55:
#             cand_for_verify = emit_candidate or top_candidate
#             verified, verify_conf, verify_rsn = ai_verify_selected_link(
#                 lookup_talent, platform, selected,
#                 cand_for_verify.get("title", "") if cand_for_verify else "",
#                 cand_for_verify.get("snippet", "") if cand_for_verify else "",
#                 title_category, title_sub_category, search_keywords,
#                 wikipedia_context=wikipedia_context,
#             )
#             print(f"[VERIFY] {platform} | {lookup_talent} | verified={verified} conf={verify_conf:.2f} | {verify_rsn}")
#             path_slug = _path_handle_slug(selected, platform)
#             cand_brand = emit_candidate or top_candidate
#             brand_handle_ok = (
#                 exp_search.get("expects_brand")
#                 and (
#                     _brand_handle_matches_row(lookup_talent, path_slug, platform)
#                     or (
#                         platform == "YouTube"
#                         and _is_youtube_channel_id_slug(path_slug)
#                         and cand_brand
#                         and _candidate_supports_brand(lookup_talent, cand_brand)
#                     )
#                 )
#             )
#             if not verified and verify_conf < AI_VERIFY_MIN_CONFIDENCE:
#                 athlete_profile_ok = (
#                     not exp_search.get("expects_brand")
#                     and _athlete_row_active(exp_search)
#                     and cand_for_verify
#                     and _candidate_has_exact_profile_title(lookup_talent, cand_for_verify)
#                     and confidence >= 0.75
#                 )
#                 if brand_handle_ok and confidence >= 0.55:
#                     print(
#                         f"[VERIFY-WAIVER] {platform} | {talent} | "
#                         "keeping brand handle match despite weak snippet verify"
#                     )
#                     reason = f"{reason} [verify waived: {verify_rsn}]"
#                 elif athlete_profile_ok:
#                     print(
#                         f"[VERIFY-WAIVER] {platform} | {lookup_talent} | "
#                         "keeping exact athlete profile-title match despite weak snippet verify"
#                     )
#                     reason = f"{reason} [verify waived: {verify_rsn}]"
#                 else:
#                     print(f"[VETO] Verify vetoed result for {platform} | {lookup_talent}")
#                     selected   = ""
#                     confidence = min(confidence, verify_conf)
#                     reason     = f"Verify pass failed: {verify_rsn}"
#                     emit_candidate = None
#             elif verified and verify_conf > 0:
#                 # Blend confidence: average of selection + verify
#                 confidence = round((confidence + verify_conf) / 2, 4)
#                 reason     = f"{reason} [verify: {verify_rsn}]"

#         # ── Final emission gate ──
#         emit, conf_out, rsn_out = decide_emitted_link(
#             lookup_talent, platform, selected or "", confidence, reason,
#             top_candidate, search_keywords, title_category, title_sub_category,
#             emit_candidate=emit_candidate,
#         )
#         return platform, emit, conf_out, rsn_out

#     except Exception as exc:
#         print(f"[ERROR] AI/Verify failed for {platform} | {lookup_talent}: {exc}")
#         # Last-resort deterministic fallback
#         if fallback:
#             rej_fb, _ = entity_profile_rejected(lookup_talent, title_category, title_sub_category, top_candidate, platform)
#             if not rej_fb:
#                 rs = candidate_rank_score(
#                     lookup_talent, top_candidate, search_keywords, title_category, title_sub_category
#                 )
#                 if rs >= MIN_RANK_SCORE_FOR_FALLBACK and talent_url_aligned(
#                     lookup_talent, fallback, title_category, title_sub_category,
#                 ):
#                     return (
#                         platform,
#                         normalize_profile_url(fallback, platform),
#                         min(0.60, rs / 22.0),
#                         f"AI error fallback (rank={rs:.1f}): {exc}",
#                     )
#         return platform, "", 0.0, f"AI error: {exc}"


# # ─────────────────────────────────────────────
# #  PROCESS ONE ROW
# # ─────────────────────────────────────────────

# def process_row(
#     df: pd.DataFrame,
#     row_label: object,
#     platform_progress: Optional[Callable[[str, str], None]] = None,
# ) -> None:
#     talent            = str(df.at[row_label, "Talent Name"] or "").strip()
#     title_category    = str(df.at[row_label, "title_category"]    or "").strip()
#     title_sub_category= str(df.at[row_label, "title_sub_category"] or "").strip()
#     wikipedia_url     = _clean_wikipedia_url(df.at[row_label, WIKIPEDIA_URL_COLUMN] if WIKIPEDIA_URL_COLUMN in df.columns else "")

#     if not talent:
#         return

#     wikipedia_context = wikipedia_identity_context(wikipedia_url) if wikipedia_url else ""
#     ambiguity = get_name_ambiguity_level(talent)
#     print(f"\n{'='*65}")
#     print(f"Processing: {talent}  [ambiguity={ambiguity}]")
#     if title_category or title_sub_category:
#         print(f"  Category: {title_category} | SubCategory: {title_sub_category}")
#     if wikipedia_context:
#         print(f"  Wikipedia: {wikipedia_context[:180]}")

#     # Resolve platforms one at a time (sequential for username-hint propagation)
#     resolved_links: Dict[str, str] = {}
#     for platform, domains in PLATFORMS.items():
#         if platform_progress:
#             platform_progress(platform, "start")

#         existing = str(df.at[row_label, platform] or "").strip()
#         if existing:
#             resolved_links[platform] = existing
#             if platform_progress:
#                 platform_progress(platform, "done")
#             continue

#         # Build username hints from what's been resolved so far
#         username_hints = extract_username_hints(resolved_links)

#         try:
#             plat_out, link, confidence, reason = search_one_platform(
#                 talent, platform, domains,
#                 title_category, title_sub_category,
#                 username_hints=username_hints,
#                 wikipedia_context=wikipedia_context,
#             )
#         except Exception as exc:
#             print(f"  [{platform}] UNEXPECTED ERROR: {exc}")
#             link, confidence, reason = "", 0.0, str(exc)

#         df.at[row_label, platform] = link
#         ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[platform] = confidence
#         df.at[row_label, PLATFORM_CONF_COLUMNS[platform]] = confidence if link else float("nan")
#         if link:
#             ROW_PLATFORM_SOURCE.setdefault(row_label, {})[platform] = "search"
#             resolved_links[platform] = link

#         print(f"  [{platform}] {link or '(blank)'} (conf={confidence:.2f}) — {reason}")
#         if platform_progress:
#             platform_progress(platform, "done")
#         time.sleep(OPENAI_DELAY_SECONDS)

#     reconcile_athlete_handles_from_confirmed_profiles(
#         df, row_label, resolved_links, title_category, title_sub_category
#     )
#     reconcile_initial_name_x_instagram_handle(
#         df, row_label, resolved_links, title_category, title_sub_category
#     )

#     # Enrich missing platforms from anchor bios / Linktree
#     enrich_row_from_anchor_profiles(df, row_label)
#     _refresh_row_aggregate_confidence(df, row_label)
#     _refresh_row_source_cell(df, row_label)


# # ─────────────────────────────────────────────
# #  API PIPELINE WRAPPERS
# # ─────────────────────────────────────────────

# def _ensure_pipeline_columns(df: pd.DataFrame) -> pd.DataFrame:
#     """Keep uploaded/name-only dataframes compatible with the resolver pipeline."""
#     if WIKIPEDIA_URL_COLUMN not in df.columns:
#         wiki_col = _find_column(
#             df,
#             "wikipedia_url",
#             "Wikipedia",
#             "Wiki URL",
#             "wiki_url",
#             "Wiki",
#             "Wikipedia Link",
#         )
#         df[WIKIPEDIA_URL_COLUMN] = df[wiki_col] if wiki_col else ""
#     for column in ("Talent Name", "title_category", "title_sub_category"):
#         if column not in df.columns:
#             df[column] = ""
#     df[WIKIPEDIA_URL_COLUMN] = df[WIKIPEDIA_URL_COLUMN].apply(_clean_wikipedia_url)
#     for platform in PLATFORMS:
#         if platform not in df.columns:
#             df[platform] = ""
#         conf_column = PLATFORM_CONF_COLUMNS[platform]
#         if conf_column not in df.columns:
#             df[conf_column] = float("nan")
#     if "Confidence" not in df.columns:
#         df["Confidence"] = float("nan")
#     if "Source" not in df.columns:
#         df["Source"] = ""
#     return df


# def run_pipeline_on_dataframe(
#     df: pd.DataFrame,
#     progress: Optional[Callable[[int, int, str], None]] = None,
#     platform_progress: Optional[Callable[[int, str, str], None]] = None,
# ) -> pd.DataFrame:
#     """
#     Run the social lookup pipeline on a prepared dataframe.
#     progress receives (1-based index, total rows, talent name).
#     """
#     df = _ensure_pipeline_columns(df.copy())
#     ROW_PLATFORM_CONFIDENCE.clear()
#     ROW_PLATFORM_SOURCE.clear()
#     ROW_USERNAME_HINTS.clear()

#     total = len(df)
#     print(f"Initialized talent dataframe with {total} row(s).")

#     for i, row_label in enumerate(df.index, start=1):
#         talent = str(df.at[row_label, "Talent Name"] or "").strip()
#         if not talent:
#             continue

#         ROW_PLATFORM_CONFIDENCE[row_label] = {}
#         ROW_PLATFORM_SOURCE[row_label] = {}
#         for platform in PLATFORMS:
#             if str(df.at[row_label, platform] or "").strip():
#                 ROW_PLATFORM_CONFIDENCE[row_label][platform] = 1.0
#                 ROW_PLATFORM_SOURCE[row_label][platform] = "input"
#                 df.at[row_label, PLATFORM_CONF_COLUMNS[platform]] = 1.0

#         if progress:
#             progress(i, total, talent)

#         def _row_platform_progress(platform: str, phase: str) -> None:
#             if platform_progress:
#                 platform_progress(i - 1, platform, phase)

#         process_row(df, row_label, platform_progress=_row_platform_progress)
#         delay = random.uniform(*REQUEST_DELAY_BETWEEN_TALENTS)
#         print(f"  [{i}/{total}] complete — sleeping {delay:.1f}s")
#         time.sleep(delay)

#     return df


# def run_pipeline_for_names(
#     names: List[str],
#     progress: Optional[Callable[[int, int, str], None]] = None,
#     platform_progress: Optional[Callable[[int, str, str], None]] = None,
# ) -> pd.DataFrame:
#     """Build a dataframe from plain names and run the lookup pipeline."""
#     clean = [str(name).strip() for name in names if name and str(name).strip()]
#     if not clean:
#         raise ValueError("At least one non-empty name is required.")
#     df = build_talent_df(clean, list(PLATFORMS.keys()))
#     return run_pipeline_on_dataframe(
#         df,
#         progress=progress,
#         platform_progress=platform_progress,
#     )


# def run_pipeline() -> pd.DataFrame:
#     return run_pipeline_on_dataframe(load_talent_table())


# # ─────────────────────────────────────────────
# #  EXCEL OUTPUT WITH FORMATTING
# # ─────────────────────────────────────────────

# def save_results(df: pd.DataFrame, output_dir: Optional[Path] = None) -> Path:
#     timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
#     base_dir = Path(output_dir) if output_dir is not None else Path(__file__).resolve().parent
#     base_dir.mkdir(parents=True, exist_ok=True)
#     output_path = base_dir / f"Talent_Social_Lookup_{timestamp}.xlsx"

#     try:
#         from openpyxl import load_workbook
#         from openpyxl.styles import PatternFill, Font, Alignment
#         from openpyxl.utils import get_column_letter
#     except ImportError:
#         df.to_excel(output_path, index=False)
#         print(f"\n✅ Saved (no formatting): {output_path}")
#         return output_path

#     df.to_excel(output_path, index=False)
#     wb = load_workbook(output_path)
#     ws = wb.active

#     # Header style
#     header_fill = PatternFill("solid", fgColor="1F4E79")
#     header_font = Font(bold=True, color="FFFFFF", size=10)
#     for cell in ws[1]:
#         cell.fill      = header_fill
#         cell.font      = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     # Identify confidence columns
#     conf_col_indices = {}
#     for col_idx, cell in enumerate(ws[1], start=1):
#         val = str(cell.value or "")
#         for p in PLATFORMS:
#             if val == PLATFORM_CONF_COLUMNS[p]:
#                 conf_col_indices[p] = col_idx

#     # Row colouring
#     low_conf_fill  = PatternFill("solid", fgColor="FFF2CC")  # yellow — risky
#     warn_fill      = PatternFill("solid", fgColor="FCE4D6")  # light orange — first-name-only
#     ok_fill        = PatternFill("solid", fgColor="E2EFDA")  # green — high confidence

#     col_headers = [str(cell.value or "") for cell in ws[1]]

#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#         talent_val = str(row[0].value or "")
#         first_only = is_first_name_only(talent_val)

#         # Find overall confidence value
#         try:
#             conf_idx    = col_headers.index("Confidence")
#             overall_val = row[conf_idx].value
#             overall_conf= float(overall_val) if overall_val not in (None, "") else None
#         except (ValueError, TypeError):
#             overall_conf = None

#         for cell in row:
#             cell.alignment = Alignment(vertical="center", wrap_text=True)
#             if first_only:
#                 cell.fill = warn_fill
#             elif overall_conf is not None:
#                 if overall_conf >= 0.85:
#                     cell.fill = ok_fill
#                 elif overall_conf < MIN_CONFIDENCE_EMIT:
#                     cell.fill = low_conf_fill

#     # Auto-width
#     for col_idx in range(1, ws.max_column + 1):
#         max_len = 0
#         for row_cell in ws.iter_rows(min_col=col_idx, max_col=col_idx):
#             for c in row_cell:
#                 try:
#                     max_len = max(max_len, len(str(c.value or "")))
#                 except Exception:
#                     pass
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

#     # Freeze header row
#     ws.freeze_panes = "A2"

#     wb.save(output_path)
#     print(f"\n✅ Saved: {output_path}")
#     return output_path


# def save_output(df: pd.DataFrame, output_dir: Optional[Path] = None) -> str:
#     """API-compatible output helper used by api_server.py."""
#     return str(save_results(df, output_dir=output_dir))


# # ─────────────────────────────────────────────
# #  MAIN
# # ─────────────────────────────────────────────

# def main() -> None:
#     if not SERPER_API_KEY:
#         print("[ERROR] SERPER_API_KEY not set. Add it to .env")
#         return
#     if not OPENAI_API_KEY:
#         print("[ERROR] OPENAI_API_KEY not set. Add it to .env")
#         return

#     df    = load_talent_table()
#     total = len(df)
#     print(f"Loaded {total} talent row(s).")

#     for idx, row_label in enumerate(df.index):
#         talent = str(df.at[row_label, "Talent Name"] or "").strip()
#         if not talent:
#             continue
#         process_row(df, row_label)
#         delay = random.uniform(*REQUEST_DELAY_BETWEEN_TALENTS)
#         print(f"  [{idx+1}/{total}] complete — sleeping {delay:.1f}s")
#         time.sleep(delay)

#     save_results(df)


# if __name__ == "__main__":
#     main()
import json
import os
import random
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple
from urllib.parse import unquote, urlparse

import pandas as pd
import requests

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# Profile discovery engine (Wikipedia + Serper + LLM verification workflow).
# Imported AFTER load_dotenv so it reads API keys from the environment.
import profile_discovery  # noqa: E402

# ================== API KEYS ==================
SERPER_API_KEY = os.environ.get("SERPER_API_KEY", "").strip()
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_CHAT_MODEL = os.environ.get("OPENAI_CHAT_MODEL", "gpt-4o-mini")

TEST_BRANDS_PATH = Path(__file__).resolve().parent / "Demo_Social.xlsx"
WIKIPEDIA_URL_COLUMN = "Wikipedia URL"

# ── NEW: Instagram URL column name as provided by client ──
INSTAGRAM_INPUT_COLUMN = "Instagram URL"

# ================== INPUT FALLBACK ==================
talent_names = [
    "Britney Vest",
    "Ari Melber",
    "Alyssa Anderson",
    "Andrea",
    "Anastasia Pagonis",
]

# ================== CONFIG ==================
RESULTS_PER_QUERY             = 10
MAX_CANDIDATES_FOR_AI         = 6
MAX_WORKERS                   = 3
REQUEST_DELAY_BETWEEN_TALENTS = (1.0, 2.0)
OPENAI_DELAY_SECONDS          = 0.4

MIN_CONFIDENCE_EMIT           = float(os.environ.get("MIN_CONFIDENCE_EMIT", "0.72"))
ANCHOR_MIN_CONFIDENCE         = float(os.environ.get("ANCHOR_MIN_CONFIDENCE", "0.86"))
# ── Lowered from 0.86 → 0.80 for early enrichment phase ──
EARLY_ENRICH_MIN_CONFIDENCE   = float(os.environ.get("EARLY_ENRICH_MIN_CONFIDENCE", "0.80"))
MIN_RANK_SCORE_FOR_FALLBACK   = float(os.environ.get("MIN_RANK_SCORE_FOR_FALLBACK", "12.0"))
AI_VERIFY_MIN_CONFIDENCE      = float(os.environ.get("AI_VERIFY_MIN_CONFIDENCE", "0.62"))

# Step 6 of the discovery workflow — configurable LLM match threshold.
PROFILE_MATCH_THRESHOLD       = profile_discovery.PROFILE_MATCH_THRESHOLD

# ── NEW: enriched links derived from bio must exceed this to be trusted ──
ENRICH_EMIT_MIN_CONFIDENCE    = float(os.environ.get("ENRICH_EMIT_MIN_CONFIDENCE", "0.75"))

FETCH_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# Per-row per-platform confidence (filled in process_row)
ROW_PLATFORM_CONFIDENCE: Dict[object, Dict[str, float]] = {}
# Per-row per-platform provenance
ROW_PLATFORM_SOURCE: Dict[object, Dict[str, str]] = {}
# Per-row resolved username hints (platform → handle string)
ROW_USERNAME_HINTS: Dict[object, Dict[str, str]] = {}
WIKIPEDIA_CONTEXT_CACHE: Dict[str, str] = {}

# ── NEW: cache of early-enriched links per row_label (platform → url) ──
# Populated during Phase-1 enrichment, consumed as extra candidates in AI step
ROW_EARLY_ENRICH_LINKS: Dict[object, Dict[str, str]] = {}

PLATFORMS: Dict[str, List[str]] = {
    "Facebook":  ["facebook.com"],
    "Instagram": ["instagram.com"],
    "X":         ["x.com", "twitter.com"],
    "TikTok":    ["tiktok.com"],
    "YouTube":   ["youtube.com"],
}

PLATFORM_CONF_COLUMNS: Dict[str, str] = {p: f"{p} Confidence" for p in PLATFORMS}

# X/Twitter path segments that are site features — NOT user handles.
_X_NON_PROFILE_HANDLES = frozenset({
    "home", "explore", "search", "notifications", "messages", "settings",
    "login", "logout", "signup", "signin", "sign_in", "register",
    "intent", "share", "compose", "post", "i", "pic", "photo", "photos",
    "hashtag", "lists", "topics", "who_to_follow", "account", "accounts",
    "analytics", "analytics_pixel", "pixel", "ads", "business", "about",
    "privacy", "tos", "help", "support", "legal", "oauth", "deck",
    "premium", "verified", "moments", "grok", "communities", "bookmarks",
    "status", "followers", "following", "likes", "media", "widgets",
    "embed", "oauth", "redirect", "sessions", "tos", "rules", "safety",
})


def _x_handle_from_url(link: str) -> str:
    """First path segment for x.com / twitter.com URLs."""
    try:
        path = (urlparse(link).path or "").strip("/")
    except Exception:
        return ""
    segs = [s for s in path.split("/") if s]
    return segs[0].lstrip("@") if len(segs) == 1 else ""


def _is_valid_x_profile_handle(handle: str) -> bool:
    if not handle:
        return False
    h = handle.lower().lstrip("@")
    if h in _X_NON_PROFILE_HANDLES:
        return False
    if h.startswith("analytics_") or h.endswith("_pixel"):
        return False
    if "." in h and h.split(".")[0] in _X_NON_PROFILE_HANDLES:
        return False
    return True


# ─────────────────────────────────────────────
#  UTILITY HELPERS
# ─────────────────────────────────────────────

def is_first_name_only(talent: str) -> bool:
    parts = re.sub(r"\s+", " ", (talent or "").strip()).split()
    return len(parts) == 1 and bool(parts[0])


def _find_column(raw: pd.DataFrame, *candidates: str) -> Optional[str]:
    cmap = {str(c).strip().lower(): c for c in raw.columns}
    for cand in candidates:
        if cand.lower() in cmap:
            return cmap[cand.lower()]
    return None


def _slug_chars(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


_KNOWN_PLATFORM_SUPPRESSIONS: Dict[str, set] = {
    "bjpowell": {"Facebook", "TikTok", "YouTube"},
}


def _platform_suppressed_for_talent(talent: str, platform: str) -> bool:
    name = re.sub(r"\s+[-–—|]\s+[A-Z0-9]{2,8}$", "", (talent or "").strip())
    return platform in _KNOWN_PLATFORM_SUPPRESSIONS.get(_slug_chars(name), set())


def _clean_wikipedia_url(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    url = str(value).strip()
    if not url or url.lower() == "nan":
        return ""
    if "wikipedia.org/wiki/" not in url.lower():
        return ""
    if url.startswith("//"):
        url = "https:" + url
    elif not re.match(r"^https?://", url, re.I):
        url = "https://" + url
    return url.split("#", 1)[0].strip()


def _clean_instagram_input_url(value: object) -> str:
    """
    Sanitise a client-provided Instagram URL.
    Accepts full URLs, bare handles (@handle or handle), profile paths.
    Returns a normalised https://www.instagram.com/<handle> or "".
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    raw = str(value).strip()
    if not raw or raw.lower() == "nan":
        return ""
    # Already a full URL
    if "instagram.com" in raw.lower():
        # Strip down to profile root
        m = re.search(r"instagram\.com/([A-Za-z0-9._]{1,30})", raw, re.I)
        if m:
            handle = m.group(1).strip("/")
            url = f"https://www.instagram.com/{handle}"
            if is_valid_profile_url(url, "Instagram"):
                return url
        return ""
    # Bare @handle or handle
    handle = raw.lstrip("@").strip("/")
    if re.match(r"^[A-Za-z0-9._]{1,30}$", handle):
        url = f"https://www.instagram.com/{handle}"
        if is_valid_profile_url(url, "Instagram"):
            return url
    return ""


def _wikipedia_title_from_url(wikipedia_url: str) -> str:
    try:
        parsed = urlparse(wikipedia_url)
    except Exception:
        return ""
    if "wikipedia.org" not in (parsed.netloc or "").lower():
        return ""
    path = parsed.path or ""
    if "/wiki/" not in path:
        return ""
    title = path.split("/wiki/", 1)[1].strip("/")
    return re.sub(r"\s+", " ", unquote(title).replace("_", " ")).strip()


def wikipedia_identity_context(wikipedia_url: str) -> str:
    """Small optional identity anchor from a provided Wikipedia URL."""
    url = _clean_wikipedia_url(wikipedia_url)
    if not url:
        return ""
    if url in WIKIPEDIA_CONTEXT_CACHE:
        return WIKIPEDIA_CONTEXT_CACHE[url]

    title = _wikipedia_title_from_url(url)
    context_parts: List[str] = []
    if title:
        context_parts.append(f"Wikipedia title: {title}")

    try:
        parsed = urlparse(url)
        page_title = (parsed.path or "").split("/wiki/", 1)[1].strip("/")
        summary_url = f"{parsed.scheme}://{parsed.netloc}/api/rest_v1/page/summary/{page_title}"
        res = requests.get(summary_url, headers=FETCH_HEADERS, timeout=8)
        if res.ok:
            payload = res.json()
            api_title   = str(payload.get("title") or "").strip()
            description = str(payload.get("description") or "").strip()
            extract     = str(payload.get("extract") or "").strip()
            if api_title and api_title.lower() != (title or "").lower():
                context_parts.insert(0, f"Wikipedia title: {api_title}")
            if description:
                context_parts.append(f"Description: {description}")
            if extract:
                context_parts.append(f"Summary: {extract[:450]}")
    except Exception as exc:
        print(f"[WARN] Wikipedia context fetch failed: {exc}")

    context = " | ".join(context_parts)[:900]
    WIKIPEDIA_CONTEXT_CACHE[url] = context
    return context


def _wikipedia_title_from_context(wikipedia_context: str) -> str:
    if not wikipedia_context:
        return ""
    m = re.search(r"Wikipedia title:\s*([^|]+)", wikipedia_context)
    return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""


def _identity_name_aliases(
    talent: str,
    title_category: str = "",
    title_sub_category: str = "",
    wikipedia_context: str = "",
) -> List[str]:
    base = re.sub(r"\s+", " ", (talent or "").strip())
    aliases: List[str] = []
    seen: set = set()

    def add(name: str) -> None:
        name = re.sub(r"\s+", " ", (name or "").strip())
        key = name.lower()
        if name and key not in seen:
            seen.add(key)
            aliases.append(name)

    add(base)
    wiki_title = _wikipedia_title_from_context(wikipedia_context)
    if wiki_title:
        add(wiki_title)

    exp = parse_entity_expectations(title_category, title_sub_category, talent=base)
    if not exp.get("expects_brand"):
        parts = _brand_name_parts(base)
        if len(parts) >= 3:
            add(" ".join(parts[:2]))
            add(f"{parts[0]} {parts[-1]}")
    return aliases


def _identity_handle_aliases(
    talent: str,
    title_category: str = "",
    title_sub_category: str = "",
    wikipedia_context: str = "",
) -> List[str]:
    handles: List[str] = []
    seen: set = set()
    for alias in _identity_name_aliases(talent, title_category, title_sub_category, wikipedia_context):
        slug = _slug_chars(alias)
        if len(slug) >= 4 and slug not in seen:
            seen.add(slug)
            handles.append(slug)
    return handles


# ─────────────────────────────────────────────
#  BRAND HELPERS  (unchanged)
# ─────────────────────────────────────────────

_BRAND_WORD_ABBREV: Dict[str, str] = {
    "basketball": "bball",
    "football": "fball",
}

_BRAND_OFFICIAL_SUFFIXES = frozenset({
    "hq", "official", "sports", "sport", "canada",
})

_BRAND_VERTICAL_SUFFIXES = frozenset({
    "football", "basketball", "cbb", "wcbb", "cfb", "nba", "nfl", "mlb", "nhl", "mls",
    "soccer", "college", "fantasy", "gaming", "news", "sport", "sports", "golf", "tennis",
    "racing", "mma", "wwe", "esports", "highlights", "podcast", "radio", "tv", "plus",
    "women", "womens", "men", "kids", "pr", "shop", "store", "fan", "fans",
})

_KNOWN_SINGLE_WORD_BRANDS = frozenset({
    "overtime", "slam",
})

_KNOWN_BRAND_PLATFORM_HANDLES: Dict[str, Dict[str, Tuple[str, ...]]] = {
    "espn": {
        "Facebook": ("ESPN",),
        "Instagram": ("espn",),
        "X": ("espn",),
        "TikTok": ("espn",),
        "YouTube": ("espn",),
    },
    "foxsports1": {
        "Instagram": ("fs1",),
        "X": ("FS1",),
    },
    "on3sports": {
        "Instagram": ("on3",),
        "X": ("On3sports",),
    },
    "overtime": {
        "Facebook": ("getovertime",),
    },
    "overtimeelite": {
        "Instagram": ("ote",),
        "YouTube": ("OvertimeElite", "ote"),
    },
    "slam": {
        "Facebook": ("SLAM",),
        "Instagram": ("slam",),
        "X": ("SLAMonline",),
        "YouTube": ("SLAM",),
    },
    "slamhs": {
        "Facebook": ("SLAMHS", "SLAMHighSchool"),
        "Instagram": ("slam_hs",),
        "X": ("SLAM_HS",),
        "YouTube": ("SLAMHS", "SLAMHighSchool"),
    },
}


def _brand_canonical_slug(talent: str) -> str:
    return _slug_chars(talent)


def _brand_name_parts(talent: str) -> List[str]:
    return re.findall(r"[a-zA-Z0-9]+", talent or "")


def _talent_name_implies_brand(talent: str) -> bool:
    t = (talent or "").strip()
    if not t:
        return False
    parts = _brand_name_parts(t)
    slug = _slug_chars(t)
    if slug in _KNOWN_BRAND_PLATFORM_HANDLES or slug in _KNOWN_SINGLE_WORD_BRANDS:
        return True
    if is_first_name_only(t):
        return False
    if len(parts) < 2:
        return False
    org_markers = (
        "sports", "sport", "network", "media", "elite", "athletic", "tribune",
        "report", "overtime", "yahoo", "bleacher", "espn", "tsn", "sportsnet",
        "publication", "broadcast", "channel", "league",
    )
    if any(m in slug for m in org_markers):
        return True
    if re.search(r"\([A-Z]{2,}\)", talent):
        return True
    if len(parts) >= 3 and parts[0].lower() == "the":
        return True
    return False


def _brand_explicit_acronyms(talent: str) -> List[str]:
    found: List[str] = []
    seen: set = set()
    for m in re.finditer(r"\(([A-Za-z0-9]{2,})\)", talent or ""):
        ac = m.group(1).lower()
        if ac not in seen:
            seen.add(ac)
            found.append(ac)
    for m in re.finditer(r"\b[A-Z]{2,}\b", talent or ""):
        ac = m.group(0).lower()
        if ac not in seen:
            seen.add(ac)
            found.append(ac)
    return found


def _brand_short_acronyms(parts: List[str]) -> List[str]:
    if len(parts) < 2:
        return []
    meaningful = [p for p in parts if p.lower() not in ("the", "and", "of", "a")]
    if len(meaningful) < 2:
        meaningful = parts
    first, last = meaningful[0], meaningful[-1]
    acronyms: List[str] = []
    if first and last:
        acronyms.append("".join(p[0].lower() for p in meaningful if p))
        if len(last) >= 2:
            acronyms.append(first[0].lower() + last[:2].lower())
        if len(last) >= 3:
            acronyms.append(first[0].lower() + last[-2:].lower())
    return acronyms


def _brand_parent_slug(talent: str) -> str:
    parts = _brand_name_parts(talent)
    if parts and parts[-1].isdigit():
        return _slug_chars("".join(parts[:-1]))
    return ""


def _normal_handle_text(handle: str) -> str:
    return (handle or "").strip().lstrip("@").rstrip("/").lower()


def _brand_platform_handles(talent: str, platform: str = "") -> List[str]:
    aliases = _KNOWN_BRAND_PLATFORM_HANDLES.get(_brand_canonical_slug(talent), {})
    ordered: List[str] = []
    seen: set = set()

    def add_many(values: Tuple[str, ...]) -> None:
        for value in values:
            key = _normal_handle_text(value)
            if key and key not in seen:
                seen.add(key)
                ordered.append(value)

    if platform:
        add_many(aliases.get(platform, ()))
    add_many(aliases.get("default", ()))
    if not platform:
        for plat_values in aliases.values():
            add_many(plat_values)
    return ordered


def _path_handle_text(link: str, platform: str) -> str:
    segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
    if not segs:
        return ""
    if platform == "YouTube" and segs[0].lower() in ("user", "channel", "c") and len(segs) >= 2:
        return segs[1].lstrip("@")
    return segs[-1].lstrip("@")


def _brand_has_strict_platform_handles(talent: str, platform: str) -> bool:
    return bool(_brand_platform_handles(talent, platform))


def _brand_search_handles(talent: str) -> List[str]:
    handles: List[str] = []
    seen: set = set()

    def add(h: str) -> None:
        h = (h or "").strip().lstrip("@").lower()
        if h and h not in seen:
            seen.add(h)
            handles.append(h)

    parts = _brand_name_parts(talent)
    add(_brand_canonical_slug(talent))
    for handle in _brand_platform_handles(talent):
        add(handle)
    for ac in _brand_explicit_acronyms(talent):
        add(ac)
        add(f"{ac}_sports")
        add(f"{ac}sports")
        add(f"{ac}_canada")
        add(f"{ac}canada")
    if parts:
        add("".join(p.lower() for p in parts))
        add("_".join(p.lower() for p in parts))
        if len(parts) >= 2:
            add(parts[0].lower() + "_" + "".join(p.lower() for p in parts[1:]))
        if parts[0].lower() == "the" and len(parts) > 1:
            add("".join(p.lower() for p in parts[1:]))
            add("_".join(p.lower() for p in parts))
            add(parts[0].lower() + "_" + "".join(p.lower() for p in parts[1:]))
            add("_".join(p.lower() for p in parts[1:]))
            add("".join(p[0].lower() for p in parts[1:] if p))
        if len(parts) >= 2:
            add("".join(p[0].lower() for p in parts if p))
        for ac in _brand_short_acronyms(parts):
            add(ac)
        if parts[-1].isdigit():
            add("".join(p[0].lower() for p in parts if p))
        if len(parts[0]) <= 4 and re.match(r"^[a-zA-Z0-9]+$", parts[0]):
            add(parts[0].lower())
        if re.search(r"\d", "".join(parts)):
            add("".join(p[0].lower() for p in parts if p and p[0].isalnum()))
        abbrev_join = []
        for p in parts:
            pl = p.lower()
            if pl in _BRAND_WORD_ABBREV:
                abbrev_join.append(_BRAND_WORD_ABBREV[pl])
            elif pl != "the":
                abbrev_join.append(p.lower())
        if abbrev_join:
            add("".join(abbrev_join))
            if parts and parts[0].lower() == "the":
                add("the" + "".join(abbrev_join))
    for h in list(handles):
        if 4 <= len(h) <= 18:
            add(f"{h}hq")
            add(f"{h}_hq")
        if 2 <= len(h) <= 14:
            add(f"{h}_official")
    return handles[:18]


def _brand_known_handle_slugs(talent: str) -> set:
    return {_slug_chars(h) for h in _brand_search_handles(talent)}


def _brand_handle_rank(talent: str, handle_slug: str) -> int:
    handle_slug = _slug_chars(handle_slug)
    for idx, handle in enumerate(_brand_search_handles(talent)):
        if _slug_chars(handle) == handle_slug:
            return idx
    return 999


def _platform_from_link(link: str) -> str:
    low = (link or "").lower()
    for plat, domains in PLATFORMS.items():
        if any(d in low for d in domains):
            return plat
    return ""


def _is_youtube_channel_id_slug(slug: str) -> bool:
    return bool(re.match(r"^UC[\w-]{10,}$", slug or "", re.I))


def _candidate_supports_brand(talent: str, candidate: dict) -> bool:
    title   = (candidate.get("title") or "")
    snippet = (candidate.get("snippet") or "")
    blob    = f"{title} {snippet}".lower()
    slug    = _slug_chars(talent)
    if slug and len(slug) >= 5 and slug[:min(10, len(slug))] in _slug_chars(blob):
        return True
    for part in _brand_name_parts(talent):
        pl = part.lower()
        if pl in ("the", "and", "of", "a") or len(pl) < 3:
            continue
        if pl in blob:
            return True
    for ac in _brand_explicit_acronyms(talent):
        if ac in blob:
            return True
    for ac in _brand_short_acronyms(_brand_name_parts(talent)):
        if len(ac) >= 3 and re.search(rf"\b{re.escape(ac.lower())}\b", blob):
            return True
    for handle in _brand_platform_handles(talent):
        h = _normal_handle_text(handle)
        if len(h) >= 3 and re.search(rf"\b{re.escape(h)}\b", blob):
            return True
    return False


def _path_handle_slug(link: str, platform: str) -> str:
    segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
    if not segs:
        return ""
    if platform == "YouTube" and segs[0].lower() in ("user", "channel", "c") and len(segs) >= 2:
        return _slug_chars(segs[1].lstrip("@"))
    return _slug_chars(segs[-1].lstrip("@"))


def _brand_handle_matches_row(talent: str, handle_slug: str, platform: str = "") -> bool:
    handle_slug = _slug_chars(handle_slug)
    if not handle_slug:
        return False
    if handle_slug in _brand_known_handle_slugs(talent):
        return True
    canonical = _brand_canonical_slug(talent)
    if handle_slug == canonical:
        return True
    if canonical and handle_slug.startswith(canonical) and len(handle_slug) > len(canonical):
        extra = handle_slug[len(canonical):]
        if extra in _BRAND_OFFICIAL_SUFFIXES:
            return True
    parent = _brand_parent_slug(talent)
    if parent and handle_slug == parent:
        return platform == "YouTube"
    first_tokens  = re.findall(r"[a-z0-9]+", (talent or "").lower())
    explicit      = _brand_explicit_acronyms(talent)
    for ac in explicit:
        if handle_slug == ac:
            return True
        if handle_slug.startswith(ac) and handle_slug[len(ac):] in _BRAND_OFFICIAL_SUFFIXES:
            return True
    if (
        first_tokens
        and handle_slug == _slug_chars(first_tokens[0])
        and re.search(r"\d", first_tokens[0])
    ):
        return True
    significant = [t for t in first_tokens if len(t) > 1 or t.isdigit()]
    if not significant or not all(t in handle_slug for t in significant):
        return False
    if canonical and handle_slug.startswith(canonical) and len(handle_slug) > len(canonical):
        extra = handle_slug[len(canonical):]
        if extra and extra not in _slug_chars(talent) and (
            extra in _BRAND_VERTICAL_SUFFIXES or len(extra) >= 2
        ):
            return False
    return True


def _brand_slug_is_vertical(canonical: str, path_slug: str, talent: str, platform: str = "") -> bool:
    path_slug = _slug_chars(path_slug)
    canonical = _slug_chars(canonical)
    if not path_slug:
        return False
    if _brand_handle_matches_row(talent, path_slug, platform):
        return False
    parent = _brand_parent_slug(talent)
    if parent and path_slug == parent and platform != "YouTube":
        return True
    if canonical and path_slug != canonical and canonical.startswith(path_slug) and len(canonical) > len(path_slug) + 1:
        return True
    if canonical and path_slug.startswith(canonical) and len(path_slug) > len(canonical):
        remainder = path_slug[len(canonical):]
        if remainder in _BRAND_VERTICAL_SUFFIXES or len(remainder) >= 2:
            return True
    return False


def profile_from_candidate_url(link: str, platform: str) -> str:
    if not link:
        return ""
    if is_valid_profile_url(link, platform):
        return normalize_profile_url(link, platform)
    low = link.lower()
    if platform == "Facebook" and "facebook.com" in low:
        m = re.search(r"facebook\.com/([^/?#]+)", low, re.I)
        if m:
            handle = m.group(1).lower()
            blocked = {
                "share", "sharer", "groups", "events", "marketplace", "gaming", "watch",
                "people", "pages", "profile.php", "public", "login",
            }
            if handle not in blocked:
                base = f"https://www.facebook.com/{m.group(1)}"
                if is_valid_profile_url(base, platform):
                    return normalize_profile_url(base, platform)
    if platform == "X" and ("x.com" in low or "twitter.com" in low):
        if "/user/status/" in low:
            return ""
        m = re.search(r"(?:x|twitter)\.com/([^/?#]+)/?", low, re.I)
        if m:
            handle = m.group(1)
            if _is_valid_x_profile_handle(handle):
                base = f"https://x.com/{handle}"
                if is_valid_profile_url(base, platform):
                    return normalize_profile_url(base, platform)
    if platform == "TikTok" and "tiktok.com" in low:
        m = re.search(r"tiktok\.com/@([^/?#]+)", low, re.I)
        if m:
            base = f"https://www.tiktok.com/@{m.group(1)}"
            if is_valid_profile_url(base, platform):
                return normalize_profile_url(base, platform)
    if platform == "YouTube" and "youtube.com" in low:
        for pattern in (
            r"youtube\.com/(@[\w.-]+)",
            r"youtube\.com/user/([\w.-]+)",
            r"youtube\.com/channel/([\w.-]+)",
            r"youtube\.com/c/([\w.-]+)",
        ):
            m = re.search(pattern, low, re.I)
            if m:
                if "/user/" in m.group(0).lower():
                    seg = f"user/{m.group(1)}"
                elif "/channel/" in m.group(0).lower():
                    seg = f"channel/{m.group(1)}"
                elif "/c/" in m.group(0).lower():
                    seg = f"c/{m.group(1)}"
                else:
                    seg = m.group(1)
                base = f"https://www.youtube.com/{seg}"
                if is_valid_profile_url(base, platform):
                    return normalize_profile_url(base, platform)
    return ""


def profile_from_candidate_text(candidate: dict, platform: str) -> str:
    title   = candidate.get("title", "") or ""
    snippet = candidate.get("snippet", "") or ""
    text    = f"{title} {snippet}"

    handles: List[str] = []
    for m in re.finditer(r"\(@([A-Za-z0-9_.-]{2,30})\)", text):
        handles.append(m.group(1))
    for m in re.finditer(r"\b@([A-Za-z0-9_.-]{2,30})\b", text):
        handles.append(m.group(1))

    blocked = {
        "instagram", "facebook", "twitter", "youtube", "tiktok", "espn",
        "nba", "ncaabasketball", "sportscenter",
    }
    blocked |= _X_NON_PROFILE_HANDLES
    for handle in handles:
        h = handle.strip().strip(".")
        if not h or h.lower() in blocked:
            continue
        if platform == "X":
            if not _is_valid_x_profile_handle(h):
                continue
            url = f"https://x.com/{h}"
        elif platform == "Instagram":
            url = f"https://www.instagram.com/{h}"
        elif platform == "TikTok":
            url = f"https://www.tiktok.com/@{h}"
        else:
            continue
        if is_valid_profile_url(url, platform):
            return normalize_profile_url(url, platform)
    return ""


# ─────────────────────────────────────────────
#  NAME AMBIGUITY
# ─────────────────────────────────────────────

_COMMON_FIRST_NAMES = {
    "andrea", "jessica", "jennifer", "ashley", "brittany", "britney",
    "samantha", "amanda", "sarah", "emily", "emma", "olivia", "megan",
    "michael", "james", "john", "david", "robert", "william", "daniel",
    "matthew", "chris", "jason", "kevin", "ryan", "brian", "tyler",
    "alex", "jordan", "taylor", "morgan", "charlie", "casey", "drew",
}


def get_name_ambiguity_level(talent: str) -> str:
    parts = re.sub(r"\s+", " ", (talent or "").strip()).lower().split()
    if len(parts) == 1:
        return "high"
    if len(parts) == 2:
        first = parts[0]
        if first in _COMMON_FIRST_NAMES:
            return "medium"
    return "low"


def _effective_min_confidence(talent: str) -> float:
    level = get_name_ambiguity_level(talent)
    if level == "high":
        return min(0.92, MIN_CONFIDENCE_EMIT + 0.10)
    if level == "medium":
        return min(0.85, MIN_CONFIDENCE_EMIT + 0.05)
    return MIN_CONFIDENCE_EMIT


# ─────────────────────────────────────────────
#  METADATA HELPERS
# ─────────────────────────────────────────────

def extract_search_keywords(title_category: str, title_sub_category: str) -> str:
    parts: List[str] = []
    for raw in (title_category, title_sub_category):
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        s = str(raw).strip()
        if not s or s.lower() == "nan":
            continue
        parts.append(s)
    if not parts:
        return ""
    text = " ".join(parts)
    text = text.replace(",", " ").replace("|", " ")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(
        r"(?i)\b(talent type|gender|talent subtype|publication type)\s*-\s*",
        " ", text,
    )
    text = re.sub(r"\s+", " ", text).strip()
    words = text.split()
    return " ".join(words[:14])[:140].strip()


def parse_entity_expectations(
    title_category: str,
    title_sub_category: str,
    talent: str = "",
) -> Dict[str, bool]:
    blob = f"{title_category or ''} {title_sub_category or ''}".lower()
    cat  = (title_category or "").lower()
    expects_brand = bool(
        re.search(r"\bpublishers?\b", cat)
        or re.search(r"\b(publication|network|brand|company|organization|organisation)\b", blob)
        or re.search(r"\btv\s*network\b", cat)
        or re.search(r"\bmedia\s*(brand|company|outlet)\b", blob)
        or _talent_name_implies_brand(talent)
    )
    return {
        "expects_brand":      expects_brand,
        "expects_male":       bool(re.search(r"gender\s*-\s*man\b", blob)),
        "expects_female":     bool(re.search(r"gender\s*-\s*woman\b", blob)),
        "expects_athlete":    False if expects_brand else bool(
            re.search(r"\bathlete\b", blob)
            or re.search(r"\bbasketball\b", blob)
            or re.search(r"\bbaseball\b", blob)
            or re.search(r"\bfootball\b", blob)
            or ("sport" in blob and not re.search(r"publication\s*type", blob))
        ),
        "expects_basketball": False if expects_brand else bool(re.search(r"\bbasketball\b", blob)),
        "expects_baseball":   False if expects_brand else bool(re.search(r"\bbaseball\b", blob)),
        "expects_musician":   bool(
            re.search(r"\bmusician\b|\bsinger\b|\brap(per)?\b|\bartist\b|\bband\b", blob)
        ),
        "expects_journalist": bool(
            re.search(r"\bjournalist\b|\bnews\b|\banchor\b|\breporter\b|\bhost\b|\bmedia\b", blob)
        ),
        "expects_executive":  bool(
            re.search(r"\bceo\b|\bcfo\b|\bcoo\b|\bexecutive\b|\bfounder\b|\bbusiness\b", blob)
        ),
        "expects_actor":      bool(
            re.search(r"\bactor\b|\bactress\b|\bfilm\b|\btelevision\b|\btv star\b", blob)
        ),
        "expects_politician": bool(
            re.search(r"\bpolitician\b|\bsenator\b|\bcongressman\b|\bgovernor\b|\bpresident\b", blob)
        ),
    }


def _metadata_tokens(search_keywords: str) -> List[str]:
    if not search_keywords:
        return []
    parts = re.split(r"[^\w]+", search_keywords.lower())
    stop = {"the", "and", "for", "type", "talent", "gender", "subtype", "publication", "network", "man", "woman"}
    return [p for p in parts if p and p not in stop and len(p) > 2]


def _talent_lookup_name(
    talent: str,
    title_category: str = "",
    title_sub_category: str = "",
) -> str:
    name = re.sub(r"\s+", " ", (talent or "").strip())
    if not name:
        return ""
    exp = parse_entity_expectations(title_category, title_sub_category, talent=name)
    if exp.get("expects_brand"):
        return name
    cleaned = re.sub(r"\s+[-–—|]\s+[A-Z0-9]{2,8}$", "", name).strip()
    return cleaned or name


def _athlete_handle_aliases(talent: str) -> List[str]:
    parts = [p.lower() for p in _brand_name_parts(talent)]
    if len(parts) < 2:
        return []
    first = _slug_chars(parts[0])
    last  = _slug_chars(parts[-1])
    if not first or not last:
        return []
    aliases: List[str] = []
    seen: set = set()

    def add(value: str) -> None:
        if value and value not in seen and len(value) >= 4:
            seen.add(value)
            aliases.append(value)

    add(first + last)
    add(f"{first}{last}page")
    add(f"{first}{last}1")
    add(first[:1] + last)
    add(f"{first[:1]}{last}1")
    if len(first) >= 2:
        add(first[:2] + last)
    # Compact handles common on X/IG (e.g. Jake Thompson -> jthomp15)
    if len(last) >= 5:
        add(first[:1] + last[:5])
        add(first[:1] + last[:5] + "15")
    if len(last) >= 4:
        add(first[:1] + last[:4])
    if len(first) >= 2 and len(last) >= 4:
        add(first[:2] + last[:4])
    return aliases[:12]


def _expected_sports_from_metadata(
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
    exp: Dict[str, bool],
) -> set:
    """Sports implied by row metadata (used to reject wrong-sport bios)."""
    blob = f"{title_category or ''} {title_sub_category or ''} {search_keywords or ''}".lower()
    expected: set = set()
    if exp.get("expects_basketball") or re.search(r"\bbasketball\b", blob):
        expected.add("basketball")
    if exp.get("expects_baseball") or re.search(r"\bbaseball\b", blob):
        expected.add("baseball")
    if re.search(r"\b(golf|golfer|pga)\b", blob) and not (
        exp.get("expects_baseball") or exp.get("expects_basketball")
    ):
        expected.add("golf")
    if re.search(r"\bfootball\b", blob) and not re.search(r"publication\s*type", blob):
        expected.add("football")
    if re.search(r"\b(baseball|mlb)\b", blob):
        expected.add("baseball")
    if re.search(r"\b(soccer|mls|fifa)\b", blob):
        expected.add("soccer")
    if re.search(r"\b(tennis|atp|wta)\b", blob):
        expected.add("tennis")
    if re.search(r"\b(hockey|nhl)\b", blob):
        expected.add("hockey")
    return expected


def _declared_profile_sports(blob: str) -> set:
    """Sports explicitly claimed in Serper title/snippet (profile bio text)."""
    b = (blob or "").lower()
    declared: set = set()
    if re.search(r"\b(golfer|golfing)\b", b) or re.search(r"\bgolf\b", b) or "handicap" in b or "club golf" in b:
        declared.add("golf")
    if re.search(
        r"\b(basketball|hoops|nba|wnba|ncaa|point guard|shooting guard|power forward|small forward|center)\b",
        b,
    ):
        declared.add("basketball")
    if re.search(r"\b(football|quarterback|wide receiver|tight end|nfl|cfb)\b", b):
        declared.add("football")
    if re.search(r"\b(baseball|pitcher|mlb)\b", b):
        declared.add("baseball")
    if re.search(r"\b(soccer|mls|fifa|midfielder|striker)\b", b):
        declared.add("soccer")
    if re.search(r"\b(tennis|atp|wta)\b", b):
        declared.add("tennis")
    if re.search(r"\b(hockey|nhl|hockey player)\b", b):
        declared.add("hockey")
    return declared


def _athlete_sport_mismatch(
    talent: str,
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
    candidate: dict,
    exp: Dict[str, bool],
) -> Tuple[bool, str]:
    if exp.get("expects_brand"):
        return False, ""
    if not (
        exp.get("expects_athlete")
        or exp.get("expects_basketball")
        or exp.get("expects_baseball")
    ):
        return False, ""
    expected = _expected_sports_from_metadata(
        title_category, title_sub_category, search_keywords, exp
    )
    if not expected:
        return False, ""
    blob = f"{candidate.get('title') or ''} {candidate.get('snippet') or ''}"
    declared = _declared_profile_sports(blob)
    if not declared:
        return False, ""
    if expected and (declared & expected):
        return False, ""
    if expected:
        return (
            True,
            f"Profile bio indicates {', '.join(sorted(declared))}; "
            f"metadata expects {', '.join(sorted(expected))}.",
        )
    # Row metadata has no sport, but bio loudly claims a single other sport (e.g. golfer).
    if exp.get("expects_athlete") and declared == {"golf"}:
        return True, "Athlete row: profile bio declares golf only (likely a different person with the same name)."
    return False, ""


def _is_fullname_plus_digits_handle(talent: str, handle_slug: str) -> bool:
    """e.g. JakeThompson141 — often a different person with the same display name."""
    parts = [p.lower() for p in _brand_name_parts(talent)]
    if len(parts) < 2 or not handle_slug:
        return False
    full = _slug_chars(parts[0] + parts[-1])
    if len(full) < 8 or not handle_slug.startswith(full):
        return False
    suffix = handle_slug[len(full):]
    return len(suffix) >= 1 and suffix.isdigit()


def _compact_athlete_handle_slugs(talent: str) -> List[str]:
    parts = [p.lower() for p in _brand_name_parts(talent)]
    if len(parts) < 2:
        return []
    first = _slug_chars(parts[0])
    last = _slug_chars(parts[-1])
    out: List[str] = []
    seen: set = set()
    for slug in (
        first[:1] + last[:5] if len(last) >= 5 else "",
        first[:1] + last[:4] if len(last) >= 4 else "",
        first[:2] + last[:5] if len(first) >= 2 and len(last) >= 5 else "",
        first[:1] + last,
    ):
        if slug and len(slug) >= 4 and slug not in seen:
            seen.add(slug)
            out.append(slug)
    return out


def _primary_sport_search_term(
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
    exp: Dict[str, bool],
) -> str:
    """Best sport keyword for Serper queries from row metadata."""
    expected = _expected_sports_from_metadata(
        title_category, title_sub_category, search_keywords, exp
    )
    for sport in ("baseball", "basketball", "football", "soccer", "hockey", "tennis", "golf"):
        if sport in expected:
            return sport
    if exp.get("expects_basketball"):
        return "basketball"
    if exp.get("expects_baseball"):
        return "baseball"
    return "athlete"


def _athlete_row_active(exp: Dict[str, bool]) -> bool:
    return bool(
        exp.get("expects_athlete")
        or exp.get("expects_basketball")
        or exp.get("expects_baseball")
    )


# ─────────────────────────────────────────────
#  CANDIDATE SIGNALS
# ─────────────────────────────────────────────

_FOLLOWER_RE = re.compile(
    r"([\d,\.]+)\s*[KkMmBb]?\s*(?:followers|subscribers|fans)", re.I
)


def _parse_follower_count(text: str) -> Optional[float]:
    m = _FOLLOWER_RE.search(text or "")
    if not m:
        return None
    raw = m.group(1).replace(",", "")
    try:
        val = float(raw)
    except ValueError:
        return None
    lower_full = m.group(0).lower()
    if "b" in lower_full:
        val *= 1_000_000_000
    elif "m" in lower_full:
        val *= 1_000_000
    elif "k" in lower_full:
        val *= 1_000
    return val


def _candidate_has_exact_profile_title(talent: str, candidate: dict) -> bool:
    title = (candidate.get("title") or "").lower()
    t = re.sub(r"\s+", " ", (talent or "").strip()).lower()
    return bool(
        t
        and re.match(rf"^{re.escape(t)}\s*\(@[A-Za-z0-9_.-]{{2,30}}\)", title)
    )


def _candidate_has_name_matching_personal_website(talent: str, candidate: dict) -> bool:
    blob = f"{candidate.get('title') or ''} {candidate.get('snippet') or ''}".lower()
    name_slug = _slug_chars(talent)
    website_match = re.search(r"website:\s*https?://(?:www\.)?([^/\s]+)", blob)
    if not website_match or not name_slug:
        return False
    website_host = website_match.group(1).lower()
    social_hosts = ("instagram.", "x.", "twitter.", "facebook.", "tiktok.", "youtube.")
    return name_slug in _slug_chars(website_host) and not any(h in website_host for h in social_hosts)


def build_candidate_signals(talent: str, candidate: dict, platform: str) -> dict:
    title   = (candidate.get("title")   or "").strip()
    snippet = (candidate.get("snippet") or "").strip()
    link    = (candidate.get("link")    or "").strip()
    blob    = f"{title} {snippet}".lower()
    path    = urlparse(link).path if link else ""
    slug    = _slug_chars(path)
    t_slug  = _slug_chars(talent)

    name_parts             = re.sub(r"\s+", " ", talent).lower().split()
    name_tokens_in_url     = [p for p in name_parts if len(p) >= 3 and _slug_chars(p) in slug]
    name_tokens_in_title   = [p for p in name_parts if p in title.lower()]
    name_tokens_in_snippet = [p for p in name_parts if p in snippet.lower()]
    full_name_in_url       = len(t_slug) >= 5 and t_slug[:min(8, len(t_slug))] in slug

    verification_signals = []
    if "official" in blob:
        verification_signals.append("official")
    if "verified" in blob or "✓" in title or "✔" in title:
        verification_signals.append("verified")
    if "blue check" in blob or "checkmark" in blob:
        verification_signals.append("blue_check")

    follower_count = _parse_follower_count(blob)

    PROFESSION_MARKERS = [
        "realtor", "real estate", "mortgage", "listing agent",
        "basketball", "nba", "wnba", "athlete", "ncaa", "espn",
        "singer", "musician", "rapper", "artist", "album",
        "journalist", "reporter", "anchor", "news anchor",
        "actor", "actress", "film", "director",
        "ceo", "founder", "executive", "entrepreneur",
        "senator", "governor", "congressman", "politician",
        "author", "writer", "novelist",
        "doctor", "physician", "surgeon", "dentist", "lawyer", "attorney",
        "digital creator", "content creator", "influencer",
    ]
    profession_signals = [p for p in PROFESSION_MARKERS if p in blob]
    url_depth          = len([s for s in path.strip("/").split("/") if s])

    handle_match = re.search(r"tiktok\.com/@(\w+)", link.lower())
    if not handle_match:
        handle_match = re.search(r"instagram\.com/(\w+)", link.lower())
    if not handle_match:
        handle_match = re.search(r"(?:x|twitter)\.com/(\w+)", link.lower())
    handle            = handle_match.group(1) if handle_match else ""
    handle_suspicious = bool(handle and (len(handle) <= 3 or handle.isdigit()))

    return {
        "name_tokens_in_url":     name_tokens_in_url,
        "name_tokens_in_title":   name_tokens_in_title,
        "name_tokens_in_snippet": name_tokens_in_snippet,
        "full_name_in_url":       full_name_in_url,
        "verification_signals":   verification_signals,
        "follower_count":         follower_count,
        "profession_signals":     profession_signals,
        "url_depth":              url_depth,
        "handle":                 handle,
        "handle_suspicious":      handle_suspicious,
        "exact_profile_title":    _candidate_has_exact_profile_title(talent, candidate),
        "personal_website_match": _candidate_has_name_matching_personal_website(talent, candidate),
        "is_valid_profile_url":   is_valid_profile_url(link, platform),
        "search_position":        int(candidate.get("search_position", 99) or 99),
        "recovered_from_text":    bool(candidate.get("recovered_from_text")),
    }


# ─────────────────────────────────────────────
#  CATEGORY DISAMBIGUATION CONTEXT
# ─────────────────────────────────────────────

def get_category_disambiguation_context(
    title_category: str,
    title_sub_category: str,
    talent: str = "",
) -> str:
    exp   = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    lines: List[str] = []

    if exp.get("expects_brand"):
        handles = ", ".join(f"@{h}" for h in _brand_search_handles(talent)[:4]) if talent else ""
        lines += [
            "CATEGORY: This row is a MEDIA BRAND / PUBLISHER / TV NETWORK (organization), not an individual person.",
            "SELECT the official organization page for this brand on this platform.",
            "CONFIRM if: URL handle matches the brand name, or title/snippet names the network/outlet/channel.",
            "CONFIRM if: verified/official signals and high follower counts typical of major media brands.",
            f"PREFERRED HANDLES (search hints): {handles or '(derive from talent name)'}",
            "REJECT if: profile is clearly an individual employee, journalist, or fan — unless the talent name is that person.",
            "REJECT if: vertical/sub-brand page (e.g. /BrandFootball, /BrandPR, /BrandCBB) when the talent is the parent brand.",
            "REJECT if: parent brand handle when talent is a distinct sub-brand (e.g. foxsports for FOX Sports 1).",
        ]
        return "\n".join(lines)

    if _athlete_row_active(exp):
        sport = _primary_sport_search_term(
            title_category, title_sub_category, "", exp
        )
        lines += [
            f"CATEGORY: This person is a {sport} ATHLETE.",
            "CONFIRM if: snippet/title matches the expected sport, NCAA/college team, draft, or pro league context.",
            "REJECT if: snippet/title mentions 'realtor', 'real estate', 'mortgage', 'listing agent', 'homes for sale', 'digital creator' (without sports context).",
            "REJECT if: the profile clearly depicts a different profession or a different sport.",
        ]
        if sport == "baseball":
            lines += [
                "CONFIRM if: mentions baseball, MLB, MiLB, pitcher, catcher, infield, college baseball, NCAA baseball.",
                "REJECT if: bio/snippet says golfer, golf, handicap, club golf, or PGA.",
                "REJECT if: bio/snippet is clearly basketball-only (NBA, hoops, point guard) with no baseball context.",
            ]
        elif sport == "basketball":
            lines += [
                "CONFIRM if: mentions basketball, NBA, WNBA, NCAA, hoops, draft, team name.",
                "REJECT if: bio/snippet says golfer, golf, handicap, club golf, or PGA.",
            ]
        lines += [
            "PREFER compact athlete handles (e.g. jthomp15) over FullName+numbers handles (e.g. JakeThompson141) when both appear.",
        ]
        if exp["expects_male"]:
            lines.append(
                "REJECT if: the profile clearly belongs to a woman when we expect a male athlete."
            )

    if exp["expects_musician"]:
        lines += [
            "CATEGORY: This person is a MUSICIAN / ARTIST.",
            "CONFIRM if: mentions music, songs, album, tour, label, artist, rapper, singer, band.",
            "REJECT if: the profile is a tribute act, cover band, or fan page.",
        ]

    if exp["expects_journalist"]:
        lines += [
            "CATEGORY: This person is a JOURNALIST / MEDIA HOST / NEWS ANCHOR.",
            "CONFIRM if: mentions journalism, news, reporting, broadcasting, anchor, host, network name.",
            "REJECT if: this is clearly a personal trainer, realtor, or unrelated person with the same name.",
        ]

    if exp["expects_actor"]:
        lines += [
            "CATEGORY: This person is an ACTOR / ACTRESS.",
            "CONFIRM if: mentions film, TV, show, movie, series, screen, Broadway, IMDb.",
            "REJECT if: the profile is a fan account.",
        ]

    if exp["expects_executive"]:
        lines += [
            "CATEGORY: This person is a BUSINESS EXECUTIVE / FOUNDER / CEO.",
            "CONFIRM if: mentions company, startup, CEO, founder, entrepreneur, board, leadership.",
        ]

    if exp["expects_politician"]:
        lines += [
            "CATEGORY: This person is a POLITICIAN / PUBLIC OFFICIAL.",
            "CONFIRM if: mentions senator, congressman, governor, mayor, representative, campaign.",
            "REJECT if: the profile is a fan/parody account.",
        ]

    if not lines:
        lines = [
            "No specific category metadata available.",
            "Use name matching and profile authenticity signals (verified, official, follower count) to disambiguate.",
        ]

    return "\n".join(lines)


# ─────────────────────────────────────────────
#  ENTITY REJECTION
# ─────────────────────────────────────────────

def entity_profile_rejected(
    talent: str,
    title_category: str,
    title_sub_category: str,
    candidate: Optional[dict],
    platform: str = "",
) -> Tuple[bool, str]:
    if not candidate:
        return False, ""

    title   = (candidate.get("title")   or "")
    snippet = (candidate.get("snippet") or "")
    link    = (candidate.get("link")    or "")
    blob    = f"{title} {snippet}".lower()
    exp     = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    search_kw = extract_search_keywords(title_category, title_sub_category)
    sport_bad, sport_why = _athlete_sport_mismatch(
        talent, title_category, title_sub_category, search_kw, candidate, exp
    )
    if sport_bad:
        return True, sport_why

    plat_for_suppression = platform or _platform_from_link(link)
    if plat_for_suppression and _platform_suppressed_for_talent(talent, plat_for_suppression):
        return True, f"{plat_for_suppression} suppressed for this talent after manual validation."

    if exp.get("expects_brand"):
        plat      = platform or _platform_from_link(link)
        path_slug = _path_handle_slug(link, plat) if plat else _slug_chars(urlparse(link).path)
        raw_handle = _normal_handle_text(_path_handle_text(link, plat)) if plat else ""
        canonical  = _brand_canonical_slug(talent)
        if _brand_slug_is_vertical(canonical, path_slug, talent, plat):
            return True, "Brand row: URL is a vertical/sub-brand or parent-mismatch handle."
        preferred_handles = {_normal_handle_text(h) for h in _brand_platform_handles(talent, plat)}
        if preferred_handles and raw_handle and raw_handle not in preferred_handles:
            if (
                plat == "YouTube"
                and _is_youtube_channel_id_slug(path_slug)
                and _candidate_supports_brand(talent, candidate)
            ):
                pass
            else:
                return True, "Brand row: this platform has a more specific official handle for this organization."
        handle = path_slug
        if (
            plat == "YouTube"
            and _is_youtube_channel_id_slug(handle)
            and _candidate_supports_brand(talent, candidate)
        ):
            pass
        elif handle and not _brand_handle_matches_row(talent, handle, plat):
            if plat == "YouTube" and _candidate_supports_brand(talent, candidate):
                pass
            else:
                return True, "Brand row: profile handle does not match this organization's known aliases."

    sport_markers = (
        "basketball", "nba", "wnba", "ncaa", "college basketball",
        "draft", "athlete", "espn", "sport", "point guard",
        "shooting guard", "forward", "center", "hoops", "nba draft",
        "football", "nfl", "soccer", "mls", "baseball", "mlb", "pitcher", "catcher",
    )
    sport_hit = any(m in blob for m in sport_markers)

    non_sport_professions = (
        "realtor", "real estate", "mortgage", "homes realty",
        "florida homes", "digital creator", "realtor sales",
        "realty & mortgage", "realty and mortgage",
        "listing agent", "homes for sale", "property management",
        "nft creator", "crypto investor", "cryptocurrency", "web3",
    )
    non_sport_hit = any(m in blob for m in non_sport_professions)

    if exp["expects_male"] and _athlete_row_active(exp):
        lower_link       = link.lower()
        profile_handle   = _path_handle_slug(link, platform or _platform_from_link(link))
        org_profile_markers = (
            "nbagleague", "gleague", "league", "team", "sports", "sport",
            "recruit", "rivals", "on3", "espn", "overtime", "athletics",
            "news", "247sports", "maxpreps", "mbb", "wbb", "mensbball",
            "womensbball", "bball", "hoops", "athletic",
        )
        t              = re.sub(r"\s+", " ", (talent or "").strip()).lower()
        name_evidence  = bool(t and (t in blob or t.replace(" ", "") in _slug_chars(blob)))
        handle_evidence = talent_url_aligned(talent, link, title_category, title_sub_category)
        if (
            profile_handle
            and any(m in profile_handle for m in org_profile_markers)
            and not handle_evidence
        ):
            return True, "Athlete row: profile handle belongs to a league/team/news account."
        if (
            "facebook.com" in lower_link
            and (profile_handle.isdigit() or "profile.php" in lower_link)
            and not (name_evidence and sport_hit)
        ):
            return True, "Athlete row: numeric/generic Facebook profile lacks identity evidence."
        if non_sport_hit and not sport_hit:
            return True, "Metadata = male athlete; result is realtor/creator/real-estate professional."
        female_hits = (
            "bobbie ", " bobbie", "brittany ", "britney ", "jessica ",
            "samantha ", "miss ", " mrs ", "she is ", "she's ", "her ",
        )
        if any(x in blob for x in female_hits) and not sport_hit:
            return True, "Profile text signals a different (female) person; talent is a male athlete."

    if exp["expects_female"] and exp["expects_athlete"] and non_sport_hit and not sport_hit:
        male_lean = (" mr ", "his ", "his own", "father", "husband")
        if any(x in blob for x in male_lean) and "woman" not in blob:
            return True, "Metadata = female athlete; result appears to be an unrelated male professional."

    if exp["expects_musician"]:
        non_music    = ("realtor", "real estate", "lawyer", "attorney", "doctor", "physician", "financial advisor")
        music_confirm = ("music", "artist", "singer", "rapper", "album", "tour", "label", "song")
        if any(m in blob for m in non_music) and not any(m in blob for m in music_confirm):
            return True, "Metadata = musician; result is clearly a non-music professional."
        if any(x in blob for x in ("tribute", "tribute band", "fan page", "unofficial")):
            return True, "Fan/tribute/unofficial page — not the artist's own profile."

    if exp["expects_journalist"]:
        non_media    = ("realtor", "real estate", "fitness trainer", "personal trainer", "chef")
        media_confirm = ("news", "journalist", "anchor", "reporter", "host", "broadcasting", "media")
        if any(m in blob for m in non_media) and not any(m in blob for m in media_confirm):
            return True, "Metadata = journalist/anchor; result is a non-media professional."

    article_signals = (
        " - wikipedia", "wikipedia.org", "imdb.com", "biography",
        "interviews", "profile of ", "article about", "story of",
        " - espn.com", " | espn", " - bleacher report",
    )
    if any(x in (link + blob).lower() for x in article_signals):
        return True, "Result appears to be an editorial article or wiki page, not a social profile."

    plat = platform or _platform_from_link(link)
    if plat == "X":
        x_handle = _x_handle_from_url(link)
        if x_handle and not _is_valid_x_profile_handle(x_handle):
            return True, f"X/Twitter URL is a site path ({x_handle}), not a user profile."

    return False, ""


# ─────────────────────────────────────────────
#  CANDIDATE RANKING
# ─────────────────────────────────────────────

def candidate_rank_score(
    talent: str,
    c: dict,
    search_keywords: str,
    title_category: str = "",
    title_sub_category: str = "",
    username_hints: Optional[Dict[str, str]] = None,
    platform: str = "",
    # NEW: bonus for candidates injected from early-enrich phase
    early_enrich_bonus: bool = False,
) -> float:
    title   = (c.get("title")   or "").lower()
    snippet = (c.get("snippet") or "").lower()
    link    = (c.get("link")    or "").lower()
    query   = (c.get("query")   or "").lower()
    t       = re.sub(r"\s+", " ", (talent or "").strip()).lower()
    score   = 0.0
    exp_rank     = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    name_aliases = _identity_name_aliases(talent, title_category, title_sub_category)
    alias_slugs  = {_slug_chars(a) for a in name_aliases if a.lower() != t}

    # ── NEW: early-enrich candidates already validated from a known profile bio ──
    if early_enrich_bonus:
        score += 8.0

    # ── Authenticity signals ──
    if "official" in title or "official" in snippet:
        score += 4.0
    if "verified" in title or "verified" in snippet or "✓" in (c.get("title") or ""):
        score += 3.0

    # ── Name match ──
    if t and t in title:
        score += 3.5
    if t and t in snippet:
        score += 2.0
    for alias in name_aliases:
        alias_l = alias.lower()
        if alias_l == t or len(alias_l) < 4:
            continue
        if alias_l in title:
            score += 3.0
        if alias_l in snippet:
            score += 1.5

    # ── URL slug alignment ──
    path_slug = _slug_chars(urlparse(link).path)
    name_slug = _slug_chars(talent)
    if name_slug and len(name_slug) >= 5 and name_slug[:min(8, len(name_slug))] in path_slug:
        score += 4.0
    else:
        for part in re.sub(r"\s+", " ", (talent or "").strip()).lower().split():
            sp = _slug_chars(part)
            if len(sp) >= 4 and sp in path_slug:
                score += 1.5
    for alias_slug in alias_slugs:
        if len(alias_slug) >= 5 and alias_slug in path_slug:
            score += 4.0
    if platform == "YouTube" and _is_youtube_channel_id_slug(_path_handle_slug(link, platform)):
        alias_hit = any(
            alias.lower() in title or alias.lower() in snippet
            for alias in name_aliases
            if len(alias) >= 4
        )
        if alias_hit:
            score += 6.0

    # ── Follower / subscriber signal ──
    follower_count = _parse_follower_count(f"{title} {snippet}")
    if follower_count:
        if follower_count >= 1_000_000:
            score += 3.0
        elif follower_count >= 100_000:
            score += 2.0
        elif follower_count >= 10_000:
            score += 1.0

    # ── Athlete-specific search evidence ──
    if not exp_rank.get("expects_brand") and _athlete_row_active(exp_rank):
        blob             = f"{title} {snippet}"
        position         = int(c.get("search_position", 99) or 99)
        profile_handle   = _path_handle_slug(link, platform)
        sport_bad, _ = _athlete_sport_mismatch(
            talent, title_category, title_sub_category, search_keywords, c, exp_rank
        )
        if sport_bad:
            score -= 45.0
        profile_title_pattern = bool(
            t
            and t in title
            and re.search(r"\(@[A-Za-z0-9_.-]{2,30}\)", c.get("title") or "")
        )
        exact_profile_title = bool(_candidate_has_exact_profile_title(talent, c))
        if profile_title_pattern and platform in {"X", "Instagram", "TikTok"} and not sport_bad:
            score += 7.0
            if position <= 3:
                score += 2.0
        if exact_profile_title and platform in {"X", "Instagram", "TikTok"} and not sport_bad:
            score += 8.0
        compact_slugs = _compact_athlete_handle_slugs(talent)
        if profile_handle and profile_handle in {_slug_chars(a) for a in _athlete_handle_aliases(talent)}:
            score += 7.0
        if profile_handle and compact_slugs:
            if profile_handle in compact_slugs:
                score += 10.0
            elif any(profile_handle.startswith(s) and len(s) >= 5 for s in compact_slugs):
                score += 8.0
        if profile_handle and _is_fullname_plus_digits_handle(talent, profile_handle):
            score -= 10.0
        if t and t in blob and any(s in query for s in ("official", "verified")):
            score += 5.0
            if position <= 3:
                score += 2.0
        if _candidate_has_name_matching_personal_website(talent, c):
            score += 6.0
        if t and t in blob and any(s in query for s in ("basketball", "athlete", "hoops")):
            score += 4.0
            if position <= 3:
                score += 3.0
            if c.get("recovered_from_text"):
                score += 5.0
        name_parts = [p for p in t.split() if len(p) >= 3]
        last_name  = name_parts[-1] if name_parts else ""
        if (
            last_name
            and last_name in blob
            and any(s in query for s in ("basketball", "athlete", "hoops"))
            and position <= 3
        ):
            score += 4.0
        if profile_handle and any(m in profile_handle for m in ("team", "league", "sports", "recruit", "espn")):
            score -= 10.0
        if platform == "Facebook" and (profile_handle.isdigit() or "profile.php" in link):
            score -= 8.0

    # ── Metadata keyword alignment ──
    sport_bias_tokens = {
        "sports", "sport", "basketball", "football", "publication", "publishers", "publisher",
    }
    for token in _metadata_tokens(search_keywords):
        if len(token) < 3:
            continue
        if exp_rank.get("expects_brand") and token in sport_bias_tokens:
            continue
        if token in title or token in snippet:
            score += 1.5

    # ── Brand handle alignment ──
    if exp_rank.get("expects_brand"):
        brand_slug = _path_handle_slug(link, platform) if platform else path_slug
        raw_handle = _normal_handle_text(_path_handle_text(link, platform)) if platform else ""
        canonical  = _brand_canonical_slug(talent)
        preferred_handles = [_normal_handle_text(h) for h in _brand_platform_handles(talent, platform)]
        if raw_handle and raw_handle in preferred_handles:
            score += 18.0
            score += max(0.0, 4.0 - preferred_handles.index(raw_handle))
        elif preferred_handles and any(_slug_chars(h) == brand_slug for h in preferred_handles):
            score -= 8.0
        if platform == "Instagram" and canonical and brand_slug == f"{canonical}hq":
            score += 13.0
        elif canonical and brand_slug == canonical:
            score += 12.0
        elif _brand_handle_matches_row(talent, brand_slug, platform):
            rank = _brand_handle_rank(talent, brand_slug)
            score += max(5.0, 11.0 - min(rank, 8))
            if brand_slug.endswith("official"):
                score -= 1.5
            for ac in _brand_explicit_acronyms(talent):
                ac_slug = _slug_chars(ac)
                if platform == "X" and brand_slug in {f"{ac_slug}sports", f"{ac_slug}sport"}:
                    score += 3.0
                elif platform == "TikTok" and brand_slug == ac_slug:
                    score += 3.0
        elif platform == "YouTube" and _is_youtube_channel_id_slug(brand_slug):
            if _candidate_supports_brand(talent, c):
                score += 11.0
        elif _candidate_supports_brand(talent, c):
            score += 5.0
        elif canonical and _brand_slug_is_vertical(canonical, brand_slug, talent, platform):
            score -= 20.0

    # ── Username hint bonus ──
    if username_hints and platform:
        for src_plat, hint in username_hints.items():
            if src_plat == platform or not hint:
                continue
            hint_slug = _slug_chars(hint)
            if hint_slug in path_slug:
                hint_bonus = 5.0
                if (
                    not exp_rank.get("expects_brand")
                    and _athlete_row_active(exp_rank)
                    and src_plat == "Facebook"
                    and platform in {"X", "Instagram", "TikTok"}
                    and hint_slug == name_slug
                ):
                    hint_bonus = 1.0
                score += hint_bonus
                break

    # ── Generic handle penalty ──
    segs = [s for s in urlparse(link).path.strip("/").split("/") if s]
    if segs:
        handle = segs[-1].lstrip("@")
        if len(handle) <= 2 or handle.isdigit():
            score -= 3.0

    # ── Hard entity rejection ──
    rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, c, platform)
    if rej:
        score -= 35.0

    return score


# ─────────────────────────────────────────────
#  QUERY BUILDING
# ─────────────────────────────────────────────

def build_queries(
    talent: str,
    platform: str,
    domains: List[str],
    search_keywords: str,
    title_category: str = "",
    title_sub_category: str = "",
    username_hints: Optional[Dict[str, str]] = None,
    wikipedia_context: str = "",
    # NEW: if True, skip exhaustive query list (fast path for verify-only calls)
    fast: bool = False,
) -> List[str]:
    kw         = (search_keywords or "").strip()
    exp        = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    queries: List[str] = []
    priority_queries: List[str] = []
    wiki_title = _wikipedia_title_from_context(wikipedia_context)
    name_aliases   = _identity_name_aliases(talent, title_category, title_sub_category, wikipedia_context)
    handle_aliases = _identity_handle_aliases(talent, title_category, title_sub_category, wikipedia_context)

    if exp.get("expects_brand"):
        platform_handles = _brand_platform_handles(talent, platform)
        for handle in platform_handles:
            for domain in domains:
                priority_queries.append(f"site:{domain}/{handle}")
                priority_queries.append(f"site:{domain}/@{handle}")
            priority_queries.append(f'"{handle}" {platform} official')
        if platform == "Instagram":
            canonical = _brand_canonical_slug(talent)
            for domain in domains:
                queries.insert(0, f"site:{domain}/{canonical}_hq")
                queries.insert(0, f"site:{domain}/{canonical}hq")
            parts_for_mix = _brand_name_parts(talent)
            if len(parts_for_mix) >= 2:
                mixed = parts_for_mix[0].lower() + "_" + "".join(p.lower() for p in parts_for_mix[1:])
                for domain in domains:
                    queries.insert(0, f"site:{domain}/{mixed}")
        if platform == "X":
            for ac in _brand_explicit_acronyms(talent):
                for domain in domains:
                    queries.insert(0, f"site:{domain}/{ac}_sports")
                    queries.insert(0, f"site:{domain}/{ac}sports")
        for handle in _brand_search_handles(talent):
            for domain in domains:
                queries.append(f"site:{domain}/{handle}")
                queries.append(f"site:{domain}/@{handle}")
            queries.append(f'"{handle}" {platform} official')
        parts = _brand_name_parts(talent)
        parent_slug = _brand_parent_slug(talent)
        if parent_slug and platform == "YouTube":
            for domain in domains:
                if "youtube" in domain:
                    queries.insert(0, f"site:{domain}/user/{parent_slug}")
                    cap = parent_slug[:1].upper() + parent_slug[1:]
                    queries.insert(0, f"site:{domain}/user/{cap}")
        if platform == "YouTube":
            for domain in domains:
                if "youtube" in domain:
                    queries.insert(0, f'site:{domain}/channel "{talent}"')
                    queries.insert(0, f'site:{domain} "{talent}" official channel')
        if parts:
            pascal = "".join((p[:1].upper() + p[1:].lower()) if p.isalpha() else p for p in parts)
            compact_pascal = "".join((p[:1].upper() + p[1:].lower()) if len(p) > 1 else p.lower() for p in parts)
            def _pascal_token(p: str) -> str:
                pl = p.lower()
                if pl in _BRAND_WORD_ABBREV:
                    ab = _BRAND_WORD_ABBREV[pl]
                    return ab[:1].upper() + ab[1:].lower()
                if p.isalpha():
                    return p[:1].upper() + p[1:].lower()
                return p
            abbrev_pascal = "".join(_pascal_token(p) for p in parts)
            for domain in domains:
                queries.insert(0, f"site:{domain}/{pascal}")
                if compact_pascal.lower() != pascal.lower():
                    queries.insert(0, f"site:{domain}/{compact_pascal}")
                if abbrev_pascal.lower() not in (pascal.lower(), compact_pascal.lower()):
                    queries.insert(0, f"site:{domain}/{abbrev_pascal}")
                if platform == "Instagram":
                    canonical = _brand_canonical_slug(talent)
                    queries.insert(0, f"site:{domain}/{canonical}_hq")
                    queries.insert(0, f"site:{domain}/{canonical}hq")
        for domain in domains:
            if wiki_title and wiki_title.lower() != talent.lower():
                queries.append(f'site:{domain} "{talent}" "{wiki_title}" official')
            queries.append(f'site:{domain} "{talent}" official')
            queries.append(f'site:{domain} "{talent}" verified')
            queries.append(f'site:{domain} "{talent}"')
        queries.append(f'"{talent}" {platform} official')
        queries.append(f'"{talent}" {platform}')
        seen: set = set()
        unique: List[str] = []
        for q in priority_queries + queries:
            if q not in seen:
                seen.add(q)
                unique.append(q)
        return unique

    for handle in handle_aliases:
        for domain in domains:
            priority_queries.append(f"site:{domain}/{handle}")
            priority_queries.append(f"site:{domain}/@{handle}")
        priority_queries.append(f'"{handle}" {platform}')

    for alias in name_aliases:
        if alias.lower() == talent.lower():
            continue
        for domain in domains:
            priority_queries.append(f'site:{domain} "{alias}" official')
            priority_queries.append(f'site:{domain} "{alias}" verified')
            priority_queries.append(f'site:{domain} "{alias}"')
            if platform == "YouTube":
                priority_queries.append(f'site:{domain} "{alias}" channel')
        priority_queries.append(f'"{alias}" {platform} official')
        priority_queries.append(f'"{alias}" {platform}')

    if _athlete_row_active(exp):
        sport_term = _primary_sport_search_term(
            title_category, title_sub_category, search_keywords, exp
        )
        for domain in domains:
            if wiki_title and wiki_title.lower() != talent.lower():
                queries.append(f'site:{domain} "{talent}" "{wiki_title}" {sport_term}')
            queries.append(f'site:{domain} "{talent}" {sport_term}')
            queries.append(f'site:{domain} "{talent}" official')
            queries.append(f'site:{domain} "{talent}" verified')
        for handle in _athlete_handle_aliases(talent):
            for domain in domains:
                priority_queries.append(f"site:{domain}/{handle}")
                priority_queries.append(f"site:{domain}/@{handle}")
            priority_queries.append(f'"{handle}" {platform} {sport_term}')
        for compact in _compact_athlete_handle_slugs(talent):
            for domain in domains:
                priority_queries.insert(0, f"site:{domain}/{compact}")
                priority_queries.insert(0, f"site:{domain}/@{compact}")

    if username_hints:
        for src_plat, hint in username_hints.items():
            if src_plat == platform or not hint:
                continue
            for domain in domains:
                queries.append(f'site:{domain} "@{hint}"')
                queries.append(f'site:{domain} "{hint}"')
            queries.append(f'"{hint}" {platform}')
            break

    if exp["expects_male"] and _athlete_row_active(exp):
        sport_term = _primary_sport_search_term(
            title_category, title_sub_category, search_keywords, exp
        )
        for domain in domains:
            queries.append(f'site:{domain} "{talent}" {sport_term}')
            queries.append(f'site:{domain} "{talent}" {sport_term} player')
            if sport_term == "baseball":
                queries.append(f'site:{domain} "{talent}" NCAA baseball')
            elif sport_term == "basketball":
                queries.append(f'site:{domain} "{talent}" NCAA basketball')

    if exp["expects_journalist"]:
        for domain in domains:
            queries.append(f'site:{domain} "{talent}" journalist anchor')
            queries.append(f'site:{domain} "{talent}" news host')

    if exp["expects_musician"]:
        for domain in domains:
            queries.append(f'site:{domain} "{talent}" music artist')
            queries.append(f'site:{domain} "{talent}" official artist')

    for domain in domains:
        if wiki_title and wiki_title.lower() != talent.lower():
            queries.append(f'site:{domain} "{talent}" "{wiki_title}" official')
        queries.append(f'site:{domain} "{talent}" official')
        queries.append(f'site:{domain} "{talent}" verified')
        queries.append(f'site:{domain} "{talent}"')
        if kw:
            queries.append(f'site:{domain} "{talent}" {kw} official')
            queries.append(f'site:{domain} "{talent}" {kw}')

    queries.append(f'"{talent}" {platform} official')
    queries.append(f'"{talent}" {platform}')
    if wiki_title and wiki_title.lower() != talent.lower():
        queries.append(f'"{talent}" "{wiki_title}" {platform} official')
    if kw:
        queries.append(f'"{talent}" {kw} {platform} official')
        queries.append(f'"{talent}" {kw} {platform}')

    seen2: set = set()
    unique2: List[str] = []
    for q in priority_queries + queries:
        if q not in seen2:
            seen2.add(q)
            unique2.append(q)

    if fast:
        return unique2[:6]
    return unique2


# ─────────────────────────────────────────────
#  DATA LOADING
# ─────────────────────────────────────────────

def _default_talent_table() -> pd.DataFrame:
    n = len(talent_names)
    data: Dict[str, List] = {
        "Talent Name":          list(talent_names),
        "title_category":       [""] * n,
        "title_sub_category":   [""] * n,
        WIKIPEDIA_URL_COLUMN:   [""] * n,
        INSTAGRAM_INPUT_COLUMN: [""] * n,
    }
    for p in PLATFORMS:
        data[p] = [""] * n
    for c in PLATFORM_CONF_COLUMNS.values():
        data[c] = [float("nan")] * n
    data["Confidence"] = [float("nan")] * n
    data["Source"]     = [""] * n
    return pd.DataFrame(data)


def load_talent_table_from_path(excel_path: Path) -> pd.DataFrame:
    excel_path = Path(excel_path)
    if not excel_path.is_file():
        raise ValueError(f"File not found: {excel_path}")
    suffix = excel_path.suffix.lower()
    try:
        raw = pd.read_csv(excel_path) if suffix == ".csv" else pd.read_excel(excel_path)
    except Exception as exc:
        raise ValueError(f"Could not read spreadsheet: {exc}") from exc
    if raw.empty:
        raise ValueError("The file has no rows.")

    name_col = _find_column(raw, "Talent Name", "Talent", "title", "Title", "Name")
    if name_col is None:
        name_col = raw.columns[0]
    cat_col  = _find_column(raw, "title_category", "de_category", "category", "Title Category")
    sub_col  = _find_column(raw, "title_sub_category", "sub_category", "Title Sub Category", "subtitle")
    wiki_col = _find_column(
        raw,
        WIKIPEDIA_URL_COLUMN,
        "wikipedia_url", "Wikipedia", "Wiki URL", "wiki_url", "Wiki", "Wikipedia Link",
    )
    # ── NEW: detect client-provided Instagram URL column ──
    ig_input_col = _find_column(
        raw,
        INSTAGRAM_INPUT_COLUMN,
        "instagram_url", "Instagram Input", "ig_url", "IG URL", "Instagram Link", "Instagram",
    )

    names_list, cat_list, sub_list, wiki_list, ig_input_list = [], [], [], [], []
    for i in range(len(raw)):
        name = str(raw.iloc[i][name_col]).strip()
        if not name or name.lower() == "nan":
            continue
        names_list.append(name)
        c = raw.iloc[i][cat_col]  if cat_col  else ""
        s = raw.iloc[i][sub_col]  if sub_col  else ""
        w = raw.iloc[i][wiki_col] if wiki_col else ""
        ig = raw.iloc[i][ig_input_col] if ig_input_col else ""
        cat_list.append("" if (isinstance(c, float) and pd.isna(c)) else str(c).strip())
        sub_list.append("" if (isinstance(s, float) and pd.isna(s)) else str(s).strip())
        wiki_list.append(_clean_wikipedia_url(w))
        ig_input_list.append(_clean_instagram_input_url(ig))

    if not names_list:
        raise ValueError("No valid talent names found.")
    n = len(names_list)
    out: Dict[str, List] = {
        "Talent Name":          names_list,
        "title_category":       cat_list,
        "title_sub_category":   sub_list,
        WIKIPEDIA_URL_COLUMN:   wiki_list,
        INSTAGRAM_INPUT_COLUMN: ig_input_list,
    }
    for p in PLATFORMS:
        out[p] = [""] * n
    for c in PLATFORM_CONF_COLUMNS.values():
        out[c] = [float("nan")] * n
    out["Confidence"] = [float("nan")] * n
    out["Source"]     = [""] * n
    return pd.DataFrame(out)


def load_talent_table() -> pd.DataFrame:
    if not TEST_BRANDS_PATH.exists():
        return _default_talent_table()
    try:
        return load_talent_table_from_path(TEST_BRANDS_PATH)
    except ValueError as exc:
        print(f"[WARN] {exc}. Using default talent_names.")
        return _default_talent_table()


def build_talent_df(names: List[str], platforms: List[str]) -> pd.DataFrame:
    data: Dict[str, List] = {"Talent Name": names}
    for p in platforms:
        data[p] = [""] * len(names)
    for p in platforms:
        data[f"{p} Confidence"] = [float("nan")] * len(names)
    data["title_category"]     = [""] * len(names)
    data["title_sub_category"] = [""] * len(names)
    data[WIKIPEDIA_URL_COLUMN]   = [""] * len(names)
    data[INSTAGRAM_INPUT_COLUMN] = [""] * len(names)
    data["Confidence"] = [float("nan")] * len(names)
    data["Source"]     = [""] * len(names)
    return pd.DataFrame(data)


# ─────────────────────────────────────────────
#  URL VALIDATION & NORMALISATION
# ─────────────────────────────────────────────

def is_valid_profile_url(link: str, platform: str) -> bool:
    if not isinstance(link, str) or not link.strip():
        return False
    u = link.strip()
    try:
        parsed = urlparse(u)
    except Exception:
        return False
    if parsed.scheme not in ("http", "https"):
        return False
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()
    full = u.lower()

    if platform == "Facebook":
        if "facebook.com" not in host:
            return False
        if any(seg in full for seg in ("/posts/", "/photos/", "/videos/", "/watch/",
                                        "/reel", "/story.php", "/permalink/")):
            return False
        if "profile.php" in path or "/people/" in path or "/pages/" in path:
            return True
        segs = [s for s in path.strip("/").split("/") if s]
        if len(segs) == 1 and segs[0] not in (
            "share", "sharer", "groups", "events", "marketplace", "gaming", "watch"
        ):
            return True
        return False

    if platform == "Instagram":
        if "instagram.com" not in host:
            return False
        if any(x in full for x in ("/p/", "/reel", "/reels/", "/stories/",
                                    "/tv/", "/explore/", "/tags/", "/locations/")):
            return False
        segs = [s for s in path.strip("/").split("/") if s]
        return len(segs) == 1

    if platform == "YouTube":
        if "youtube.com" not in host and "youtu.be" not in host:
            return False
        if any(x in full for x in ("/watch", "/shorts", "/playlist", "/results",
                                    "/live/", "/feed/", "/attribution_link")):
            return False
        return "/@" in full or "/channel/" in full or "/c/" in full or "/user/" in full

    if platform == "X":
        if "x.com" not in host and "twitter.com" not in host:
            return False
        if "/status/" in full or "/i/" in full or "/intent/" in full or "/search" in full:
            return False
        segs = [s for s in path.strip("/").split("/") if s]
        if len(segs) != 1:
            return False
        return _is_valid_x_profile_handle(segs[0])

    if platform == "TikTok":
        if "tiktok.com" not in host:
            return False
        if any(x in full for x in ("/video/", "/tag/", "/music/", "/discover", "/foryou")):
            return False
        return bool(re.search(r"tiktok\.com/@[^/]+/?$", full))

    return False


def normalize_profile_url(url: str, platform: str) -> str:
    if not url or not isinstance(url, str):
        return ""
    u = url.strip()
    if platform == "YouTube":
        u = u.replace("://m.youtube.com", "://www.youtube.com")
        u = u.replace("://music.youtube.com", "://www.youtube.com")
        if "youtube.com" in u and "www." not in urlparse(u).netloc and "m." not in urlparse(u).netloc:
            u = u.replace("://youtube.com", "://www.youtube.com")
    return u.rstrip("/")


def talent_url_aligned(
    talent: str,
    link: str,
    title_category: str = "",
    title_sub_category: str = "",
) -> bool:
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand"):
        plat        = _platform_from_link(link)
        handle_slug = _path_handle_slug(link, plat) if plat else _slug_chars(urlparse(link).path)
        return _brand_handle_matches_row(talent, handle_slug, plat)
    t = _slug_chars(talent)
    if len(t) < 4:
        return False
    path_compact = _slug_chars(urlparse(link).path)
    if len(t) >= 6 and t[:min(8, len(t))] in path_compact:
        return True
    for part in re.sub(r"\s+", " ", (talent or "").strip()).lower().split():
        if len(part) < 3:
            continue
        sp = _slug_chars(part)
        if len(sp) >= 5 and sp in path_compact:
            return True
    return False


def first_valid_profile_link(candidates: List[dict], platform: str) -> str:
    for item in candidates:
        if is_valid_profile_url(item.get("link", ""), platform):
            return normalize_profile_url(item["link"], platform)
    return ""


def sort_candidates_for_ai(
    talent: str,
    candidates: List[dict],
    search_keywords: str,
    title_category: str = "",
    title_sub_category: str = "",
    username_hints: Optional[Dict[str, str]] = None,
    platform: str = "",
) -> List[dict]:
    return sorted(
        candidates,
        key=lambda c: -candidate_rank_score(
            talent, c, search_keywords, title_category, title_sub_category,
            username_hints=username_hints, platform=platform,
            early_enrich_bonus=bool(c.get("from_early_enrich")),
        ),
    )


def _prefer_recovered_athlete_handle(
    talent: str,
    platform: str,
    selected: str,
    candidates: List[dict],
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
) -> Optional[dict]:
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand") or not _athlete_row_active(exp):
        return None
    if platform not in {"X", "Instagram", "TikTok"}:
        return None

    selected_norm  = normalize_profile_url(selected or "", platform).rstrip("/")
    selected_score = -999.0
    for c in candidates:
        if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == selected_norm:
            selected_score = candidate_rank_score(
                talent, c, search_keywords, title_category, title_sub_category, platform=platform
            )
            break

    full_name = re.sub(r"\s+", " ", (talent or "").strip()).lower()
    for c in candidates[:4]:
        if not c.get("recovered_from_text"):
            continue
        blob = f"{c.get('title', '')} {c.get('snippet', '')}".lower()
        if full_name not in blob:
            continue
        rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, c, platform)
        if rej:
            continue
        score = candidate_rank_score(
            talent, c, search_keywords, title_category, title_sub_category, platform=platform
        )
        if score >= selected_score - 2.0:
            return c
    return None


def _prefer_non_mismatch_athlete_profile(
    talent: str,
    platform: str,
    selected: str,
    candidates: List[dict],
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
) -> Optional[dict]:
    """
    If AI picked a same-name wrong sport (e.g. golfer for a basketball row),
    switch to the best candidate that passes sport checks — or force blank.
    """
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand") or not _athlete_row_active(exp):
        return None
    if platform not in {"X", "Instagram", "TikTok", "Facebook"}:
        return None

    selected_norm = normalize_profile_url(selected or "", platform).rstrip("/")
    selected_cand: Optional[dict] = None
    for c in candidates:
        if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == selected_norm:
            selected_cand = c
            break
    if not selected_cand:
        return None

    bad, why = _athlete_sport_mismatch(
        talent, title_category, title_sub_category, search_keywords, selected_cand, exp
    )
    if not bad:
        return None

    print(f"  [SPORT-FIX] {platform} | {talent} | rejecting {selected_norm}: {why}")

    best_alt: Optional[dict] = None
    best_score = -999.0
    for c in candidates[:8]:
        rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, c, platform)
        if rej:
            continue
        sm, _ = _athlete_sport_mismatch(
            talent, title_category, title_sub_category, search_keywords, c, exp
        )
        if sm:
            continue
        rs = candidate_rank_score(
            talent, c, search_keywords, title_category, title_sub_category,
            username_hints=None, platform=platform,
            early_enrich_bonus=bool(c.get("from_early_enrich")),
        )
        if rs > best_score:
            best_score = rs
            best_alt = c

    if best_alt and best_score >= MIN_RANK_SCORE_FOR_FALLBACK - 4.0:
        return best_alt
    return {"link": "", "_force_blank": True}


def _prefer_personal_website_athlete_profile(
    talent: str,
    platform: str,
    selected: str,
    candidates: List[dict],
    title_category: str,
    title_sub_category: str,
    search_keywords: str,
) -> Optional[dict]:
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand") or not _athlete_row_active(exp):
        return None
    if platform not in {"X", "Instagram", "TikTok"} or not candidates:
        return None

    top = candidates[0]
    if not (
        _candidate_has_exact_profile_title(talent, top)
        and _candidate_has_name_matching_personal_website(talent, top)
    ):
        return None

    selected_norm = normalize_profile_url(selected or "", platform).rstrip("/")
    top_norm      = normalize_profile_url(top.get("link", ""), platform).rstrip("/")
    if selected_norm == top_norm:
        return None

    selected_candidate = next(
        (c for c in candidates if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == selected_norm),
        None,
    )
    if selected_candidate and _candidate_has_name_matching_personal_website(talent, selected_candidate):
        return None

    top_score = candidate_rank_score(
        talent, top, search_keywords, title_category, title_sub_category, platform=platform
    )
    selected_score = (
        candidate_rank_score(
            talent, selected_candidate, search_keywords, title_category, title_sub_category, platform=platform
        )
        if selected_candidate else 0.0
    )
    if top_score >= selected_score:
        return top
    return None


# ─────────────────────────────────────────────
#  USERNAME HINT EXTRACTION
# ─────────────────────────────────────────────

def extract_username_hints(resolved_links: Dict[str, str]) -> Dict[str, str]:
    hints: Dict[str, str] = {}
    for plat, url in resolved_links.items():
        if not url:
            continue
        url = url.strip().rstrip("/")
        m = re.search(r"tiktok\.com/@([\w.]+)", url, re.I)
        if m:
            hints[plat] = m.group(1)
            continue
        m = re.search(r"youtube\.com/@([\w.]+)", url, re.I)
        if m:
            hints[plat] = m.group(1)
            continue
        m = re.search(r"youtube\.com/c/([\w.]+)", url, re.I)
        if m:
            hints[plat] = m.group(1)
            continue
        m = re.search(r"(?:instagram|twitter|x|facebook)\.com/(@?)([\w.]+)", url, re.I)
        if m:
            hints[plat] = m.group(2)
            continue
    return hints


# ─────────────────────────────────────────────
#  AI — MAIN SELECTION
# ─────────────────────────────────────────────

def ai_select_best_profile(
    talent: str,
    platform: str,
    candidates: List[dict],
    entity_category: str,
    entity_sub_category: str,
    search_keywords: str,
    username_hints: Optional[Dict[str, str]] = None,
    wikipedia_context: str = "",
    # NEW: early-enriched link for this platform, if any
    early_enrich_link: str = "",
) -> dict:
    """
    Two-phase AI selection with early-enrich injection.
    Phase 1 — classify each candidate ACCEPT / MAYBE / REJECT.
    Phase 2 — pick single best from ACCEPT; fall to MAYBE only if conf ≥ 0.75.
    If early_enrich_link is provided, it is injected as a high-priority candidate.
    """
    if not candidates and not early_enrich_link:
        return {"best_link": "", "confidence": 0.0, "reason": "No candidates provided."}

    exp_sel    = parse_entity_expectations(entity_category, entity_sub_category, talent=talent)
    cat_context = get_category_disambiguation_context(entity_category, entity_sub_category, talent=talent)

    # ── Inject early-enrich link as a top candidate if not already present ──
    if early_enrich_link:
        norm_enrich = normalize_profile_url(early_enrich_link, platform).rstrip("/")
        existing_norms = {normalize_profile_url(c.get("link", ""), platform).rstrip("/") for c in candidates}
        if norm_enrich not in existing_norms:
            enrich_candidate = {
                "link":             early_enrich_link,
                "title":            f"[Bio-enriched] {talent} on {platform}",
                "snippet":          "Discovered from a confirmed high-confidence anchor profile bio.",
                "query":            "bio_enrich",
                "search_position":  0,
                "recovered_from_text": False,
                "from_early_enrich": True,
            }
            candidates = [enrich_candidate] + list(candidates)
            print(f"  [EARLY-ENRICH-INJECT] {platform} | {talent} | injected {early_enrich_link[:80]}")

    # Enrich candidates with pre-computed signals
    enriched_candidates = []
    for c in candidates:
        signals = build_candidate_signals(talent, c, platform)
        enriched_candidates.append({
            "link":             c.get("link", ""),
            "title":            c.get("title", ""),
            "snippet":          c.get("snippet", ""),
            "signals":          signals,
            "query":            c.get("query", ""),
            "from_early_enrich": bool(c.get("from_early_enrich")),
        })

    hint_lines = []
    if username_hints:
        for src_plat, hint in username_hints.items():
            hint_lines.append(
                f"  • On {src_plat} we already found @{hint} — prioritise candidates whose URL contains this handle."
            )
    # ── NEW: explicitly tell AI about the early-enrich link ──
    if early_enrich_link:
        hint_lines.append(
            f"  • BIO-ENRICHED CANDIDATE: {early_enrich_link} was discovered in the bio of a "
            "confirmed high-confidence anchor profile. Treat this as strong corroborating evidence."
        )
    hint_block = (
        "CROSS-PLATFORM HINTS:\n" + "\n".join(hint_lines)
        if hint_lines else ""
    )

    if exp_sel.get("expects_brand"):
        system_msg = (
            "You are an expert social media profile resolver for media brands, publishers, and TV networks. "
            "Your job is to identify the single official organization profile on each platform. "
            "\n\nCORE RULE: A blank cell is ALWAYS better than a wrong link. "
            "When uncertain, return empty string for best_link. "
            "\nNEVER select: posts, videos, reels, individual employee accounts, fan pages, "
            "news articles, or vertical sub-brands unless the talent name is that vertical."
        )
    else:
        system_msg = (
            "You are an expert social media profile resolver working for a talent research firm. "
            "Your job is to identify the single official, active social media profile for a real public figure. "
            "\n\nCORE RULE: A blank cell is ALWAYS better than a wrong link. "
            "When uncertain, return empty string for best_link. "
            "\nNEVER select: posts, videos, reels, shorts, news articles, Wikipedia pages, fan pages, "
            "tribute accounts, or profiles that clearly belong to a different person."
        )

    user_msg = f"""
TALENT: "{talent}"
PLATFORM: {platform}
SEARCH KEYWORDS FROM METADATA: {search_keywords or "(none)"}
WIKIPEDIA IDENTITY CONTEXT: {wikipedia_context or "(not provided)"}
KNOWN NAME ALIASES: {", ".join(_identity_name_aliases(talent, entity_category, entity_sub_category, wikipedia_context)) or "(none)"}

CATEGORY CONTEXT:
{cat_context}

{hint_block}

CANDIDATES (each has pre-computed signals to help you):
{json.dumps(enriched_candidates, indent=2, ensure_ascii=True)}

INSTRUCTIONS:
Step 1 — For EACH candidate, classify it as:
  ACCEPT  → very likely the correct official profile for this talent
  MAYBE   → possible but uncertain
  REJECT  → wrong person, fan page, article, or content URL

Use these signals in order of importance:
  1. from_early_enrich=true candidates: STRONGLY prefer — they were found in a confirmed anchor profile bio
  2. signals.name_tokens_in_url and signals.full_name_in_url (strongest identifier)
  3. Wikipedia identity context, when provided, to disambiguate the exact person/brand
  4. signals.verification_signals (official, verified, blue_check)
  5. Cross-platform username hints (if handle from another platform appears in URL)
  6. signals.profession_signals matching the expected category
  7. signals.follower_count (higher = more credible public figure)
  8. signals.name_tokens_in_title and signals.name_tokens_in_snippet

Step 2 — From all ACCEPT candidates, choose the one with the most signals.
  If no ACCEPT, choose from MAYBE only if confidence ≥ 0.75.
  If no suitable candidate: return best_link="" and confidence<0.40.

BLANK RULES (return best_link="" if any of these apply):
  • All candidates appear to be the wrong person
  • The talent name is very common and no candidate clearly confirms identity
  • The best candidate's profession_signals conflict with the expected category
  • The best candidate is clearly a fan/tribute/unofficial page
  • You cannot distinguish between 2+ legitimate people with the same name
  • For MEDIA BRANDS: reject parent-brand URLs when talent is a sub-brand

OUTPUT FORMAT — strict JSON only, no markdown, no extra keys:
{{
  "phase1_evaluation": [
    {{"link": "...", "verdict": "ACCEPT|MAYBE|REJECT", "reason": "short reason"}}
  ],
  "best_link": "URL or empty string",
  "confidence": 0.0,
  "reason": "one sentence"
}}
"""

    body = {
        "model": OPENAI_CHAT_MODEL,
        "temperature": 0,
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user",   "content": user_msg},
        ],
    }
    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json",
    }
    response = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers=headers, json=body, timeout=60,
    )
    response.raise_for_status()
    content = response.json()["choices"][0]["message"]["content"]
    parsed  = _extract_json_obj(content)

    best_link  = str(parsed.get("best_link", "") or "").strip()
    confidence = float(parsed.get("confidence", 0.0))
    reason     = str(parsed.get("reason", "") or "").strip()
    confidence = max(0.0, min(1.0, confidence))

    for ev in (parsed.get("phase1_evaluation") or []):
        verdict = ev.get("verdict", "?")
        link_ev = (ev.get("link") or "")[:80]
        print(f"  [PHASE1] {verdict:6s} | {link_ev} — {ev.get('reason','')}")

    candidate_links = {normalize_profile_url(c.get("link", ""), platform).rstrip("/")
                       for c in candidates}
    if best_link:
        norm = normalize_profile_url(best_link, platform).rstrip("/")
        if norm not in candidate_links:
            print(f"[WARN] AI returned a link not in candidates — discarding: {best_link}")
            best_link  = ""
            confidence = min(confidence, 0.30)
            reason     = f"AI link not in candidate set (discarded). {reason}"

    # ── If AI chose the early-enrich candidate, boost confidence slightly ──
    if best_link and early_enrich_link:
        if normalize_profile_url(best_link, platform).rstrip("/") == normalize_profile_url(early_enrich_link, platform).rstrip("/"):
            confidence = min(1.0, confidence + 0.05)
            reason = f"{reason} [early-enrich confirmed]"

    return {"best_link": best_link, "confidence": confidence, "reason": reason or "No confident match."}


# ─────────────────────────────────────────────
#  AI — VERIFICATION PASS
# ─────────────────────────────────────────────

def ai_verify_selected_link(
    talent: str,
    platform: str,
    link: str,
    title: str,
    snippet: str,
    entity_category: str,
    entity_sub_category: str,
    search_keywords: str,
    wikipedia_context: str = "",
) -> Tuple[bool, float, str]:
    """
    Quick second AI call: 'Does this specific URL definitively belong to [talent]?'
    Returns (verified: bool, adjusted_confidence: float, reason: str).
    """
    exp_v       = parse_entity_expectations(entity_category, entity_sub_category, talent=talent)
    cat_context = get_category_disambiguation_context(entity_category, entity_sub_category, talent=talent)
    signals     = build_candidate_signals(talent, {"link": link, "title": title, "snippet": snippet}, platform)

    if exp_v.get("expects_brand"):
        system_msg = (
            "You are a fact-checker verifying whether a social URL is the official page for a "
            "media brand/publisher/network (organization). Answer with strict JSON only."
        )
        verify_q  = f"Does this URL clearly belong to the organization \"{talent}\" on {platform}?"
        verify_no = (
            "Answer NO if: clearly a different brand, an individual employee/fan account, "
            "a vertical sub-page (PR/CFB/etc.) when the talent is the parent brand, "
            "or a news article — not the org profile."
        )
    else:
        system_msg = (
            "You are a fact-checker verifying whether a specific social media URL belongs "
            "to a specific public figure. Answer with strict JSON only."
        )
        verify_q  = f"Does this URL clearly and definitively belong to the talent named above on {platform}?"
        verify_no = (
            "Answer NO if:\n"
            "  • This is clearly a different person\n"
            "  • This is a fan/tribute/unofficial page\n"
            "  • This is a news article or Wikipedia page\n"
            "  • There is not enough evidence to confirm identity"
        )
    user_msg = f"""
TALENT: "{talent}"
PLATFORM: {platform}
CATEGORY CONTEXT: {cat_context}
SEARCH KEYWORDS: {search_keywords or "(none)"}
WIKIPEDIA IDENTITY CONTEXT: {wikipedia_context or "(not provided)"}
KNOWN NAME ALIASES: {", ".join(_identity_name_aliases(talent, entity_category, entity_sub_category, wikipedia_context)) or "(none)"}

URL TO VERIFY: {link}
PAGE TITLE:    {title}
SNIPPET:       {snippet}
PRE-COMPUTED SIGNALS: {json.dumps(signals, ensure_ascii=True)}

QUESTION: {verify_q}

Answer YES only if you are confident this is the official profile for this entity.
{verify_no}

Output strict JSON only:
{{"verified": true/false, "confidence": 0.0, "reason": "one sentence"}}
"""
    body = {
        "model": OPENAI_CHAT_MODEL,
        "temperature": 0,
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user",   "content": user_msg},
        ],
    }
    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json",
    }
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers, json=body, timeout=40,
        )
        response.raise_for_status()
        parsed     = _extract_json_obj(response.json()["choices"][0]["message"]["content"])
        verified   = bool(parsed.get("verified", False))
        conf       = max(0.0, min(1.0, float(parsed.get("confidence", 0.0))))
        verify_rsn = str(parsed.get("reason", "")).strip()
        return verified, conf, verify_rsn
    except Exception as exc:
        print(f"[WARN] Verify call failed: {exc}")
        return True, 0.0, f"Verify skipped ({exc})"


# ─────────────────────────────────────────────
#  EMISSION GATE
# ─────────────────────────────────────────────

def decide_emitted_link(
    talent: str,
    platform: str,
    selected: str,
    confidence: float,
    reason: str,
    top_candidate: Optional[dict],
    search_keywords: str,
    title_category: str = "",
    title_sub_category: str = "",
    emit_candidate: Optional[dict] = None,
) -> Tuple[str, float, str]:
    effective_min = _effective_min_confidence(talent)
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if _platform_suppressed_for_talent(talent, platform):
        return "", 0.0, f"{platform} suppressed for this talent after manual validation."

    if not selected or selected == "Not Found":
        return "", confidence, reason or "No selection."

    selected = normalize_profile_url(selected, platform)
    if not is_valid_profile_url(selected, platform):
        return "", 0.0, "Rejected: not a valid profile/channel URL."

    if emit_candidate:
        rej, why = entity_profile_rejected(talent, title_category, title_sub_category, emit_candidate, platform)
        if rej:
            return "", min(confidence, 0.12), why

    if confidence >= effective_min:
        return selected, confidence, reason

    # Fallback: strong deterministic rank + URL name alignment
    if top_candidate is not None:
        rej_fb, rej_msg = entity_profile_rejected(talent, title_category, title_sub_category, top_candidate, platform)
        if rej_fb:
            return "", confidence, f"Omitted: {rej_msg}"
        rs   = candidate_rank_score(talent, top_candidate, search_keywords, title_category, title_sub_category)
        link = top_candidate.get("link", "")
        prof = profile_from_candidate_url(link, platform) or link
        url_ok = (
            rs >= MIN_RANK_SCORE_FOR_FALLBACK
            and talent_url_aligned(talent, prof, title_category, title_sub_category)
            and is_valid_profile_url(prof, platform)
        )
        if exp.get("expects_brand"):
            path_slug = _path_handle_slug(prof, platform)
            url_ok = url_ok or (
                rs >= MIN_RANK_SCORE_FOR_FALLBACK - 2.0
                and _brand_handle_matches_row(talent, path_slug, platform)
                and is_valid_profile_url(prof, platform)
            )
        if url_ok:
            fb_conf = min(0.85, max(confidence, rs / 20.0))
            if exp.get("expects_brand"):
                fb_conf = max(fb_conf, 0.80)
            return (
                normalize_profile_url(prof, platform),
                fb_conf,
                f"Strong search rank + URL match ({rs:.1f}): {reason}",
            )

    return "", confidence, f"Omitted (below {effective_min:.2f}): {reason}"


# ─────────────────────────────────────────────
#  SERPER SEARCH
# ─────────────────────────────────────────────

def serper_search(query: str, num_results: int = 10) -> List[dict]:
    url     = "https://google.serper.dev/search"
    headers = {"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"}
    payload = {"q": query, "num": max(1, min(num_results, 10))}
    response = requests.post(url, headers=headers, json=payload, timeout=30)
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        try:
            detail = response.json().get("message") or response.text
        except ValueError:
            detail = response.text
        raise RuntimeError(f"Serper search failed ({response.status_code}): {detail}") from exc
    data = response.json()
    return [
        {
            "title":   item.get("title", "") or "",
            "snippet": item.get("snippet", "") or "",
            "link":    item.get("link", "") or "",
        }
        for item in data.get("organic", [])
    ]


def _extract_json_obj(text: str) -> dict:
    if not text:
        raise ValueError("Empty OpenAI response.")
    start = text.find("{")
    end   = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("No JSON object found in OpenAI response.")
    return json.loads(text[start: end + 1])


# ─────────────────────────────────────────────
#  HTML ENRICHMENT
# ─────────────────────────────────────────────

URL_IN_TEXT_RE = re.compile(r"https?://[^\s\"\'<>\)\]]+", re.I)


def fetch_html(url: str) -> str:
    try:
        r = requests.get(url, headers=FETCH_HEADERS, timeout=20, allow_redirects=True)
        r.raise_for_status()
        if len(r.content) > 2_500_000:
            return ""
        return r.text or ""
    except Exception as exc:
        print(f"[WARN] fetch failed {url[:90]}… : {exc}")
        return ""


def extract_urls_from_html(html: str) -> List[str]:
    if not html:
        return []
    found: set = set()
    for m in URL_IN_TEXT_RE.finditer(html):
        u = m.group(0).rstrip(".,);\\]}\"'")
        if u.startswith("http"):
            found.add(u.split("&utm_")[0])
    for m in re.finditer(r'href\s*=\s*["\']([^"\']+)["\']', html, re.I):
        h = m.group(1).strip()
        if h.startswith("http"):
            found.add(h.split("&utm_")[0])
    return list(found)


def _platform_for_discovered_url(url: str) -> Optional[str]:
    for plat in PLATFORMS:
        if is_valid_profile_url(url, plat):
            return plat
    u = url.lower()
    if "linktr.ee/" in u or "linktree.com/" in u or "lnk.bio" in u or "beacons.ai" in u:
        return "__link_hub__"
    return None


def _extract_visible_social_handles(html: str) -> Dict[str, str]:
    if not html:
        return {}
    text = re.sub(r"\\u([0-9a-fA-F]{4})", lambda m: chr(int(m.group(1), 16)), html)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    out: Dict[str, str] = {}
    patterns = {
        "Instagram": r"(?:instagram|ig)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
        "TikTok":    r"(?:tiktok|tik tok)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
        "X":         r"(?:twitter|x\.com)[^A-Za-z0-9@._-]{0,40}@?([A-Za-z0-9._]{3,30})",
    }
    blocked = {"instagram", "tiktok", "twitter", "facebook", "youtube", "official"}
    for platform, pattern in patterns.items():
        for m in re.finditer(pattern, text, re.I):
            handle = m.group(1).strip("._-")
            if handle.lower() in blocked:
                continue
            out[platform] = handle
            break
    return out


def _url_for_platform_handle(platform: str, handle: str) -> str:
    handle = (handle or "").strip().lstrip("@")
    if not handle:
        return ""
    if platform == "Instagram":
        return f"https://www.instagram.com/{handle}"
    if platform == "TikTok":
        return f"https://www.tiktok.com/@{handle}"
    if platform == "X":
        return f"https://x.com/{handle}"
    if platform == "Facebook":
        return f"https://www.facebook.com/{handle}"
    return ""


def extract_social_links_from_page(page_url: str, source_platform: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    to_fetch  = [page_url]
    if source_platform == "YouTube":
        base = page_url.split("?")[0].rstrip("/")
        if "/@" in base or "/channel/" in base or "/c/" in base or "/user/" in base:
            if "/about" not in base:
                to_fetch.append(base + "/about")
    hubs_fetched = 0
    seen_fetch: set = set()
    for u in to_fetch:
        u = u.strip()
        if not u or u in seen_fetch:
            continue
        seen_fetch.add(u)
        html = fetch_html(u)
        for raw in extract_urls_from_html(html):
            raw  = raw.strip().rstrip(".,);")
            plat = _platform_for_discovered_url(raw)
            if plat and plat != "__link_hub__" and plat not in out:
                out[plat] = normalize_profile_url(raw, plat)
            elif plat == "__link_hub__" and hubs_fetched < 3:
                hubs_fetched += 1
                for raw2 in extract_urls_from_html(fetch_html(raw)):
                    raw2 = raw2.strip().rstrip(".,);")
                    p2   = _platform_for_discovered_url(raw2)
                    if p2 and p2 != "__link_hub__" and p2 not in out:
                        out[p2] = normalize_profile_url(raw2, p2)
        for plat, handle in _extract_visible_social_handles(html).items():
            if plat in out:
                continue
            url = _url_for_platform_handle(plat, handle)
            if url and is_valid_profile_url(url, plat):
                out[plat] = normalize_profile_url(url, plat)
    return out


# ─────────────────────────────────────────────
#  PHASE-1 EARLY ENRICHMENT  (NEW — runs before AI decisions)
# ─────────────────────────────────────────────

def run_early_enrichment(
    df: pd.DataFrame,
    row_label: object,
    resolved_links: Dict[str, str],
) -> Dict[str, str]:
    """
    Phase-1 enrichment: fires as soon as ANY platform is resolved with confidence
    >= EARLY_ENRICH_MIN_CONFIDENCE (0.80).

    Returns a dict {platform: url} of links discovered from the anchor's bio.
    These are stored in ROW_EARLY_ENRICH_LINKS and injected as extra candidates
    into the AI prompt for each still-unresolved platform.

    Critically this runs BEFORE AI makes decisions on other platforms, so AI gets
    richer candidate pools — fixing the architectural flaw in the original code.

    SAFETY: a discovered link is only stored, never directly written to the df here.
    Writing only happens in the conservative post-gate Phase-2 enrichment. This
    prevents a wrong anchor from directly cascading into the output.
    """
    anchor_order = ["Instagram", "YouTube", "X", "Facebook", "TikTok"]
    confs        = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
    talent       = str(df.at[row_label, "Talent Name"] or "")

    already_cached = ROW_EARLY_ENRICH_LINKS.get(row_label)
    if already_cached is not None:
        return already_cached

    best_plat, best_url, best_c = None, "", 0.0
    for p in anchor_order:
        url = str(df.at[row_label, p] or "").strip()
        if not url:
            continue
        c = float(confs.get(p, 0.0))
        if c < EARLY_ENRICH_MIN_CONFIDENCE:
            continue
        if c > best_c:
            best_plat, best_url, best_c = p, url, c

    if not best_url or not best_plat:
        ROW_EARLY_ENRICH_LINKS[row_label] = {}
        return {}

    print(f"[EARLY-ENRICH] {talent} <- Phase-1 anchor {best_plat} (conf={best_c:.2f})")
    try:
        discovered = extract_social_links_from_page(best_url, best_plat)
    except Exception as exc:
        print(f"[WARN] early enrich failed: {exc}")
        ROW_EARLY_ENRICH_LINKS[row_label] = {}
        return {}

    # Only keep platforms not yet resolved
    filtered: Dict[str, str] = {}
    for plat, link in discovered.items():
        if plat not in PLATFORMS:
            continue
        if _platform_suppressed_for_talent(talent, plat):
            continue
        if str(df.at[row_label, plat] or "").strip():
            continue
        if not is_valid_profile_url(link, plat):
            continue
        filtered[plat] = link
        print(f"  [EARLY-ENRICH] found {plat}: {link[:80]}")

    ROW_EARLY_ENRICH_LINKS[row_label] = filtered
    return filtered


# ─────────────────────────────────────────────
#  PHASE-2 POST-GATE ENRICHMENT  (conservative write-through)
# ─────────────────────────────────────────────

def enrich_row_from_anchor_profiles(df: pd.DataFrame, row_label: object) -> None:
    """
    Phase-2 enrichment: conservative last-resort fill for still-blank platforms.
    Requires anchor confidence >= ANCHOR_MIN_CONFIDENCE (0.86, stricter than Phase-1).
    FIXED: confidence inheritance capped so enriched links can NEVER themselves become
    Phase-2 anchors (prevents cascade amplification of wrong anchors).
    """
    anchor_order = ["Instagram", "YouTube", "X", "Facebook", "TikTok"]
    confs        = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
    best_plat, best_url, best_c = None, "", 0.0
    for p in anchor_order:
        url = str(df.at[row_label, p] or "").strip()
        if not url:
            continue
        # ── FIXED: only use search/input/handle_reconcile sources as Phase-2 anchors ──
        src = ROW_PLATFORM_SOURCE.get(row_label, {}).get(p, "")
        if src in ("bio_enrich", "x_instagram_handle_match"):
            continue   # enriched results must not re-anchor
        c = float(confs.get(p, 0.0))
        if c < ANCHOR_MIN_CONFIDENCE:
            continue
        if c > best_c:
            best_plat, best_url, best_c = p, url, c

    if not best_url or not best_plat:
        return

    talent = str(df.at[row_label, "Talent Name"] or "")
    print(f"[ENRICH-P2] {talent} <- anchor {best_plat} (conf={best_c:.2f})")
    try:
        discovered = extract_social_links_from_page(best_url, best_plat)
    except Exception as exc:
        print(f"[WARN] enrich-p2 failed: {exc}")
        return

    if not discovered and best_plat == "Facebook":
        fb_handle  = _handle_slug_from_profile(best_url, "Facebook")
        alias_slugs = {
            _slug_chars(h)
            for h in _identity_handle_aliases(
                talent,
                str(df.at[row_label, "title_category"] or ""),
                str(df.at[row_label, "title_sub_category"] or ""),
                wikipedia_identity_context(
                    df.at[row_label, WIKIPEDIA_URL_COLUMN]
                    if WIKIPEDIA_URL_COLUMN in df.columns else ""
                ),
            )
        }
        if fb_handle and fb_handle in alias_slugs:
            for tgt in ("Instagram", "TikTok"):
                guessed = _url_for_platform_handle(tgt, fb_handle)
                if guessed and is_valid_profile_url(guessed, tgt):
                    discovered.setdefault(tgt, guessed)

    for tgt, link in discovered.items():
        if tgt not in PLATFORMS:
            continue
        if _platform_suppressed_for_talent(talent, tgt):
            continue
        if str(df.at[row_label, tgt] or "").strip():
            continue
        if not is_valid_profile_url(link, tgt):
            continue
        # ── FIXED: hard cap so enriched confidence < ANCHOR_MIN_CONFIDENCE ──
        # Prevents Phase-2 enriched links from triggering another enrichment round.
        conf_value = round(min(ANCHOR_MIN_CONFIDENCE - 0.01, best_c * 0.94), 3)
        # ── NEW: must also meet ENRICH_EMIT_MIN_CONFIDENCE ──
        if conf_value < ENRICH_EMIT_MIN_CONFIDENCE:
            print(f"  [ENRICH-P2 SKIP] {tgt} conf {conf_value:.3f} below emit gate")
            continue
        df.at[row_label, tgt] = link
        ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[tgt] = conf_value
        df.at[row_label, PLATFORM_CONF_COLUMNS[tgt]]           = conf_value
        ROW_PLATFORM_SOURCE.setdefault(row_label, {})[tgt]     = "bio_enrich"
        print(f"  [ENRICH-P2] filled {tgt} from bio/link hub (conf={conf_value:.3f})")
    _refresh_row_aggregate_confidence(df, row_label)


def _refresh_row_aggregate_confidence(df: pd.DataFrame, row_label: object) -> None:
    parts = [
        float(ROW_PLATFORM_CONFIDENCE.get(row_label, {}).get(p, 0.0))
        for p in PLATFORMS
        if str(df.at[row_label, p] or "").strip()
    ]
    if parts:
        df.at[row_label, "Confidence"] = round(sum(parts) / len(parts), 4)


def _refresh_row_source_cell(df: pd.DataFrame, row_label: object) -> None:
    parts = []
    for p in PLATFORMS:
        if not str(df.at[row_label, p] or "").strip():
            continue
        src = ROW_PLATFORM_SOURCE.get(row_label, {}).get(p, "")
        if src:
            parts.append(f"{p}:{src}")
    df.at[row_label, "Source"] = "; ".join(parts)


def _handle_slug_from_profile(url: str, platform: str) -> str:
    if platform == "TikTok":
        m = re.search(r"tiktok\.com/@([^/?#]+)", url or "", re.I)
        return _slug_chars(m.group(1)) if m else ""
    return _path_handle_slug(url, platform)


def _find_exact_handle_profile(
    handle: str,
    talent: str,
    platform: str,
    domains: List[str],
    title_category: str,
    title_sub_category: str,
) -> Optional[dict]:
    if _platform_suppressed_for_talent(talent, platform):
        return None
    handle_clean = (handle or "").strip().lstrip("@")
    handle_slug  = _slug_chars(handle_clean)
    if len(handle_slug) < 4:
        return None

    queries: List[str] = []
    for domain in domains:
        queries.append(f"site:{domain}/{handle_clean}")
        queries.append(f"site:{domain}/@{handle_clean}")
        queries.append(f'site:{domain} "{handle_clean}"')
    queries.append(f'"{handle_clean}" {platform}')

    seen: set = set()
    for query in queries:
        try:
            results = serper_search(query, num_results=RESULTS_PER_QUERY)
        except Exception as exc:
            print(f"[WARN] handle reconcile failed '{query}': {exc}")
            continue
        for pos, item in enumerate(results, start=1):
            raw_link = item.get("link", "")
            prof = (
                profile_from_candidate_url(raw_link, platform)
                or profile_from_candidate_text(item, platform)
                or raw_link
            )
            if not prof or prof in seen or not is_valid_profile_url(prof, platform):
                continue
            seen.add(prof)
            if _handle_slug_from_profile(prof, platform) != handle_slug:
                continue
            cand = {**item, "link": prof, "search_position": pos, "query": query}
            rej, _ = entity_profile_rejected(talent, title_category, title_sub_category, cand, platform)
            if not rej:
                return cand
        time.sleep(0.2)
    return None


def _is_initial_style_person_name(talent: str) -> bool:
    parts = _brand_name_parts(talent)
    if len(parts) < 2:
        return False
    first = parts[0]
    return bool(len(first) <= 3 and first.upper() == first and re.search(r"[A-Z]", first))


def reconcile_initial_name_x_instagram_handle(
    df: pd.DataFrame,
    row_label: object,
    resolved_links: Dict[str, str],
    title_category: str,
    title_sub_category: str,
) -> None:
    talent = str(df.at[row_label, "Talent Name"] or "").strip()
    if not _is_initial_style_person_name(talent):
        return
    exp = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand"):
        return
    confs  = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
    x_url  = resolved_links.get("X", "")
    x_conf = float(confs.get("X", 0.0))
    if not x_url or x_conf < 0.85:
        return
    x_handle = _handle_slug_from_profile(x_url, "X")
    if len(x_handle) < 4:
        return

    current_ig        = str(df.at[row_label, "Instagram"] or "").strip()
    current_ig_handle = _handle_slug_from_profile(current_ig, "Instagram") if current_ig else ""
    if current_ig_handle == x_handle:
        return

    cand = _find_exact_handle_profile(
        x_handle, talent, "Instagram", PLATFORMS["Instagram"],
        title_category, title_sub_category,
    )
    if not cand:
        return
    score = candidate_rank_score(
        talent, cand, extract_search_keywords(title_category, title_sub_category),
        title_category, title_sub_category, platform="Instagram",
    )
    if score < 5.0:
        return

    link       = normalize_profile_url(cand["link"], "Instagram")
    conf_value = round(min(0.92, max(0.86, x_conf * 0.98)), 3)
    print(f"  [INITIAL-HANDLE-FIX] Instagram | {talent} | using X @{x_handle} -> {link[:90]}")
    df.at[row_label, "Instagram"] = link
    ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})["Instagram"] = conf_value
    df.at[row_label, PLATFORM_CONF_COLUMNS["Instagram"]] = conf_value
    ROW_PLATFORM_SOURCE.setdefault(row_label, {})["Instagram"] = "x_instagram_handle_match"
    resolved_links["Instagram"] = link


def reconcile_athlete_handles_from_confirmed_profiles(
    df: pd.DataFrame,
    row_label: object,
    resolved_links: Dict[str, str],
    title_category: str,
    title_sub_category: str,
) -> None:
    talent = str(df.at[row_label, "Talent Name"] or "").strip()
    exp    = parse_entity_expectations(title_category, title_sub_category, talent=talent)
    if exp.get("expects_brand") or not _athlete_row_active(exp):
        return

    confs   = ROW_PLATFORM_CONFIDENCE.get(row_label, {})
    anchors = [
        (p, resolved_links.get(p, ""), float(confs.get(p, 0.0)))
        for p in ("X", "Instagram", "TikTok")
        if resolved_links.get(p) and float(confs.get(p, 0.0)) >= 0.85
    ]
    for src_plat, src_url, src_conf in sorted(anchors, key=lambda x: -x[2]):
        handle = _handle_slug_from_profile(src_url, src_plat)
        if len(handle) < 4:
            continue
        for target_plat, domains in PLATFORMS.items():
            if target_plat == src_plat or target_plat not in {"Instagram", "X", "TikTok", "YouTube"}:
                continue
            current      = str(df.at[row_label, target_plat] or "").strip()
            current_conf = float(confs.get(target_plat, 0.0))
            if current and _handle_slug_from_profile(current, target_plat) == handle:
                continue
            if current and current_conf >= 0.55:
                continue
            cand = _find_exact_handle_profile(
                handle, talent, target_plat, domains, title_category, title_sub_category
            )
            if not cand:
                continue
            link       = normalize_profile_url(cand["link"], target_plat)
            conf_value = round(min(0.92, src_conf * 0.98), 3)
            print(f"  [HANDLE-RECONCILE] {target_plat} | {talent} | @{handle} from {src_plat} -> {link[:90]}")
            df.at[row_label, target_plat] = link
            ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[target_plat] = conf_value
            df.at[row_label, PLATFORM_CONF_COLUMNS[target_plat]] = conf_value
            ROW_PLATFORM_SOURCE.setdefault(row_label, {})[target_plat] = "handle_reconcile"
            resolved_links[target_plat] = link


# ─────────────────────────────────────────────
#  SEARCH ONE PLATFORM  (with Phase-1 enrich injection)
# ─────────────────────────────────────────────

def search_one_platform(
    talent: str,
    platform: str,
    domains: List[str],
    title_category: str,
    title_sub_category: str,
    username_hints: Optional[Dict[str, str]] = None,
    wikipedia_context: str = "",
    # NEW: early-enrich cache for this row
    early_enrich_cache: Optional[Dict[str, str]] = None,
) -> Tuple[str, str, float, str]:
    search_keywords = extract_search_keywords(title_category, title_sub_category)
    lookup_talent   = _talent_lookup_name(talent, title_category, title_sub_category)
    if _platform_suppressed_for_talent(lookup_talent, platform):
        return platform, "", 0.0, f"{platform} suppressed for this talent after manual validation."

    exp_search     = parse_entity_expectations(title_category, title_sub_category, talent=lookup_talent)
    all_candidates: List[dict] = []
    seen_links: set = set()

    # ── NEW: early-enrich link for this platform, if any ──
    early_enrich_link = (early_enrich_cache or {}).get(platform, "")

    # ── If early-enrich already gave us a confident link, use fast query path ──
    queries = build_queries(
        lookup_talent, platform, domains, search_keywords,
        title_category, title_sub_category,
        username_hints=username_hints,
        wikipedia_context=wikipedia_context,
        fast=bool(early_enrich_link),   # fast=True skips exhaustive query list
    )

    for query in queries:
        try:
            results = serper_search(query, num_results=RESULTS_PER_QUERY)
            print(f"[QUERY] {platform} | {lookup_talent} | '{query}' -> {len(results)} raw results")
            for pos, item in enumerate(results, start=1):
                raw_link = item.get("link", "")
                prof     = profile_from_candidate_url(raw_link, platform)
                recovered_from_text = False
                if not prof:
                    prof = profile_from_candidate_text(item, platform)
                    recovered_from_text = bool(prof)
                if not prof:
                    prof = raw_link
                if not prof or prof in seen_links:
                    continue
                if not is_valid_profile_url(prof, platform):
                    continue
                seen_links.add(prof)
                all_candidates.append({
                    **item,
                    "link":                prof,
                    "search_position":     pos,
                    "query":               query,
                    "recovered_from_text": recovered_from_text,
                })
        except Exception as exc:
            print(f"[WARN] Serper failed '{query}': {exc}")
            fatal_markers = (
                "not enough credits", "unauthorized", "invalid api key",
                "forbidden", "quota", "billing",
            )
            if any(marker in str(exc).lower() for marker in fatal_markers):
                raise RuntimeError(str(exc)) from exc
        if len(all_candidates) >= RESULTS_PER_QUERY * 2:
            break
        time.sleep(0.2)

    valid_candidates = [c for c in all_candidates if is_valid_profile_url(c.get("link", ""), platform)]
    if exp_search.get("expects_brand"):
        canonical   = _brand_canonical_slug(lookup_talent)
        brand_clean: List[dict] = []
        for c in valid_candidates:
            clink     = c.get("link", "")
            path_slug = _path_handle_slug(clink, platform)
            if _brand_slug_is_vertical(canonical, path_slug, lookup_talent, platform):
                print(f"  [BRAND] Skip vertical handle: {(clink or '')[:90]}")
                continue
            rej_brand, why_brand = entity_profile_rejected(lookup_talent, title_category, title_sub_category, c, platform)
            if rej_brand:
                print(f"  [BRAND] Skip mismatch: {(clink or '')[:90]} | {why_brand}")
                continue
            brand_clean.append(c)
        valid_candidates = brand_clean

    valid_candidates = sort_candidates_for_ai(
        lookup_talent, valid_candidates, search_keywords, title_category, title_sub_category,
        username_hints=username_hints, platform=platform,
    )
    top_candidates = valid_candidates[:MAX_CANDIDATES_FOR_AI]

    ctx = f" | kw: {search_keywords}" if search_keywords else ""
    print(f"[INFO] {platform} | {lookup_talent}{ctx} -> {len(top_candidates)} candidates for AI"
          + (f" | early-enrich: {early_enrich_link[:60]}" if early_enrich_link else ""))

    if not top_candidates and not early_enrich_link:
        return platform, "", 0.0, "No valid profile/channel URLs in search results."

    top_candidate = top_candidates[0] if top_candidates else None
    fallback      = first_valid_profile_link(top_candidates, platform)

    try:
        ai_result  = ai_select_best_profile(
            lookup_talent, platform, top_candidates,
            title_category, title_sub_category, search_keywords,
            username_hints=username_hints,
            wikipedia_context=wikipedia_context,
            early_enrich_link=early_enrich_link,   # NEW: inject early-enrich link
        )
        selected   = ai_result["best_link"]
        confidence = ai_result["confidence"]
        reason     = ai_result["reason"]

        if top_candidates:
            recovered_pick = _prefer_recovered_athlete_handle(
                lookup_talent, platform, selected, top_candidates,
                title_category, title_sub_category, search_keywords,
            )
            if recovered_pick:
                selected   = recovered_pick["link"]
                confidence = max(confidence, 0.88)
                reason     = f"{reason} | Preferred @handle recovered from search result text."

            personal_site_pick = _prefer_personal_website_athlete_profile(
                lookup_talent, platform, selected, top_candidates,
                title_category, title_sub_category, search_keywords,
            )
            if personal_site_pick:
                selected   = personal_site_pick["link"]
                confidence = max(confidence, 0.90)
                reason     = f"{reason} | Preferred top-ranked athlete profile with name-matching personal website."

            sport_fix = _prefer_non_mismatch_athlete_profile(
                lookup_talent, platform, selected, top_candidates,
                title_category, title_sub_category, search_keywords,
            )
            if sport_fix:
                if sport_fix.get("_force_blank"):
                    selected   = ""
                    confidence = 0.0
                    reason     = f"{reason} | Rejected wrong-sport same-name profile (e.g. golfer vs basketball)."
                else:
                    selected   = sport_fix["link"]
                    confidence = max(confidence, 0.82)
                    reason     = f"{reason} | Switched to non-conflicting sport/handle match."

        if not selected and fallback:
            rej_fb, _ = entity_profile_rejected(
                lookup_talent, title_category, title_sub_category,
                top_candidate or {}, platform,
            )
            if not rej_fb:
                selected   = fallback
                confidence = min(confidence, 0.42)
                reason     = (reason or "") + " | AI empty; using top-ranked candidate."

        elif selected and not is_valid_profile_url(selected, platform):
            print(f"[WARN] AI returned non-profile URL; trying fallback.")
            rej_fb, _ = entity_profile_rejected(
                lookup_talent, title_category, title_sub_category,
                top_candidate or {}, platform,
            )
            if fallback and not rej_fb:
                selected   = fallback
                confidence = min(confidence, 0.42)
                reason     = f"{reason} (AI URL invalid; fallback used)"
            else:
                selected   = ""
                confidence = 0.0

        # ── Hard entity check on the selected candidate ──
        emit_candidate: Optional[dict] = None
        if selected:
            sel_norm = normalize_profile_url(selected, platform).rstrip("/")
            cand     = next(
                (c for c in top_candidates
                 if normalize_profile_url(c.get("link", ""), platform).rstrip("/") == sel_norm),
                None,
            )
            # ── NEW: also check early-enrich injected candidate ──
            if cand is None and early_enrich_link:
                norm_enrich = normalize_profile_url(early_enrich_link, platform).rstrip("/")
                if sel_norm == norm_enrich:
                    cand = {
                        "link":    early_enrich_link,
                        "title":   f"Bio-enriched: {lookup_talent}",
                        "snippet": "Discovered from anchor bio.",
                        "from_early_enrich": True,
                    }
            emit_candidate = cand
            if cand:
                rej, why = entity_profile_rejected(
                    lookup_talent, title_category, title_sub_category, cand, platform
                )
                if rej:
                    # ── Early-enrich candidates get a grace: only veto if NOT from a confirmed input ──
                    if cand.get("from_early_enrich"):
                        print(f"[WARN] Early-enrich candidate for {platform} rejected by entity check: {why}")
                    print(f"[REJECT] {platform} | {lookup_talent} | {why}")
                    selected = ""; confidence = min(confidence, 0.15); reason = why
                    emit_candidate = None

        # ── Verification pass ──
        if selected and confidence >= 0.55:
            cand_for_verify = emit_candidate or top_candidate
            verified, verify_conf, verify_rsn = ai_verify_selected_link(
                lookup_talent, platform, selected,
                cand_for_verify.get("title", "") if cand_for_verify else "",
                cand_for_verify.get("snippet", "") if cand_for_verify else "",
                title_category, title_sub_category, search_keywords,
                wikipedia_context=wikipedia_context,
            )
            print(f"[VERIFY] {platform} | {lookup_talent} | verified={verified} conf={verify_conf:.2f} | {verify_rsn}")
            path_slug   = _path_handle_slug(selected, platform)
            cand_brand  = emit_candidate or top_candidate
            brand_handle_ok = (
                exp_search.get("expects_brand")
                and (
                    _brand_handle_matches_row(lookup_talent, path_slug, platform)
                    or (
                        platform == "YouTube"
                        and _is_youtube_channel_id_slug(path_slug)
                        and cand_brand
                        and _candidate_supports_brand(lookup_talent, cand_brand)
                    )
                )
            )
            if not verified and verify_conf < AI_VERIFY_MIN_CONFIDENCE:
                athlete_profile_ok = (
                    not exp_search.get("expects_brand")
                    and _athlete_row_active(exp_search)
                    and cand_for_verify
                    and _candidate_has_exact_profile_title(lookup_talent, cand_for_verify)
                    and confidence >= 0.75
                )
                # ── NEW: early-enrich waiver — bio-origin links get extra trust ──
                early_enrich_ok = (
                    bool(emit_candidate and emit_candidate.get("from_early_enrich"))
                    and confidence >= 0.70
                )
                if brand_handle_ok and confidence >= 0.55:
                    print(f"[VERIFY-WAIVER] {platform} | brand handle match")
                    reason = f"{reason} [verify waived: {verify_rsn}]"
                elif athlete_profile_ok:
                    print(f"[VERIFY-WAIVER] {platform} | exact athlete profile title")
                    reason = f"{reason} [verify waived: {verify_rsn}]"
                elif early_enrich_ok:
                    print(f"[VERIFY-WAIVER] {platform} | early-enrich bio origin")
                    reason = f"{reason} [verify waived: bio-enrich origin — {verify_rsn}]"
                else:
                    print(f"[VETO] Verify vetoed result for {platform} | {lookup_talent}")
                    selected   = ""
                    confidence = min(confidence, verify_conf)
                    reason     = f"Verify pass failed: {verify_rsn}"
                    emit_candidate = None
            elif verified and verify_conf > 0:
                confidence = round((confidence + verify_conf) / 2, 4)
                reason     = f"{reason} [verify: {verify_rsn}]"

        # ── Final emission gate ──
        emit, conf_out, rsn_out = decide_emitted_link(
            lookup_talent, platform, selected or "", confidence, reason,
            top_candidate, search_keywords, title_category, title_sub_category,
            emit_candidate=emit_candidate,
        )
        return platform, emit, conf_out, rsn_out

    except Exception as exc:
        print(f"[ERROR] AI/Verify failed for {platform} | {lookup_talent}: {exc}")
        if fallback and top_candidate:
            rej_fb, _ = entity_profile_rejected(
                lookup_talent, title_category, title_sub_category, top_candidate, platform
            )
            if not rej_fb:
                rs = candidate_rank_score(
                    lookup_talent, top_candidate, search_keywords, title_category, title_sub_category
                )
                if rs >= MIN_RANK_SCORE_FOR_FALLBACK and talent_url_aligned(
                    lookup_talent, fallback, title_category, title_sub_category,
                ):
                    return (
                        platform,
                        normalize_profile_url(fallback, platform),
                        min(0.60, rs / 22.0),
                        f"AI error fallback (rank={rs:.1f}): {exc}",
                    )
        return platform, "", 0.0, f"AI error: {exc}"


# ─────────────────────────────────────────────
#  PROCESS ONE ROW  (restructured pipeline)
# ─────────────────────────────────────────────

def process_row(
    df: pd.DataFrame,
    row_label: object,
    platform_progress: Optional[Callable[[str, str], None]] = None,
) -> None:
    """Resolve all platforms for one talent via the Wikipedia + Serper + LLM workflow.

    Discovery logic lives in profile_discovery.py. This function only:
      • reads the row inputs (Talent Name + Wikipedia URL),
      • honours any client-provided platform cells (treated as INPUT),
      • writes results back into the existing wide-Excel columns, and
      • refreshes the aggregate Confidence + Source cells.

    Per-platform source ∈ {WIKIPEDIA, LLM_VERIFIED, INPUT}; blank when NOT_FOUND.
    """
    talent             = str(df.at[row_label, "Talent Name"] or "").strip()
    title_category     = str(df.at[row_label, "title_category"]    or "").strip()
    title_sub_category = str(df.at[row_label, "title_sub_category"] or "").strip()
    wikipedia_url      = _clean_wikipedia_url(
        df.at[row_label, WIKIPEDIA_URL_COLUMN] if WIKIPEDIA_URL_COLUMN in df.columns else ""
    )

    if not talent:
        return

    print(f"\n{'='*65}")
    print(f"Processing: {talent}")
    if title_category or title_sub_category:
        print(f"  Category: {title_category} | SubCategory: {title_sub_category}")
    if wikipedia_url:
        print(f"  Wikipedia URL: {wikipedia_url}")

    # ═══════════════════════════════════════════════════════════
    # STEP 0 — Honour client-provided platform cells as INPUT
    # (highest priority — confidence 1.0; skips discovery for that platform).
    # Includes the dedicated "Instagram URL" input column for backward compat.
    # ═══════════════════════════════════════════════════════════
    prefilled: Dict[str, str] = {}

    if INSTAGRAM_INPUT_COLUMN in df.columns:
        ig_input = _clean_instagram_input_url(df.at[row_label, INSTAGRAM_INPUT_COLUMN])
        if ig_input and not str(df.at[row_label, "Instagram"] or "").strip():
            prefilled["Instagram"] = ig_input
            print(f"  [INPUT] Instagram <- {ig_input} (conf=1.00, source=input)")

    for platform in PLATFORMS:
        existing = str(df.at[row_label, platform] or "").strip()
        if existing and platform not in prefilled:
            prefilled[platform] = existing

    for platform, url in prefilled.items():
        df.at[row_label, platform] = url
        ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[platform] = 1.0
        ROW_PLATFORM_SOURCE.setdefault(row_label, {})[platform]     = "INPUT"
        df.at[row_label, PLATFORM_CONF_COLUMNS[platform]] = 1.0

    # ═══════════════════════════════════════════════════════════
    # STEP 1 + 2 — Wikipedia metadata + existing Wikipedia/Wikidata socials
    # ═══════════════════════════════════════════════════════════
    metadata = profile_discovery.extract_wikipedia_metadata(wikipedia_url, talent)
    wiki_socials = profile_discovery.get_wikipedia_socials(
        talent, wikipedia_url, title_category, title_sub_category,
        target_platforms=list(PLATFORMS.keys()),
    )

    # ═══════════════════════════════════════════════════════════
    # STEP 3-6 — Per platform: Serper search -> candidates -> LLM verify -> gate
    # ═══════════════════════════════════════════════════════════
    for platform in PLATFORMS:
        if platform_progress:
            platform_progress(platform, "start")

        if platform in prefilled:                 # already provided as INPUT
            if platform_progress:
                platform_progress(platform, "done")
            continue

        try:
            result = profile_discovery.discover_platform(
                talent, platform, metadata,
                wiki_url=wiki_socials.get(platform),
                threshold=PROFILE_MATCH_THRESHOLD,
            )
        except RuntimeError:
            raise                                 # fatal API error — fail the job
        except Exception as exc:                  # noqa: BLE001 — defensive
            print(f"  [{platform}] UNEXPECTED ERROR: {exc}")
            result = {
                "profile_url": "", "source": profile_discovery.SOURCE_NOT_FOUND,
                "confidence": 0.0, "reasoning": str(exc),
            }

        link       = result.get("profile_url", "") or ""
        confidence = float(result.get("confidence", 0.0) or 0.0)
        source     = result.get("source", profile_discovery.SOURCE_NOT_FOUND)
        reason     = result.get("reasoning", "")
        candidates = result.get("candidate_urls", []) or []

        df.at[row_label, platform] = link
        ROW_PLATFORM_CONFIDENCE.setdefault(row_label, {})[platform] = confidence
        df.at[row_label, PLATFORM_CONF_COLUMNS[platform]] = confidence if link else float("nan")
        if link:
            ROW_PLATFORM_SOURCE.setdefault(row_label, {})[platform] = source

        print(f"  [{platform}] {link or '(blank)'} (conf={confidence:.2f}, {source}) — {reason}")
        if candidates:
            print(f"           candidates ({len(candidates)}): {candidates}")
        if platform_progress:
            platform_progress(platform, "done")
        time.sleep(OPENAI_DELAY_SECONDS)

    # ═══════════════════════════════════════════════════════════
    # STEP 7 — Aggregate row Confidence + Source cells
    # ═══════════════════════════════════════════════════════════
    _refresh_row_aggregate_confidence(df, row_label)
    _refresh_row_source_cell(df, row_label)


# ─────────────────────────────────────────────
#  API PIPELINE WRAPPERS
# ─────────────────────────────────────────────

def _ensure_pipeline_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep uploaded/name-only dataframes compatible with the resolver pipeline."""
    if WIKIPEDIA_URL_COLUMN not in df.columns:
        wiki_col = _find_column(
            df, "wikipedia_url", "Wikipedia", "Wiki URL", "wiki_url", "Wiki", "Wikipedia Link",
        )
        df[WIKIPEDIA_URL_COLUMN] = df[wiki_col] if wiki_col else ""
    if INSTAGRAM_INPUT_COLUMN not in df.columns:
        ig_col = _find_column(
            df, "instagram_url", "Instagram Input", "ig_url", "IG URL",
            "Instagram Link", "Instagram",
        )
        df[INSTAGRAM_INPUT_COLUMN] = df[ig_col] if ig_col else ""
    for column in ("Talent Name", "title_category", "title_sub_category"):
        if column not in df.columns:
            df[column] = ""
    df[WIKIPEDIA_URL_COLUMN]   = df[WIKIPEDIA_URL_COLUMN].apply(_clean_wikipedia_url)
    df[INSTAGRAM_INPUT_COLUMN] = df[INSTAGRAM_INPUT_COLUMN].apply(_clean_instagram_input_url)
    for platform in PLATFORMS:
        if platform not in df.columns:
            df[platform] = ""
        conf_column = PLATFORM_CONF_COLUMNS[platform]
        if conf_column not in df.columns:
            df[conf_column] = float("nan")
    if "Confidence" not in df.columns:
        df["Confidence"] = float("nan")
    if "Source" not in df.columns:
        df["Source"] = ""
    return df


def run_pipeline_on_dataframe(
    df: pd.DataFrame,
    progress: Optional[Callable[[int, int, str], None]] = None,
    platform_progress: Optional[Callable[[int, str, str], None]] = None,
) -> pd.DataFrame:
    df = _ensure_pipeline_columns(df.copy())
    ROW_PLATFORM_CONFIDENCE.clear()
    ROW_PLATFORM_SOURCE.clear()
    ROW_USERNAME_HINTS.clear()
    ROW_EARLY_ENRICH_LINKS.clear()   # NEW: clear early-enrich cache

    total = len(df)
    print(f"Initialized talent dataframe with {total} row(s).")

    for i, row_label in enumerate(df.index, start=1):
        talent = str(df.at[row_label, "Talent Name"] or "").strip()
        if not talent:
            continue

        ROW_PLATFORM_CONFIDENCE[row_label] = {}
        ROW_PLATFORM_SOURCE[row_label]     = {}

        # Mark any pre-filled platform cells as input-provided
        for platform in PLATFORMS:
            if str(df.at[row_label, platform] or "").strip():
                ROW_PLATFORM_CONFIDENCE[row_label][platform] = 1.0
                ROW_PLATFORM_SOURCE[row_label][platform]     = "input"
                df.at[row_label, PLATFORM_CONF_COLUMNS[platform]] = 1.0

        if progress:
            progress(i, total, talent)

        def _row_platform_progress(platform: str, phase: str) -> None:
            if platform_progress:
                platform_progress(i - 1, platform, phase)

        process_row(df, row_label, platform_progress=_row_platform_progress)
        delay = random.uniform(*REQUEST_DELAY_BETWEEN_TALENTS)
        print(f"  [{i}/{total}] complete — sleeping {delay:.1f}s")
        time.sleep(delay)

    return df


def run_pipeline_for_names(
    names: List[str],
    progress: Optional[Callable[[int, int, str], None]] = None,
    platform_progress: Optional[Callable[[int, str, str], None]] = None,
) -> pd.DataFrame:
    clean = [str(name).strip() for name in names if name and str(name).strip()]
    if not clean:
        raise ValueError("At least one non-empty name is required.")
    df = build_talent_df(clean, list(PLATFORMS.keys()))
    return run_pipeline_on_dataframe(df, progress=progress, platform_progress=platform_progress)


def run_pipeline() -> pd.DataFrame:
    return run_pipeline_on_dataframe(load_talent_table())


# ─────────────────────────────────────────────
#  EXCEL OUTPUT WITH FORMATTING
# ─────────────────────────────────────────────

def save_results(df: pd.DataFrame, output_dir: Optional[Path] = None) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir  = Path(output_dir) if output_dir is not None else Path(__file__).resolve().parent
    base_dir.mkdir(parents=True, exist_ok=True)
    output_path = base_dir / f"Talent_Social_Lookup_{timestamp}.xlsx"

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        df.to_excel(output_path, index=False)
        print(f"\n✅ Saved (no formatting): {output_path}")
        return output_path

    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    conf_col_indices = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value or "")
        for p in PLATFORMS:
            if val == PLATFORM_CONF_COLUMNS[p]:
                conf_col_indices[p] = col_idx

    low_conf_fill = PatternFill("solid", fgColor="FFF2CC")
    warn_fill     = PatternFill("solid", fgColor="FCE4D6")
    ok_fill       = PatternFill("solid", fgColor="E2EFDA")

    col_headers = [str(cell.value or "") for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        talent_val = str(row[0].value or "")
        first_only = is_first_name_only(talent_val)

        try:
            conf_idx    = col_headers.index("Confidence")
            overall_val = row[conf_idx].value
            overall_conf = float(overall_val) if overall_val not in (None, "") else None
        except (ValueError, TypeError):
            overall_conf = None

        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if first_only:
                cell.fill = warn_fill
            elif overall_conf is not None:
                if overall_conf >= 0.85:
                    cell.fill = ok_fill
                elif overall_conf < MIN_CONFIDENCE_EMIT:
                    cell.fill = low_conf_fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_cell in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row_cell:
                try:
                    max_len = max(max_len, len(str(c.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    return output_path


def save_output(df: pd.DataFrame, output_dir: Optional[Path] = None) -> str:
    """API-compatible output helper used by api_server.py."""
    return str(save_results(df, output_dir=output_dir))


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main() -> None:
    if not SERPER_API_KEY:
        print("[ERROR] SERPER_API_KEY not set. Add it to .env")
        return
    if not OPENAI_API_KEY:
        print("[ERROR] OPENAI_API_KEY not set. Add it to .env")
        return

    df    = load_talent_table()
    total = len(df)
    print(f"Loaded {total} talent row(s).")

    for idx, row_label in enumerate(df.index):
        talent = str(df.at[row_label, "Talent Name"] or "").strip()
        if not talent:
            continue
        process_row(df, row_label)
        delay = random.uniform(*REQUEST_DELAY_BETWEEN_TALENTS)
        print(f"  [{idx+1}/{total}] complete — sleeping {delay:.1f}s")
        time.sleep(delay)

    save_results(df)


if __name__ == "__main__":
    main()